import re
from openpyxl import load_workbook
from collections import Counter

BASE = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求'
XLSX = BASE + r'\天融信vsPA功能对比.xlsx'
HTML = BASE + r'\fw-comparison-report\fw-comparison-report.html'

wb = load_workbook(XLSX)
mx = wb['功能对比矩阵']
rq = wb['需求对比表']
wv = wb['待验证移交清单']
html = open(HTML, encoding='utf-8').read()

R = []  # 评审结果 (红线/检查项, 结论, 详情)
def check(name, ok, detail=''):
    R.append((name, 'PASS' if ok else 'FAIL', detail))

# ============ 实际数据 ============
rows = []
for r in range(4, mx.max_row + 1):
    no = mx.cell(r, 1).value
    if no is not None and str(no).strip().isdigit():
        rows.append([str(mx.cell(r, c).value or '') for c in range(1, 13)])
ts = Counter(x[4] for x in rows)
pa = Counter(x[6] for x in rows)
diff = Counter(x[8] for x in rows)

reqs = []
for r in range(4, 62):
    if rq.cell(r, 1).value:
        reqs.append([str(rq.cell(r, c).value or '') for c in range(1, 15)])
rconcl = Counter()
for x in reqs:
    k = x[11].split('：')[0]
    if 'TS占优' in k: k = '双方支持·TS占优'
    rconcl[k] += 1

vitems = []
for r in range(4, wv.max_row + 1):
    if wv.cell(r, 2).value:
        vitems.append([str(wv.cell(r, c).value or '') for c in range(1, 9)])
vmethod = Counter(x[4] for x in vitems)
vprio = Counter(x[5] for x in vitems)

# ============ 红线1 范围纯度 ============
spec_words = ['吞吐', '并发连接', '延迟', 'Gbps', 'pps', 'IOPS']
# 豁免排除性声明上下文（"不含性能/吞吐…"）
real_hits = []
for w in spec_words:
    for m in re.finditer(w, html):
        ctx = html[max(0, m.start() - 30):m.start() + 15]
        if '不含' not in ctx and '不引入' not in ctx:
            real_hits.append((w, ctx.strip()))
check('红线1 范围纯度（报告无性能结论）', not real_hits, f'非排除性声明命中: {real_hits[:3] or "无"}')
# 矩阵说明列含规格词的行：须为功能参数描述或待查标注，不得给出性能对比结论
mx_spec = [x[0] for x in rows if any(w in x[5] + x[7] for w in spec_words)]
spec_concl = [x[0] for x in rows if any(w in x[8] for w in spec_words)]  # 差异列出现规格词=性能结论
check('红线1b 矩阵无性能对比结论', not spec_concl, f'差异定性列含规格词行: {spec_concl or "无"}（说明列{len(mx_spec)}行为功能参数描述/待查标注）')

# ============ 红线2 功能全覆盖 ============
check('红线2 功能全覆盖（139点/双侧枚举）', len(rows) == 139, f'矩阵点数={len(rows)}')
single_side = [x for x in rows if x[4] == '不支持' or x[6] == '不支持']
check('红线2b 单侧独有/不支持项均独立条目', all(x[3] for x in single_side), f'双侧不支持项 {len(single_side)} 行功能点列均非空')

# ============ 红线3 同维度同标尺 ============
STD = {'完全支持', '部分支持', '需订阅授权', '不支持', '待验证'}
bad_ts = [x[0] for x in rows if x[4] not in STD]
bad_pa = [x[0] for x in rows if x[6] not in STD]
check('红线3 同标尺（五级枚举）', not bad_ts and not bad_pa, f'非法TS值:{bad_ts[:5]} 非法PA值:{bad_pa[:5]}')

# ============ 红线4 证据可溯源 ============
def ev_ok(e):
    src = 'CHM' in e or 'webhelp' in e or 'admin' in e or 'Web帮助' in e or 'Datasheet' in e or '官网' in e or 'Release' in e or '检索' in e
    anchor = '[' in e or re.search(r'p\d', e) or '页' in e or 'topic' in e or 'URL' in e or '留档' in e or '留痕' in e or '无命中' in e
    return src and anchor
no_ev = [x[0] for x in rows if not ev_ok(x[9])]
check('红线4 证据可溯源（文档名+章节/页码或否定证据留痕）', not no_ev, f'证据格式存疑行: {no_ev[:8]}')

# ============ 红线5 数据带时间 ============
no_date = [x[0] for x in rows if x[10] != '2026-08-15']
check('红线5 数据带时间', not no_date, f'日期异常行: {no_date[:8] or "无（全部2026-08-15，当日数据无需复核提示）"}')

# ============ 红线6 缺失不编造 ============
tbd_rows = [x for x in rows if x[4] == '待验证']
check('红线6a 待验证点有登记', all(('待验证' in x[11] or 'V' in x[11] or '查' in x[11]) for x in tbd_rows), f'待验证{len(tbd_rows)}点备注/V编号覆盖')
no_rows = [x for x in rows if x[4] == '不支持']
neg_ev = [x[0] for x in no_rows if '无命中' not in x[5] and '无相关' not in x[5] and '未见' not in x[5] and '无提及' not in x[5]]
check('红线6b 不支持点有否定证据', not neg_ev, f'否定证据缺失行: {neg_ev}')

# ============ 红线7 实测优先（待验证不升级） ============
vstat = Counter(x[6] for x in vitems)
check('红线7a V清单状态未升级', set(vstat) <= {'待验证', ''}, f'V状态分布: {dict(vstat)}')
# 报告中无"待验证...已支持"升级句式
upgrade = re.findall(r'待验证[^。]{0,20}(已是|已经支持|确认支持)', html)
check('红线7b 报告无升级表述', not upgrade, f'升级句式: {upgrade or "无"}')

# ============ 红线8 客观中立 ============
hype = [w for w in ['业界领先', '遥遥领先', '碾压', '完爆', '世界第一', '顶级'] if w in html]
check('红线8 客观中立', not hype, f'营销词命中: {hype or "无"}')

# ============ 红线9 结论有锚点 ============
refs = {m for m in re.findall(r'#(\d{1,3})\b', html) if len(m) <= 3}  # 排除hex色值
valid = {x[0] for x in rows}
bad_refs = [r for r in refs if r not in valid]
rrefs = set(re.findall(r'\bR(\d+)\b', html))
bad_r = [r for r in rrefs if not (1 <= int(r) <= 58)]
check('红线9 结论锚点有效', not bad_refs and not bad_r, f'无效#引用:{bad_refs[:5]} 无效R引用:{bad_r[:5]}')

# ============ 报告数字 vs 实际数据交叉 ============
nums = {
    '139': len(rows) == 139,
    '94/19/15/11': (ts['完全支持'], ts['部分支持'], ts['待验证'], ts['不支持']) == (94, 19, 15, 11),
    'PA 101/10/12/10/6': (pa['完全支持'], pa['需订阅授权'], pa['部分支持'], pa['不支持'], pa['待验证']) == (101, 10, 12, 10, 6),
    '差异46/31/25/17/16/2': (diff['仅命名不同'], diff['PA优势'], diff['实现路径不同'], diff['待验证'], diff['天融信优势'], diff['天融信优势(待验证)']) == (46, 31, 25, 17, 16, 2),
    '需求58': len(reqs) == 58,
    'V清单92': len(vitems) == 92,
    '高优先级14': vprio['高'] == 14,
}
for k, ok in nums.items():
    check(f'交叉验证 {k}', ok, '')

# 报告大类支持率抽查
rate_ok = True
detail = []
expect = {'A': (90.6, 78.1), 'B': (68.4, 84.2), 'C': (90.9, 59.1), 'E': (42.9, 71.4), 'F': (66.7, 95.2), 'G': (38.5, 69.2)}
cat_stat = {}
for x in rows:
    c = x[1][:1]
    cat_stat.setdefault(c, Counter())[x[4]] += 1
    cat_stat.setdefault('PA' + c, Counter())[x[6]] += 1
for c, (ets, epa) in expect.items():
    n = sum(cat_stat[c].values())
    ats = round(cat_stat[c]['完全支持'] / n * 100, 1)
    apa = round(cat_stat['PA' + c]['完全支持'] / sum(cat_stat['PA' + c].values()) * 100, 1)
    if (ats, apa) != (ets, epa):
        rate_ok = False
        detail.append(f'{c}: 报告{ets}/{epa} 实际{ats}/{apa}')
check('交叉验证 八大类支持率', rate_ok, '; '.join(detail))

# ============ 输出 ============
fails = [x for x in R if x[1] == 'FAIL']
with open(BASE + r'\_scripts\p6_review_result.txt', 'w', encoding='utf-8') as f:
    f.write('P6 评审程序化核验结果（2026-08-15）\n' + '=' * 60 + '\n')
    for name, res, d in R:
        f.write(f'[{res}] {name}' + (f' | {d}' if d else '') + '\n')
    f.write('=' * 60 + f'\n总计: {len(R)} 项 | PASS: {len(R)-len(fails)} | FAIL: {len(fails)}\n')
print(f'总计: {len(R)} 项 | PASS: {len(R)-len(fails)} | FAIL: {len(fails)}')
for name, res, d in R:
    print(f'[{res}] {name}' + (f' | {d}' if d else ''))
