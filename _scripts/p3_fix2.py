from openpyxl import load_workbook

F = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求\对比工作底稿.xlsx'
wb = load_workbook(F)

# 1) 矩阵#39(r45)：差异说明与备注去除P2试填残留（C5-C8/C10-C11已为P3正式内容）
ws = wb['功能对比矩阵']
r = 45
assert str(ws.cell(r, 1).value).strip() == '39' and ws.cell(r, 5).value == '完全支持'
ws.cell(r, 9).value = '天融信优势'  # TS内置IPS规则集 vs PA需Threat Prevention订阅
ws.cell(r, 12).value = None
print('矩阵#39 差异→天融信优势, 备注已清')

# 2) 待验证清单：表头行R3去掉V1编号，数据行R4起重编V1..Vn
wv = wb['待验证移交清单']
assert '来源' in str(wv.cell(3, 2).value or '')
wv.cell(3, 1).value = None
n = 0
for r in range(4, wv.max_row + 1):
    if wv.cell(r, 2).value or wv.cell(r, 3).value:
        n += 1
        wv.cell(r, 1).value = f'V{n}'
print(f'待验证清单：表头已去编号，数据重编 V1-V{n}')

wb.save(F)

# 3) 复核
wb2 = load_workbook(F)
wv2 = wb2['待验证移交清单']
ids = [wv2.cell(r, 1).value for r in range(4, wv2.max_row + 1) if wv2.cell(r, 2).value]
print('复核: 表头R3编号=', wv2.cell(3, 1).value, '| 数据条数=', len(ids),
      '| 连续=', ids == [f'V{i}' for i in range(1, len(ids) + 1)])
ws2 = wb2['功能对比矩阵']
print('复核: #39 差异=', ws2.cell(45, 9).value, '| 备注=', ws2.cell(45, 12).value)
