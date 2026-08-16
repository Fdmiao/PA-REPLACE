# -*- coding: utf-8 -*-
from openpyxl import load_workbook
base="/Users/个人资料/trx/2026/PA替代"
wb=load_workbook(base+"/PA替代需求列表.xlsx", data_only=True)
print("sheets:", wb.sheetnames)
ws=wb["工作表1"]
print("尺寸:", ws.max_row, "x", ws.max_column)
print("\n=== 前12行完整内容 ===")
for r in range(1, 13):
    vals=[]
    for c in range(1, ws.max_column+1):
        v=ws.cell(r,c).value
        if v is not None:
            vals.append(f"{ws.cell(r,c).coordinate}={str(v)[:40]}")
    print(f"R{r}:", " | ".join(vals) if vals else "(空)")
wb.close()