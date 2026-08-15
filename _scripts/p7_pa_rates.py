# -*- coding: utf-8 -*-
from collections import Counter
from openpyxl import load_workbook
import os

BASE = r"e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求"
wb = load_workbook(os.path.join(BASE, "天融信vsPA功能对比.xlsx"), data_only=True)
mx = wb['功能对比矩阵']
cat = {}
for r in range(4, mx.max_row + 1):
    c2 = str(mx.cell(r, 2).value or '').strip()
    ts = str(mx.cell(r, 5).value or '').strip()
    if not c2 or not ts:
        continue
    k = c2[:1]
    cat.setdefault(k, {'ts': Counter(), 'pa': Counter()})
    cat[k]['ts'][ts] += 1
    cat[k]['pa'][str(mx.cell(r, 7).value or '').strip()] += 1
for k in sorted(cat):
    n_ts = sum(cat[k]['ts'].values())
    n_pa = sum(cat[k]['pa'].values())
    rts = cat[k]['ts']['完全支持'] / n_ts * 100
    rpa = cat[k]['pa']['完全支持'] / n_pa * 100
    print(f"{k}: 点数{n_ts} TS完全率{rts:.1f}% PA完全率{rpa:.1f}% PA分布{dict(cat[k]['pa'])}")

rm = wb['需求映射']
print("\n需求映射 max_row:", rm.max_row)
n = sum(1 for r in range(4, rm.max_row + 1) if rm.cell(r, 2).value)
print("需求映射数据行(col2非空):", n)
n1 = sum(1 for r in range(4, rm.max_row + 1) if rm.cell(r, 1).value)
print("需求映射数据行(col1非空):", n1)
# 看表头
for c in range(1, 8):
    print(f"  col{c}: {rm.cell(3, c).value}")
