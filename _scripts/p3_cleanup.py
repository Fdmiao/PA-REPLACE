from openpyxl import load_workbook

F = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求\对比工作底稿.xlsx'
wb = load_workbook(F)

# 1) 矩阵备注列清除修正错位残留标记
ws = wb['功能对比矩阵']
cleaned = []
for r in range(4, ws.max_row + 1):
    note = ws.cell(r, 12).value
    if note and 'P3填充(修正错位)' in str(note):
        no = ws.cell(r, 1).value
        rest = str(note).replace('P3填充(修正错位)', '').strip(' ;；,，')
        ws.cell(r, 12).value = rest if rest else None
        cleaned.append(no)
print('矩阵备注清理:', cleaned)

# 2) 待验证清单删除假阳性 V11/V12/V13/V14/V17/V18
wv = wb['待验证移交清单']
FALSE_POS = {'V11', 'V12', 'V13', 'V14', 'V17', 'V18'}
drop_rows = []
for r in range(3, wv.max_row + 1):  # 数据从第3行起(1标题/2表头)
    vid = wv.cell(r, 1).value
    if vid and str(vid).strip() in FALSE_POS:
        drop_rows.append(r)
print('待验证删除行:', drop_rows, [wv.cell(r,1).value for r in drop_rows])
for r in sorted(drop_rows, reverse=True):
    wv.delete_rows(r)

# 3) 重新编号 V 列
n = 0
for r in range(3, wv.max_row + 1):
    if wv.cell(r, 2).value or wv.cell(r, 3).value:
        n += 1
        wv.cell(r, 1).value = f'V{n}'
print('待验证清单重编后条数:', n)

wb.save(F)
print('saved.')

# 4) 验证：V编号连续性 + 总数
wb2 = load_workbook(F)
wv2 = wb2['待验证移交清单']
ids = [wv2.cell(r, 1).value for r in range(3, wv2.max_row + 1) if wv2.cell(r, 2).value]
expect = [f'V{i}' for i in range(1, len(ids) + 1)]
print('编号连续:', ids == expect, '| 总数:', len(ids))
