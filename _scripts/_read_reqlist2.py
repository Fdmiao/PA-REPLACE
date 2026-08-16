# -*- coding: utf-8 -*-
from openpyxl import load_workbook
base="/Users/个人资料/trx/2026/PA替代"
wb=load_workbook(base+"/PA替代需求列表.xlsx", data_only=True)
ws=wb["工作表1"]
# 打印所有非空行 A/B/C 摘要
print("=== 工作表1 全部行 (A/B/C/E/F/G) ===")
for r in range(1, ws.max_row+1):
    a=ws.cell(r,1).value; b=ws.cell(r,2).value; c=ws.cell(r,3).value
    e=ws.cell(r,5).value; f=ws.cell(r,6).value; g=ws.cell(r,7).value
    if a or b or c:
        print(f"R{r}: A={str(a)[:18] if a else ''} | B={str(b)[:18] if b else ''} | C={str(c)[:26] if c else ''} | E={e} | F={str(f)[:14] if f else ''} | G={g}")
wb.close()