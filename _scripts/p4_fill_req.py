import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font

BASE = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求'
TODAY = '2026-08-15'
wb = load_workbook(BASE + r'\对比工作底稿.xlsx')
src = load_workbook(BASE + r'\PA替代需求列表.xlsx', read_only=True)

ws = wb['需求对比表']      # R3表头，R4-R61数据
wm = wb['需求映射']         # R3起数据
mx = wb['功能对比矩阵']     # R4起数据
ss = src['工作表1']         # R2-R59数据

FULL, PART, SUB, NO, TBD = '完全支持', '部分支持', '需订阅授权', '不支持', '待验证'
FILL = {FULL: PatternFill('solid', fgColor='E2EFDA'), PART: PatternFill('solid', fgColor='FFF2CC'),
        SUB: PatternFill('solid', fgColor='E4DFEC'), NO: PatternFill('solid', fgColor='FBDBDB'),
        TBD: PatternFill('solid', fgColor='EDEDED')}

# 矩阵行索引：序号→行
mrow = {}
for r in range(4, mx.max_row + 1):
    no = mx.cell(r, 1).value
    if no is not None and str(no).strip().isdigit():
        mrow[int(str(no).strip())] = r

# 需求映射：R编号→(主F, 结论, 取证说明)
rmap = {}
for r in range(3, 61):
    rid = wm.cell(r, 1).value
    if rid is None:
        continue
    c3 = str(wm.cell(r, 3).value or '')
    nums = re.findall(r'([A-H])?(\d+)', c3)
    main_f = int(nums[0][1]) if nums else None
    rmap[str(rid).strip()] = (main_f, str(wm.cell(r, 5).value or ''), str(wm.cell(r, 6).value or ''))

# 原表备注：R编号(行序)→备注
snote = {}
for i, r in enumerate(range(2, 60), 1):
    note = ss.cell(r, 6).value
    if note and str(note).strip() and str(note).strip() != 'None':
        snote[i] = str(note).strip().replace('\n', '；')


def diff_label(concl, mfr):
    mdiff = str(mx.cell(mrow[mfr], 9).value or '').strip() if mfr in mrow else ''
    if 'TS占优' in concl:
        return '天融信优势'
    if '天融信缺口' in concl:
        return '天融信缺口'
    if 'PA优势' in concl and 'TS待实测' in concl:
        return 'PA优势·TS待实测'
    if 'PA优势' in concl:
        return 'PA优势'
    if '待查资料' in concl:
        return '待查资料（规格项）'
    if '待实测' in concl:
        return (mdiff + '（待实测）') if mdiff else '待实测'
    return mdiff or '仅命名不同'


n_fill = 0
for i, r in enumerate(range(4, 62), 1):  # 需求对比表行 ← R1..R58
    rid = f'R{i}'
    if rid not in rmap:
        continue
    main_f, concl, note = rmap[rid]
    assert main_f in mrow, f'{rid} 主F{main_f} 不在矩阵'
    mr = mrow[main_f]
    # C6 原状态并入原表备注
    if i in snote:
        cur = str(ws.cell(r, 6).value or '')
        if '原表备注' not in cur:
            ws.cell(r, 6).value = f'{cur}（原表备注：{snote[i]}）'
    # C8-C11 双侧支持度与说明
    ws.cell(r, 8).value = mx.cell(mr, 5).value
    ws.cell(r, 8).fill = FILL.get(str(mx.cell(mr, 5).value), PatternFill())
    ws.cell(r, 9).value = mx.cell(mr, 6).value
    ws.cell(r, 10).value = mx.cell(mr, 7).value
    ws.cell(r, 10).fill = FILL.get(str(mx.cell(mr, 7).value), PatternFill())
    ws.cell(r, 11).value = mx.cell(mr, 8).value
    # C12 差异定性 + 取证说明
    ws.cell(r, 12).value = f'{diff_label(concl, main_f)}：{note}'
    # C13/C14
    ws.cell(r, 13).value = f'[{rid}→主证据矩阵#{main_f}] ' + str(mx.cell(mr, 10).value or '')
    ws.cell(r, 14).value = TODAY
    # 说明列换行
    for c in (9, 11, 12):
        ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical='top')
    n_fill += 1

wb.save(BASE + r'\对比工作底稿.xlsx')
print(f'需求对比表已填充 {n_fill}/58 行')

# 复核
wb2 = load_workbook(BASE + r'\对比工作底稿.xlsx')
w2 = wb2['需求对比表']
miss = [i for i, r in enumerate(range(4, 62), 1) if not w2.cell(r, 8).value or not w2.cell(r, 12).value]
print('未填行:', miss if miss else '无')
dist = {}
for r in range(4, 62):
    d = str(w2.cell(r, 12).value or '').split('：')[0]
    if d:
        dist[d] = dist.get(d, 0) + 1
print('差异定性分布:', dict(sorted(dist.items(), key=lambda x: -x[1])))
