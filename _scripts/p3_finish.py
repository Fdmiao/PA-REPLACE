# -*- coding: utf-8 -*-
"""P3-7 收尾：需求映射 58 行双重取证结论 + 待验证移交清单汇总"""
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

F = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求\对比工作底稿.xlsx'
TODAY = '2026-08-15'

# ============ 1. 需求映射 58 行：col5=结论 col6=取证说明 ============
# 结论前缀：已取证·双方支持 / 天融信缺口 / PA优势 / 天融信优势 / 待实测 / 待查资料
M = {
 1: ('已取证·双方支持', 'F84双侧完全支持（TS外发邮件账户+PA邮件通知）；SMTPS加密细节待实测→V清单'),
 2: ('已取证·天融信缺口', 'F83：TS无公开XML/REST API文档（CHM仅"第三方管理接口"），FireMon对接依赖API通道'),
 3: ('已取证·天融信缺口', 'F83：Ansible经API+Playbook对接同受API缺口影响'),
 4: ('已取证·天融信缺口', 'F83：RESTful API为R2/R3的总通道，PA XML+REST双API完全支持'),
 5: ('已取证·双方支持', 'F84：TS日志查看+PA日志筛选均支持；源/目的IP字段级过滤待实测→V清单'),
 6: ('已取证·双方支持', '同R5（目的IP过滤）'),
 7: ('已取证·双方支持', '同R5（源端口过滤）'),
 8: ('已取证·双方支持', '同R5（目的端口过滤）'),
 9: ('已取证·待实测', 'F84：PA日志字段体系有据（会话开始/结束时间）；TS连接日志字段明细待实测→V清单'),
 10: ('已取证·待实测', '同R9（会话结束时间）'),
 11: ('已取证·待实测', '同R9：PA解密日志含结束原因字段；TS待实测'),
 12: ('已取证·双方支持', 'F91：TS License升级+PA订阅续期；升级不中断业务待实测→V清单'),
 13: ('已取证·PA优势·TS待实测', 'B24+B36：PA App-ID全球库；TS国际应用（Facebook）识别覆盖待抽样实测→V清单'),
 14: ('已取证·PA优势·TS待实测', 'B24+B36：YouTube，同R13港澳应用抽样清单'),
 15: ('已取证·PA优势·TS待实测', 'B24+B36：Twitter(X)，同R13'),
 16: ('已取证·PA优势·TS待实测', 'B24+B36：WhatsApp，同R13'),
 17: ('已取证·PA优势·TS待实测', 'B24+B36：TikTok，同R13'),
 18: ('已取证·PA优势·TS待实测', 'B24+B36：Zoom，同R13'),
 19: ('已取证·PA优势·TS待实测', 'B24+B36：Steam，同R13'),
 20: ('已取证·PA优势·TS待实测', 'B24+B27+B36：Office365含子应用识别；TS覆盖待实测'),
 21: ('已取证·PA优势·TS待实测', 'B24+B36：TeamViewer，同R13'),
 22: ('已取证·双方支持', 'E74：TS密码+证书/UKey/短信与PA MFA供应商体系均完全支持'),
 23: ('已取证·待查资料', 'A5：地址组嵌套层数（8层）属容量规格，双方需查datasheet→V清单'),
 24: ('已取证·待实测', 'F92：TS升级服务器+PA更新服务器均支持；经代理服务器升级双方手册未载→V清单'),
 25: ('已取证·天融信缺口', 'F135：TS无配置锁机制（检索无命中）；PA Config Lock/Commit Lock完全支持——PA优势'),
 26: ('已取证·双方支持', 'A10+F80：TS独立管理路由分离显式支持；PA MGT带外+服务路由'),
 27: ('已取证·双方支持', 'A10+F80：带外管理认证路由同R26'),
 28: ('已取证·待查资料', 'F90：双方手册均无第三方配置转换工具章节；需查迁移工具（PA Expedition/天融信迁移工具）→V清单'),
 29: ('已取证·待查资料', 'F90：PA配置转换（本替代项目反向场景：PA→天融信工具为关键）→V清单'),
 30: ('已取证·待查资料', 'F90：Hillstone转换工具同R28'),
 31: ('已取证·待实测', 'B33：双方QoS完全支持；RED/尾丢弃队列算法细节未载→V清单'),
 32: ('已取证·待实测', 'B33：尾丢弃同R31'),
 33: ('已取证·双方支持·TS占优', 'A23：TS组播路由完全支持/PA部分支持；IGMP随A23证据'),
 34: ('已取证·双方支持·TS占优', 'A23：MLD（IPv6组播）随A23'),
 35: ('已取证·双方支持·TS占优', 'A23：PIM-SM随A23'),
 36: ('已取证·双方支持·TS占优', 'A23：PIM-DM随A23'),
 37: ('已取证·双方支持', 'A2+A5：策略源/目的域名（TS域名对象+PA FQDN对象）'),
 38: ('已取证·待实测', 'C39+A23：组播报文IPS检测双方手册未显式载明→V清单'),
 39: ('已取证·双方支持', 'D57：TS内置分类库/PA PAN-DB订阅，实现路径不同'),
 40: ('已取证·待实测', 'C53+B28：PA手册建议阻止QUIC（无法解密）；HTTP3解密WAF双方均未见→V清单'),
 41: ('已取证·待实测', 'C53+F84：XFF原始IP解析上报双方手册未载→V清单'),
 42: ('已取证·双方支持', 'C46：TS联动第三方沙箱/PA WildFire云订阅（需订阅授权），仅命名不同'),
 43: ('已取证·待实测', 'F80：PA多语言显式支持；TS界面英文切换待实测→V清单'),
 44: ('已取证·待实测', 'B24+F92：TS日志语言中英文可选；规则库/界面语言待实测'),
 45: ('已取证·待查资料', 'F80：PA中英文手册齐备；TS英文版手册需向原厂查询→V清单'),
 46: ('已取证·双方支持', 'F84：TS日志语言中/英文可选显式支持；PA英文原生'),
 47: ('已取证·待实测', 'F86：报表英文输出TS待实测'),
 48: ('已取证·待查资料', 'G97：网络层/IPS/DLP/IPSec加速能力属硬件规格，需查双方datasheet→V清单'),
 49: ('已取证·待查资料', 'G97：IPS加速同R48'),
 50: ('已取证·待查资料', 'G97：DLP内容检查加速同R48'),
 51: ('已取证·待查资料', 'G97：IPSec硬件加速同R48'),
 52: ('已取证·双方支持', 'F87+F88：TS显式支持整机/策略/管理员/实时四级配置同步；PA配置同步完整'),
 53: ('已取证·待实测', 'A19+F87：虚系统级HA（跨设备虚系统互备）双方手册未载→V清单'),
 54: ('已取证·待查资料', 'A5：NAT策略组容量512条属规格项，需查datasheet'),
 55: ('已取证·待查资料', 'G96：25G/40G/2.5G/5G接口速率与国产型号需查产品目录→V清单'),
 56: ('已取证·PA优势·TS待实测', 'G96：PA收发器监控（show transceiver光衰）有据；TS光衰查看待实测→V清单'),
 57: ('已取证·双方支持', 'E77：TS DPD+链路备份+策略路由/PA Tunnel Monitor+Path Monitoring；切换时间待实测'),
 58: ('已取证·PA优势', 'B124：PA解密镜像接口+许可证完全支持；TS部分支持（解密排除有据，镜像输出待实测）'),
}

wb = load_workbook(F)
ws = wb['需求映射']
GREEN = PatternFill('solid', fgColor='E2EFDA')
RED = PatternFill('solid', fgColor='FBDBDB')
YEL = PatternFill('solid', fgColor='FFF2CC')
for r in range(3, 61):
    rid = str(ws.cell(r, 1).value).strip()  # R1..R58
    n = int(rid[1:])
    concl, note = M[n]
    ws.cell(r, 5, concl + f'（{TODAY} P3矩阵取证）')
    ws.cell(r, 6, note)
    if '天融信缺口' in concl:
        ws.cell(r, 5).fill = RED
    elif '待实测' in concl or '待查资料' in concl:
        ws.cell(r, 5).fill = YEL
    else:
        ws.cell(r, 5).fill = GREEN

# ============ 2. 待验证移交清单 ============
mt = wb['功能对比矩阵']
# 方法映射：按关键词归类 T1~T8 或"资料核对"
def method_of(txt, fp):
    t = fp + txt
    if any(k in t for k in ['识别', '覆盖', '应用库', '样本']): return 'T1 应用识别实测'
    if any(k in t for k in ['解密', 'SSL', '镜像', 'QUIC', 'HTTP3']): return 'T2 SSL解密测试'
    if any(k in t for k in ['IPS', '沙箱', '威胁', '检出', 'XFF', 'WAF']): return 'T3 攻击样本回放'
    if any(k in t for k in ['URL', '分类']): return 'T4 URL分类实测'
    if any(k in t for k in ['VPN', 'IPSec', 'IKE', '国密', '隧道', 'SM2', 'DPD', '链路切换']): return 'T5 VPN实测'
    if any(k in t for k in ['HA', '会话', '切换', '同步']): return 'T6 HA实测'
    if any(k in t for k in ['API', 'syslog', '日志', 'SIEM', '对接']): return 'T7 日志与API'
    if any(k in t for k in ['界面', '报表', '配置', '升级', 'License', '授权', '管理']): return 'T8 管理任务'
    return '资料核对（L3）'

items = []  # (来源, 功能点, 原因, 方法, 优先级)
seen = set()
HIGH = ['API', '会话', 'HA', '识别', '国密', '商用密码', 'IKEv2', 'IPSec', '型号', '覆盖', '认证', '信创', '销售许可']
for r in range(4, mt.max_row + 1):
    no = mt.cell(r, 1).value
    cat = mt.cell(r, 2).value
    if cat is None or str(no) == '序号' or (isinstance(no, str) and not no.strip().isdigit()):
        continue
    fp = str(mt.cell(r, 4).value or '')
    ts, ps = mt.cell(r, 5).value or '', mt.cell(r, 7).value or ''
    note = str(mt.cell(r, 12).value or '')
    reasons = []
    if ts == '待验证': reasons.append('天融信侧待验证')
    if ps == '待验证': reasons.append('PA侧待验证')
    m = re.match(r'^R\d+\s*(.+)$', note)
    rnote = m.group(1) if m else (note if (note and note != 'P3填充') else '')
    if reasons:
        reason = '；'.join(reasons) + ('。' + rnote if rnote else '')
        key = f'M{no}'
        if key not in seen:
            seen.add(key)
            pri = '高' if any(k in fp + reason for k in HIGH) else '中'
            items.append((f'矩阵#{no}', fp, reason, method_of(reason, fp), pri))
    elif rnote and not note.endswith('P3填充'):
        key = f'M{no}'
        if key not in seen:
            seen.add(key)
            items.append((f'矩阵#{no}', fp, rnote, method_of(rnote, fp), '中'))

# 需求表衍生的待实测项（R清单中"待实测/待查资料"合并同类）
REQ_V = [
 ('需求R13-R21', '港澳国际应用识别覆盖（Facebook/YouTube/Twitter/WhatsApp/TikTok/Zoom/Steam/O365/TeamViewer）', 'TS识别库国际应用覆盖度无量化数据；PA App-ID全球库', 'T1 应用识别实测', '高'),
 ('需求R5-R8', '日志字段级过滤（源/目的IP、端口）', '双方支持日志能力但字段级过滤明细未逐项确认', 'T7 日志与API', '中'),
 ('需求R9-R11', '连接日志字段（会话开始/结束时间、结束原因）', 'PA有据；TS字段明细待实测', 'T7 日志与API', '中'),
 ('需求R1', 'SMTPS加密邮件告警', '双方邮件告警有据，SMTPS加密方式未载', 'T7 日志与API', '中'),
 ('需求R12', 'License平滑升级（不中断业务）', '升级方式有据，平滑性未载', 'T8 管理任务', '中'),
 ('需求R23/R54', '对象与策略容量（地址组嵌套8层/NAT策略组512条）', '容量规格手册不载', '资料核对（L3）', '中'),
 ('需求R24', '经代理服务器升级规则库', '双方手册未载代理升级方式', 'T8 管理任务', '中'),
 ('需求R28-R30', '第三方配置转换工具（Fortinet/PA/Hillstone→天融信）', '手册均无转换工具章节；替代项目核心迁移手段', '资料核对（L3）', '高'),
 ('需求R31-R32', 'QoS拥塞算法（RED/尾丢弃）', '队列算法细节未载', 'T8 管理任务', '低'),
 ('需求R38', '组播报文IPS检测', '双方未显式载明', 'T3 攻击样本回放', '中'),
 ('需求R40', 'HTTP3/QUIC WAF检测', 'PA建议阻止QUIC；解密检测双方未见', 'T2 SSL解密测试', '高'),
 ('需求R41', 'XFF原始客户端IP解析上报', '双方手册未载', 'T3 攻击样本回放', '中'),
 ('需求R43-R45/R47', '界面/规则库/报表英文与英文手册', 'PA多语言+中英手册有据；TS英文能力待确认', 'T8 管理任务', '中'),
 ('需求R48-R51', '硬件加速（网络层/IPS/DLP/IPSec）', '硬件规格手册不载', '资料核对（L3）', '中'),
 ('需求R53', '虚系统级HA（跨设备互备）', '双方手册未载', 'T6 HA实测', '中'),
 ('需求R55', '接口速率清单（25G/40G/2.5G/5G）与国产型号', '产品目录范围', '资料核对（L3）', '中'),
 ('需求R56', '接口光衰信息查看', 'PA show transceiver有据；TS待实测', 'T8 管理任务', '低'),
 ('需求R57', 'IPSec链路切换时间', '机制双方有据；切换耗时未量化', 'T5 VPN实测', '高'),
 ('需求R58', '解密流量镜像输出', 'TS部分支持，镜像输出接口待实测', 'T2 SSL解密测试', '中'),
]
for it in REQ_V:
    key = it[0]
    if key not in seen:
        seen.add(key)
        items.append(it)

pv = wb['待验证移交清单']
# 重建数据区
if pv.max_row > 3:
    pv.delete_rows(4, pv.max_row - 3)
hdr_font = Font(bold=True)
for i, (src, fp, reason, method, pri) in enumerate(items, 1):
    r = 3 + i
    pv.cell(r, 1, f'V{i}')
    pv.cell(r, 2, src)
    pv.cell(r, 3, fp)
    pv.cell(r, 4, reason)
    pv.cell(r, 5, method)
    pv.cell(r, 6, pri)
    pv.cell(r, 7, '待验证')
    pv.cell(r, 8, '')
    if pri == '高':
        pv.cell(r, 6).fill = RED
# 列宽
for col, w in zip('ABCDEFGH', [8, 18, 34, 46, 16, 8, 10, 20]):
    pv.column_dimensions[col].width = w

wb.save(F)
print(f'需求映射 58 行已更新；待验证清单 {len(items)} 项')
print('高优先级：', sum(1 for i in items if i[4] == '高'))
