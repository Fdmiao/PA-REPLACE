from openpyxl import load_workbook

BASE = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求'
for F in [BASE + r'\天融信vsPA功能对比.xlsx', BASE + r'\对比工作底稿.xlsx']:
    wb = load_workbook(F)
    mx = wb['功能对比矩阵']
    fixed = []
    for r in range(4, mx.max_row + 1):
        no = str(mx.cell(r, 1).value or '')
        if no == '116':
            mx.cell(r, 10).value = '天融信CHM:全文检索无命中（P3否定证据留痕）；PA:Web帮助>Network Profiles>MACsec(11.1)'
            fixed.append('#116证据')
        elif no == '36':
            cur = str(mx.cell(r, 12).value or '')
            if 'V12' not in cur:
                mx.cell(r, 12).value = cur + '（V12/V74）'
            fixed.append('#36备注')
    try:
        wb.save(F)
        print(F.split('\\')[-1], '→', fixed)
    except PermissionError:
        print(F.split('\\')[-1], '→', fixed, '[保存失败: 文件被占用]')
