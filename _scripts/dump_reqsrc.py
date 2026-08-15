from openpyxl import load_workbook

SRC = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求\PA替代需求列表.xlsx'
wb = load_workbook(SRC, read_only=True)
ws = wb['工作表1']
out = []
for r in range(2, ws.max_row + 1):
    rid = ws.cell(r, 3).value
    note = ws.cell(r, 6).value
    status = ws.cell(r, 7).value
    if note or status:
        out.append(f'{r-1:02d} | {str(rid)[:35]} | 备注:{str(note)[:60]} | 状态:{status}')
with open(r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求\_scripts\req_src_notes.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(out))
print('rows with note/status:', len(out), '/', ws.max_row - 1)
