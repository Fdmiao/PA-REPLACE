from openpyxl import load_workbook

BASE = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求'
wb = load_workbook(BASE + r'\对比工作底稿.xlsx')
ws = wb['需求对比表']

lines = []
for r in range(4, ws.max_row + 1):
    l1, l2, l3, spec = [ws.cell(r, c).value for c in (1, 2, 3, 4)]
    if l3:
        lines.append(f"R{r - 3} | {l1} / {l2} / {l3} | {str(spec)[:60] if spec else ''}")

with open(BASE + r'\_scripts\req58.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))
print('reqs:', len(lines))
