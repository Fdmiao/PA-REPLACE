from openpyxl import load_workbook
from collections import Counter

BASE = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求'
wb = load_workbook(BASE + r'\天融信vsPA功能对比.xlsx')
src = load_workbook(BASE + r'\PA替代需求列表.xlsx', read_only=True)

rq = wb['需求对比表']
ss = src['工作表1']

dst = []   # 底稿 C6 (去原表备注括号)
origin = []  # 原表 C7
for i, r in enumerate(range(4, 62)):
    v = str(rq.cell(r, 6).value or '')
    dst.append(v.split('（原表备注')[0].strip())
for r in range(2, 60):
    origin.append(str(ss.cell(r, 7).value or '').strip())

print('底稿C6分布:', dict(Counter(dst)))
print('原表C7分布:', dict(Counter(origin)))
diff = [(i + 1, origin[i], dst[i]) for i in range(58) if origin[i] != dst[i]]
print(f'\n不一致 {len(diff)} 条:')
for rid, o, d in diff:
    print(f'  R{rid}: 原表[{o}] vs 底稿[{d}]')
