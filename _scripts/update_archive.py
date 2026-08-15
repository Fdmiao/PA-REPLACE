from openpyxl import load_workbook

BASE = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求'
F = BASE + r'\天融信vsPA功能对比.xlsx'

archive = {
    2: '_sources/web/topsec/topsec-topngfw-product.html（254KB）',
    6: '_sources/web/pa/panos-12-1-new-features.html（659KB，新特性页快照；其余按功能点逐页取证）',
    7: '_sources/docs/pa/pan-os-admin-11-0-zh-cn.pdf（3.4MB）',
    8: '_sources/web/pa/pa-ngfw-hardware.html（1.0MB）',
    10: '_sources/web/pa/panos-certifications.html（286KB）',
    11: '_sources/docs/pa/cc-st_vid11482-vr.pdf（414KB）',
    13: '_sources/web/topsec/topsec-news-4441-cert.html（234KB）',
    15: '_sources/docs/pa/pa-3400-series-datasheet.pdf（431KB）',
}

wb = load_workbook(F)
ws = wb['资料登记表']
for r in range(4, 32):
    no = ws.cell(r, 1).value
    if no in archive:
        ws.cell(r, 10, archive[no])
wb.save(F)
print('archive paths updated')
