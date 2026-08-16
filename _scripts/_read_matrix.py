# -*- coding: utf-8 -*-
from openpyxl import load_workbook
base="/Users/个人资料/trx/2026/PA替代"
wb=load_workbook(base+"/天融信vsPA功能对比.xlsx", data_only=True)
ws=wb["功能对比矩阵"]
print("列:", [ws.cell(3,c).value for c in range(1,ws.max_column+1)])
# 打印全部功能点 A,B,C,D,E,F,G,H,I,J
for r in range(4, ws.max_row+1):
    no=ws.cell(r,1).value
    if no is None or not str(no).strip(): continue
    cat=ws.cell(r,2).value; mod=ws.cell(r,3).value; fn=ws.cell(r,4).value
    e=ws.cell(r,5).value; f=ws.cell(r,6).value; g=ws.cell(r,7).value; h=ws.cell(r,8).value; i=ws.cell(r,9).value
    print(f"#{no} [{cat}|{mod}] {fn}")
    if f: print(f"     TS: {str(f)[:110]}")
    if h: print(f"     PA: {str(h)[:110]}")
    if i: print(f"     差异: {str(i)[:90]}")
wb.close()