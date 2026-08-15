# -*- coding: utf-8 -*-
# P7 汇总报告数据：矩阵大类/V清单/资料/文件清单
import os
from collections import Counter
from openpyxl import load_workbook

BASE = r"e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求"
wb = load_workbook(os.path.join(BASE, "天融信vsPA功能对比.xlsx"), data_only=True)

# --- 矩阵大类 ---
mx = wb['功能对比矩阵']
cat_names = {}
cat_cnt = Counter()
ts_by_cat = {}
for r in range(4, mx.max_row + 1):
    c2 = str(mx.cell(r, 2).value or '').strip()
    ts = str(mx.cell(r, 5).value or '').strip()
    if not c2:
        continue
    key = c2[:1]
    if c2 and not ts:  # 大类标题行
        cat_names[key] = c2
    else:
        cat_cnt[key] += 1
        ts_by_cat.setdefault(key, Counter())[ts] += 1
print("== 大类计数 ==")
for k in sorted(cat_names):
    n = cat_cnt.get(k, 0)
    full = ts_by_cat.get(k, Counter())['完全支持']
    print(f"  {cat_names[k]}: {n} 点, TS完全支持率 {full/n*100:.1f}%" if n else f"  {cat_names[k]}: 0")

# --- V清单 ---
wv = wb['待验证移交清单']
print("\n== V清单表头 ==")
for c in range(1, 10):
    v = wv.cell(3, c).value
    if v is not None:
        print(f"  col{c}: {str(v)[:40]}")
vp = Counter()
vt = Counter()
hi = []
for r in range(4, wv.max_row + 1):
    if not wv.cell(r, 1).value:
        continue
    p = str(wv.cell(r, 6).value or '')
    t = str(wv.cell(r, 5).value or '')
    vp[p] += 1
    vt[t[:2]] += 1
    if '高' in p:
        hi.append(f"{wv.cell(r,1).value}: {str(wv.cell(r,4).value or '')[:46]}")
print("\n== V优先级 ==", dict(vp))
print("== V验证方式前缀 ==", dict(vt))
print("== 高优清单 ==")
for h in hi:
    print(f"  {h}")

# --- 资料登记 ---
src = wb['资料登记表']
n_src = sum(1 for r in range(4, src.max_row + 1) if src.cell(r, 2).value)
print("\n== 资料登记份数 ==", n_src)

# --- 需求映射 ---
rm = wb['需求映射']
n_map = sum(1 for r in range(4, rm.max_row + 1) if rm.cell(r, 2).value)
print("== 需求映射条数 ==", n_map)

# --- 范围基线要素 ---
rb = wb['范围基线']
print("\n== 范围基线内容(前12行col1) ==")
cnt = 0
for r in range(1, min(rb.max_row + 1, 40)):
    v = rb.cell(r, 1).value
    if v and cnt < 12:
        print("  " + str(v)[:70])
        cnt += 1

# --- 交付物文件 ---
print("\n== 交付物文件 ==")
for name in ["天融信vsPA功能对比.xlsx", "对比工作底稿.xlsx", "对比执行计划.md",
             r"fw-comparison-report\fw-comparison-report.html"]:
    p = os.path.join(BASE, name)
    print(f"  {name}: {'%.0f KB' % (os.path.getsize(p)/1024) if os.path.exists(p) else 'MISSING'}")
for d in ["_sources", "_scripts"]:
    p = os.path.join(BASE, d)
    if os.path.isdir(p):
        n = sum(len(fs) for _, _, fs in os.walk(p))
        sz = sum(os.path.getsize(os.path.join(dp, f)) for dp, _, fs in os.walk(p) for f in fs) / 1048576
        print(f"  {d}\\: {n} 个文件, {sz:.1f} MB")

# --- 报告章节标题 ---
html = open(os.path.join(BASE, r"fw-comparison-report\fw-comparison-report.html"), encoding='utf-8').read()
import re
print("\n== 报告 h2 章节 ==")
for m in re.findall(r'<h2[^>]*>(.*?)</h2>', html)[:12]:
    print("  " + re.sub(r'<[^>]+>', '', m)[:60])
