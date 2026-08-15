import re
from openpyxl import load_workbook

BASE = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求'
wb = load_workbook(BASE + r'\天融信vsPA功能对比.xlsx')
mx = wb['功能对比矩阵']
wv = wb['待验证移交清单']
html = open(BASE + r'\fw-comparison-report\fw-comparison-report.html', encoding='utf-8').read()

# --- 红线1: 报告中 吞吐/延迟 上下文 ---
print('=== 红线1：报告规格词上下文 ===')
for w in ['吞吐', '延迟']:
    for m in re.finditer(w, html):
        print(f'  [{w}] ...{html[max(0,m.start()-40):m.start()+40]}...'.replace('\n', ' '))

# --- 红线1b: 矩阵含规格词的3行 ---
print('\n=== 红线1b：矩阵规格词行 ===')
for r in range(4, mx.max_row + 1):
    no = mx.cell(r, 1).value
    if no is None or not str(no).strip().isdigit():
        continue
    tsd, pad = str(mx.cell(r, 6).value or ''), str(mx.cell(r, 8).value or '')
    if any(w in tsd + pad for w in ['吞吐', '并发连接', '延迟', 'Gbps', 'pps']):
        print(f'  #{no} TS支持度={mx.cell(r,5).value} 差异={mx.cell(r,9).value}')
        print(f'     TS说明: {tsd[:100]}')

# --- 红线4: 5行证据内容 ---
print('\n=== 红线4：证据格式存疑行 ===')
for r in range(4, mx.max_row + 1):
    no = str(mx.cell(r, 1).value or '')
    if no in ('5', '60', '104', '107', '116'):
        print(f'  #{no}: {str(mx.cell(r, 10).value or "")[:110]}')

# --- 红线6a: 待验证点 V编号登记情况 ---
print('\n=== 红线6a：待验证15点登记情况 ===')
# V清单来源列
vsrc = set()
for r in range(4, wv.max_row + 1):
    if wv.cell(r, 2).value:
        s = str(wv.cell(r, 2).value)
        for m in re.findall(r'#(\d+)', s):
            vsrc.add(m)
for r in range(4, mx.max_row + 1):
    if str(mx.cell(r, 5).value) == '待验证':
        no = str(mx.cell(r, 1).value)
        note = str(mx.cell(r, 12).value or '')
        print(f'  #{no} {str(mx.cell(r,4).value)[:25]} | 备注:{note[:45]} | V清单登记:{"是" if no in vsrc else "否"}')

# --- 红线9: 确认误报（hex色值） ---
print('\n=== 红线9：#引用核实 ===')
refs = re.findall(r'>([^<]*)#(\d+)', html)
hexish = [m for m in re.findall(r'#(\d+)', html) if len(m) > 4]
print('长度>4的#数字（hex色值特征）:', hexish)
real_refs = {m for m in re.findall(r'#(\d{1,3})\b', html) if len(m) <= 3}
wb2 = load_workbook(BASE + r'\天融信vsPA功能对比.xlsx')
mx2 = wb2['功能对比矩阵']
valid = {str(mx2.cell(r, 1).value) for r in range(4, mx2.max_row + 1) if mx2.cell(r, 1).value}
bad = [x for x in real_refs if x not in valid]
print('真实#引用无效项:', bad or '无')
