# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import math

base="/Users/个人资料/trx/2026/PA替代"
path=base+"/PA替代需求列表.xlsx"
wb=load_workbook(path)
ws=wb["策略需求"]

# 0) 先取消全部合并，避免 MergedCell 只读
for rng in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(rng))

# 1) 向下填充 A(一级) B(二级) 值，使每组每行都有值，便于整组合并
for r in range(2, ws.max_row+1):
    if not ws.cell(r,1).value:
        ws.cell(r,1).value=ws.cell(r-1,1).value
    if not ws.cell(r,2).value:
        ws.cell(r,2).value=ws.cell(r-1,2).value

# 2) 表头样式
thin=Side(style="thin", color="000000")
border=Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_font=Font(name="等线", size=12, bold=True)
hdr_align=Alignment(horizontal="center", vertical="center", wrap_text=True)
for c in range(1, 9):
    cell=ws.cell(1, c)
    cell.font=hdr_font
    cell.alignment=hdr_align

# 3) 数据区样式：居中列 A/B/E/G/H，左对齐换行列 C/D/F
center_cols={1,2,5,7,8}
data_font=Font(name="等线", size=11)
for r in range(2, ws.max_row+1):
    for c in range(1, 9):
        cell=ws.cell(r, c)
        cell.font=data_font
        if c in center_cols:
            cell.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            cell.alignment=Alignment(horizontal="left", vertical="center", wrap_text=True)

# 4) 合并：A/B 整组合并；E/F/G/H 仅在组内合并连续相同值（参考工作表1）
def merge_runs(col, within_group=False):
    r=2
    while r<=ws.max_row:
        val=ws.cell(r, col).value
        end=r
        while end+1<=ws.max_row and ws.cell(end+1, col).value==val:
            if within_group and ws.cell(end+1, 2).value!=ws.cell(end, 2).value:
                break
            end+=1
        if end>r:
            ws.merge_cells(start_row=r, start_column=col, end_row=end, end_column=col)
        r=end+1

merge_runs(1)   # A 一级需求：整列"策略" → A2:A33
merge_runs(2)   # B 二级需求：每组一个合并块
for col in (5,6,7,8):
    merge_runs(col, within_group=True)  # E/F/G/H 组内合并

# 5) 合并后统一补边框（MergedCell 需重新描边）
for r in range(1, ws.max_row+1):
    for c in range(1, 9):
        ws.cell(r, c).border=border

# 6) 行高：按 C/D 列最长文本估算换行行数，最小 30
chars_per_line={3:24, 4:27}
for r in range(2, ws.max_row+1):
    max_lines=1
    for c, cpl in chars_per_line.items():
        v=ws.cell(r, c).value
        if v:
            lines=0
            for seg in str(v).split("\n"):
                lines+=max(1, math.ceil(len(seg)/cpl))
            max_lines=max(max_lines, lines)
    ws.row_dimensions[r].height=max(30, max_lines*15+6)
ws.row_dimensions[1].height=20

wb.save(path)
print("格式化完成")
print("合并范围数:", len(ws.merged_cells.ranges))
for rng in sorted(ws.merged_cells.ranges, key=lambda x:(x.min_col,x.min_row)):
    print(rng)
