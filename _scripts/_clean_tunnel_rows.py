# -*- coding: utf-8 -*-
from openpyxl import load_workbook

path = "/Users/个人资料/trx/2026/PA替代/PA替代需求列表.xlsx"
wb = load_workbook(path, rich_text=True)
ws = wb["策略需求差异"]

# 行19：修复句尾悬空逗号
ws["D19"] = "已解密/未解密/非TLS流量均可内联转发第三方安全链"

# 行20：删旧留新（红字版），E20补GRE与GTP-U前提，状态改部分支持
ws["B20"] = "隧道检测(Tunnel Inspection)"
ws["C20"] = "支持 GRE/GTP-U/非加密 IPSec/VXLAN 明文隧道内层流量检测"
ws["D20"] = "解封装并深度检测 GRE、GTP-U、NULL/AH IPSec、VXLAN 隧道内的嵌套流量"
ws["E20"] = ("PA隧道检测支持GRE/GTP-U/非加密IPSec（NULL加密/AH传输模式）/VXLAN明文隧道内层检测，"
             "GRE与非加密IPSec全型号支持，GTP-U仅支持GTP的防火墙可用（PA Web帮助p172-174）；"
             "天融信GRE隧道解封装后内层报文按普通流量检测（天融信手册《GRE隧道》页），"
             "VXLAN仅隧道封装转发与三层网关无内层安全检测，GTP检索无命中；对应矩阵#128/37")
ws["G20"] = "部分支持"

# 行21：删旧留新；D21红字与D20重复，按p174"最大隧道检测级别"实际描述重写，E21同步
ws["C21"] = "支持最大隧道检测级别控制（一级/两级）"
ws["D21"] = "指定检测一级或两级（Tunnel In Tunnel）封装级别，超过最大级别的封装数据包可选丢弃"
ws["E21"] = ("PA隧道检测可指定一级/两级（Tunnel In Tunnel）封装检测级别，超限封装数据包可选丢弃，"
             "VXLAN仅外层一级检测（PA Web帮助p174）；天融信手册无隧道检测级别控制（检索无命中）")

wb.save(path)
print("cleaned")
