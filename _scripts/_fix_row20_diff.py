# -*- coding: utf-8 -*-
from openpyxl import load_workbook

path = "/Users/个人资料/trx/2026/PA替代/PA替代需求列表.xlsx"
wb = load_workbook(path, rich_text=True)
ws = wb["策略需求差异"]

ws["C20"] = "支持明文隧道内层流量检测（GTP-U/VXLAN/非加密IPSec）"
ws["D20"] = "解封装并深度检测GTP-U/VXLAN/NULL加密(AH)IPSec隧道内层嵌套流量"
ws["E20"] = ("PA隧道检测支持GTP-U/VXLAN/非加密IPSec（NULL加密/AH传输模式）明文隧道内层检测，"
             "GTP-U仅支持GTP的防火墙可用（PA Web帮助p172-174）；"
             "天融信VXLAN仅隧道封装转发与三层网关、无内层安全检测（天融信手册《业务接入端》《三层VXLAN网关》页），"
             "GTP检测检索无命中；GRE天融信解封装后内层按普通流量检测（天融信手册《GRE隧道》页）已支持，"
             "故不列入本差异项；对应矩阵#128/37")
ws["G20"] = "不支持"

wb.save(path)
print("fixed")
