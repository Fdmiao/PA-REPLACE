from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

BASE = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求'
OUT = BASE + r'\天融信vsPA功能对比.xlsx'
TODAY = '2026-08-15'

F_TITLE = Font(name='微软雅黑', size=14, bold=True, color='1F2937')
F_HEAD = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
F_BODY = Font(name='微软雅黑', size=10, color='1F2937')
F_SECTION = Font(name='微软雅黑', size=11, bold=True, color='1F2937')
FILL_HEAD = PatternFill('solid', fgColor='1F4E79')
FILL_ZEBRA = PatternFill('solid', fgColor='F7F9FC')
FILL_KPI = PatternFill('solid', fgColor='EAF2FF')
THIN = Side(style='thin', color='D9DEE7')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)


def style_table(ws, head_row, max_row, max_col, widths, zebra_from=None):
    for c in range(1, max_col + 1):
        cell = ws.cell(head_row, c)
        cell.font = F_HEAD
        cell.fill = FILL_HEAD
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]
    zebra_from = zebra_from or (head_row + 1)
    for r in range(zebra_from, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.font = F_BODY
            cell.alignment = WRAP
            cell.border = BORDER
            if (r - zebra_from) % 2 == 1:
                cell.fill = FILL_ZEBRA
    ws.freeze_panes = ws.cell(head_row + 1, 1)


wb = Workbook()

# ---------- Sheet1 需求对比表 ----------
ws1 = wb.active
ws1.title = '需求对比表'
ws1['A1'] = '天融信 NGFW × Palo Alto NGFW · PA替代需求逐条对比（表式沿用 PA替代需求列表）'
ws1['A1'].font = F_TITLE
ws1.merge_cells('A1:N1')
ws1['A2'] = f'范围：产品功能层面；数据截止 {TODAY}；支持度五级：完全支持/部分支持/需订阅授权/不支持/待验证；每格证据见"证据出处"列'
ws1['A2'].font = Font(name='微软雅黑', size=9, color='6B7280')
ws1.merge_cells('A2:N2')

heads1 = ['一级需求', '二级需求', '三级需求', '需求规格', '优先级', '原状态', '负责人',
          '天融信/支持度', '天融信/参数与说明', 'PA/支持度', 'PA/参数与说明',
          '差异定性', '证据出处', '数据日期']
for c, h in enumerate(heads1, 1):
    ws1.cell(3, c, h)

df = pd.read_excel(BASE + r'\PA替代需求列表.xlsx', header=None)
rows = df.iloc[1:].values.tolist()
r = 4
last_l1 = last_l2 = last_l3 = ''
req_no = 0
for row in rows:
    l1, l2, l3, spec, pri, note, status, owner = [str(x) if pd.notna(x) else '' for x in row]
    if l1:
        last_l1 = l1
    if l2:
        last_l2 = l2
    if l3:
        last_l3 = l3
    req_no += 1
    vals = [last_l1, last_l2, last_l3, spec, pri, status, owner,
            '', '', '', '', '', '', '']
    for c, v in enumerate(vals, 1):
        ws1.cell(r, c, v)
    r += 1

style_table(ws1, 3, r - 1, 14, [10, 11, 26, 30, 6, 9, 8, 11, 34, 11, 34, 18, 26, 11])

# ---------- Sheet2 功能对比矩阵 ----------
ws2 = wb.create_sheet('功能对比矩阵')
ws2['A1'] = '天融信 NGFW × Palo Alto NGFW · 功能对比矩阵（A~H 大类，功能全集）'
ws2['A1'].font = F_TITLE
ws2.merge_cells('A1:L1')
ws2['A2'] = f'范围：功能能力 + 硬件/平台规格；性能（吞吐/并发/新建/延迟）不对比；数据截止 {TODAY}'
ws2['A2'].font = Font(name='微软雅黑', size=9, color='6B7280')
ws2.merge_cells('A2:L2')

heads2 = ['序号', '大类', '功能模块', '功能点', '天融信/支持度', '天融信/参数与说明',
          'PA/支持度', 'PA/参数与说明', '差异说明', '证据出处（文档+章节/页码）', '数据日期', '备注/验证编号']
for c, h in enumerate(heads2, 1):
    ws2.cell(3, c, h)

df2 = pd.read_excel(BASE + r'\防火墙功能对比矩阵_空表.xlsx', sheet_name='对比矩阵', header=None)
matrix_rows = df2.iloc[3:].values.tolist()
r = 4
for row in matrix_rows:
    vals = [str(x) if pd.notna(x) else '' for x in row[:12]]
    if vals[0] and not vals[2]:
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(r, c, v)
            cell.font = F_SECTION
            cell.fill = PatternFill('solid', fgColor='DCE6F1')
            cell.border = BORDER
            cell.alignment = WRAP
    else:
        for c, v in enumerate(vals, 1):
            ws2.cell(r, c, v)
    r += 1

style_table(ws2, 3, r - 1, 12, [6, 5, 10, 24, 11, 34, 11, 34, 18, 26, 11, 12])

# ---------- Sheet3 待验证移交清单 ----------
ws3 = wb.create_sheet('待验证移交清单')
ws3['A1'] = '待验证移交清单（矩阵/需求表中标"待验证"项登记于此，移交 P5 真机实测）'
ws3['A1'].font = F_TITLE
ws3.merge_cells('A1:H1')
heads3 = ['编号', '来源（矩阵序号/需求序号）', '功能点', '待验证原因', '验证方式（映射T1~T8）', '优先级', '状态', '结果记录（实测后填）']
for c, h in enumerate(heads3, 1):
    ws3.cell(3, c, h)
for i in range(1, 21):
    ws3.cell(3 + i, 1, i)
style_table(ws3, 3, 23, 8, [6, 16, 24, 28, 24, 8, 10, 28])

# ---------- Sheet4 资料登记表 ----------
ws4 = wb.create_sheet('资料登记表')
ws4['A1'] = '资料登记表（每收集一份文档登记一行；引用时记录章节/页码）'
ws4['A1'].font = F_TITLE
ws4.merge_cells('A1:J1')
heads4 = ['编号', '文档名称', '厂商', '类型', '版本', '发布日期', '获取渠道（含检索式/URL）', '对应大类', '登记日期', '本地归档']
for c, h in enumerate(heads4, 1):
    ws4.cell(3, c, h)

docs = [
    [1, '天融信防火墙系统一本通（NGFW一本通）', '天融信', 'L2 配置手册', '文档编号 V3.2406.37098', '©2025（文档版2024-06）',
     '本地文件：天融信防火墙系统一本通.chm（hh.exe -decompile 解包）', '全大类', TODAY, '_sources/topsec_manual/（431页HTML）'],
    [2, 'TopNGFW 下一代防火墙产品页', '天融信', 'L1 产品规格', '官网在售', '2025-12-22（产品文档上传日）',
     'https://www.topsec.com.cn/products/TopNGFW.html', '全大类', TODAY, '—'],
    [3, 'NGFW 产品文档（官网公开PDF）', '天融信', 'L1 规格/白皮书', '待核对', '2025-12-22（上传日）',
     'https://www.topsec.com.cn/uploads/2025-12-22/878eeb3b-4826-4dff-8faf-936401a619921766392900815.pdf', '全大类', TODAY, '待下载归档'],
    [4, 'PAN-OS 12.1 New Features（特性索引）', 'Palo Alto', 'L3 版本特性', '12.1', '2026-07-09（更新）',
     'https://docs.paloaltonetworks.com/ngfw/new-features/by-version/panos/12-1', '全大类', TODAY, '—'],
    [5, 'PAN-OS 12.1 Features Introduced（Release Notes）', 'Palo Alto', 'L3 Release Notes', '12.1（12.1.5于2026-03发布）', '2026-03',
     'https://origin-docs.paloaltonetworks.com/ngfw/release-notes/12-1/features-introduced-in-pan-os', '全大类', TODAY, '—'],
]
r = 4
for d in docs:
    for c, v in enumerate(d, 1):
        ws4.cell(r, c, v)
    r += 1
for i in range(len(docs) + 1, 26):
    ws4.cell(3 + i, 1, i)
style_table(ws4, 3, 28, 10, [6, 32, 10, 12, 20, 14, 40, 10, 11, 24])

# ---------- Sheet5 范围基线 ----------
ws5 = wb.create_sheet('范围基线')
ws5['A1'] = 'P0 · 对比范围基线（已锁定，执行期不可漂移）'
ws5['A1'].font = F_TITLE
ws5.merge_cells('A1:C1')
heads5 = ['基线项', '内容（已填写）', '确认来源与日期']
for c, h in enumerate(heads5, 1):
    ws5.cell(3, c, h)

baseline = [
    ['对比定位', '纯技术调研：产品功能层面差异，不服务于选型决策；不引入商业/采购权重',
     '《fw-comparison-workflow.html》P0 节，2026-08-15'],
    ['报告读者', '技术团队（网络工程/安全），用于 PA 替代需求的能力映射', '项目背景：PA替代需求列表.xlsx'],
    ['天融信 · 对比对象', 'NGFW 产品线功能全集：猎豹系列（如 NGFW4000-UF）、擎天系列（Ⅵ/III/X 超T级）、昆仑系列（信创）；虚拟化防火墙/工业防火墙形态仅登记适用性',
     'topsec.com.cn/products/TopNGFW.html，2026-08-15'],
    ['天融信 · 软件版本', 'OS 平台 NGTOS（64位多核并行）；手册基线《NGFW一本通》文档编号 V3.2406.37098（©2025）；OS 具体版本号官网未公开，建议 400-777-0777 核实',
     '本地 CHM《一本通》文档约定页 + 官网产品页，2026-08-15'],
    ['天融信 · 授权/模块', '全授权假设：应用识别特征库、病毒库、攻击规则库（IPS）、URL网址库、僵木蠕特征库、僵尸网络特征库、威胁情报等全部模块授权与升级服务生效',
     '官网产品页 + 代理商资料（猎豹系列标配许可项），2026-08-15'],
    ['Palo Alto · 对比对象', 'PA 在售产品线功能全集：PA-400/1400/3400/5400/7000 系列及 VM-Series；型号仅作功能适用标注',
     'docs.paloaltonetworks.com，2026-08-15'],
    ['Palo Alto · 软件版本', 'PAN-OS 12.1（2025-08 发布的最新大版本）；最新维护版 12.1.5（2026-03 发布）',
     'docs.paloaltonetworks.com/ngfw/release-notes/12-1，2026-07-09 更新，检索日 2026-08-15'],
    ['Palo Alto · 订阅清单', '全订阅假设：Threat Prevention、Advanced Threat Prevention、Advanced URL Filtering、WildFire/Advanced WildFire、DNS Security、SD-WAN、IoT Security 等在售安全订阅全部开启；依赖订阅的功能仍按"需订阅/授权"级如实标注',
     'PAN-OS 12.1 New Features 页订阅项清单，2026-08-15'],
    ['功能范围', 'A基础防火墙/B应用识别与控制/C威胁防护/D内容安全/E·VPN/F管理与运维/G硬件与平台规格/H合规与生态；功能能力+硬件平台规格，性能（吞吐/并发/新建/延迟）不对比',
     '《fw-comparison-workflow.html》3.2 示例框架（归纳后可扩充）'],
    ['需求映射', 'PA替代需求列表.xlsx 58 条三级需求逐条映射至功能点，天融信侧"当前状态"（已支持/部分交付/待交付/待评估/不支持）与手册双重取证',
     'PA替代需求列表.xlsx（港澳市场 PA 替代项目）'],
    ['场景关注', '港澳市场：Facebook/WhatsApp/TikTok/Threads/Instagram/LinkedIn/Snapchat/LINE/Teams/Twitch/EPIC/Gemini/OpenChat 等应用识别；英文界面/手册；Fortinet/PA/Hillstone 配置转换工具',
     'PA替代需求列表.xlsx 应用控制模块说明列'],
    ['数据截止日期', '2026-08-15（所有证据不晚于该日期，矩阵逐格标注数据日期）', '本基线'],
    ['实测处理', '本次不做真机实测（P5 移交）：文档无法确证项标"待验证"登记移交清单，附 T1~T8 定制用例', '执行计划第三节'],
    ['性能项处理', '不纳入对比；官方标称值仅登记于备注并注明"仅登记、不对比"', '《fw-comparison-workflow.html》质量红线'],
]
r = 4
for b in baseline:
    for c, v in enumerate(b, 1):
        ws5.cell(r, c, v)
    r += 1
style_table(ws5, 3, r - 1, 3, [18, 72, 34])
ws5.cell(4, 1).fill = FILL_KPI
ws5.cell(4, 2).font = Font(name='微软雅黑', size=10, bold=True, color='1F2937')

wb.save(OUT)
print('saved:', OUT)
print('需求行数:', req_no, '| 矩阵行数:', r - 4)
