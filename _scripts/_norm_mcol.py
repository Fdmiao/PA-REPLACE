# -*- coding: utf-8 -*-
import re, os
from openpyxl import load_workbook

base="/Users/个人资料/trx/2026/PA替代"
path=base+"/天融信vsPA功能对比.xlsx"
wb=load_workbook(path, data_only=True)
ws=wb["需求对比表"]
manual_files=set(os.listdir(base+"/_sources/topsec_manual"))
ADMIN_PAGES=1390; WEBHELP_PAGES=1104

def norm_id(i):
    i=i.strip().lower()
    if i.endswith('.html'):
        i=i[:-5]
    return i

def parse_ts(ts_raw):
    items=[]
    pos=0
    for m in re.finditer(r'([^\[\]+]+)\[([^\]]+)\]', ts_raw):
        title=m.group(1).strip()
        ids=m.group(2)
        for one in re.split(r'[/,]', ids):
            one=one.strip()
            if not one: continue
            if re.fullmatch(r'\d+(-\d+)?', one):
                items.append(f"{title}[p:{one}]")
            else:
                items.append(f"{title}[e:{norm_id(one)}]")
        pos=m.end()
    tail=ts_raw[pos:].strip().strip('+')
    # 前面未匹配的游离文字
    if ts_raw[:ts_raw.find('[') if '[' in ts_raw else len(ts_raw)].strip():
        pass
    if tail:
        items.append(tail)
    s="TS:"+" + ".join(items) if items else "TS:"+ts_raw.strip()
    return s

def parse_pa(pa_raw):
    segs=[]; notes=[]
    # 去前缀 PA:
    body=pa_raw
    def grab(typename):
        found=[]
        for m in re.finditer(re.escape(typename)+r'\s+([0-9pP/,\-\.]+)\s*([^，;；]*?)(?=[，,;；]|admin|webhelp|Web帮助|\+Datasheet|$|\))', body):
            pg=m.group(1); desc=m.group(2).strip()
            refs=[]
            for p in re.split(r'[/，,\.]', pg):
                p=p.strip().lstrip('pP')
                if p: refs.append(f"[p:{p}]")
            found.append(("".join(refs), desc))
        return found
    w=re.findall(r'webhelp\s+([0-9pP/,\-\.]+)\s*([^，;；]*?)(?=[，,;；]|admin|Web帮助|$|\))', body, re.S)
    a=re.findall(r'admin\s+([0-9pP/，,\.\-]+)\s*([^，;；]*?)(?=[，,;；]|webhelp|Web帮助|$|\))', body, re.S)
    web_refs=[r for r,_ in w]
    admin_refs=[r for r,_ in a]
    def fmt_refs(refs):
        out=[]
        for g in refs:
            for p in re.split(r'[/，,]', g.strip()):
                p=p.strip().lower().lstrip('p')
                if p: out.append(f"[p:{p}]")
        return "".join(out)
    if web_refs:
        segs.append("PA-WEBHELP:"+fmt_refs(web_refs))
    if admin_refs:
        segs.append("PA-ADMIN:"+fmt_refs(admin_refs))
    if 'Datasheet' in body or 'DataSheet' in body:
        segs.append("PA-DS:规格表")
    # 路径型说明(>)
    for pm in re.finditer(r'(?:Web帮助)>[^\[]+?(?=[，;；]|$)', body):
        notes.append(pm.group(0))
    # 未被 webhelp/admin 覆盖的零散页码(如 p839/p741 混排)
    remaining=re.sub(r'(?:webhelp|admin)\s+[0-9pP/，,\.\-]+','',body)
    remaining=re.sub(r'PA:|Datasheet\s*规格表','',remaining)
    remaining=remaining.replace('（否定证据：中文手册无WAF专章）','').replace('（仅账户级）','')
    remaining=remaining.strip().strip('，;；,')
    remaining=remaining.strip().rstrip('+')
    if remaining and remaining not in notes:
        notes.append(remaining)
    for n in notes:
        segs.append("PA-NOTE:"+n)
    return segs

def norm_ev(rid, ev):
    out=[]
    m=re.search(r'\[(R\d+)[→>][^\]]*#(\d+)\]', ev)
    req=f"[{m.group(1)}→M#{m.group(2)}]" if m else rid
    out.append(req)
    rest=re.sub(r'\[R\d+[→>][^\]]*\]\s*','',ev)
    notes=[]
    for pat in ['（否定证据：中文手册无WAF专章）','（仅账户级）','（型号信息缺失）','（插槽表述仅限PA-7000）']:
        if pat in rest:
            notes.append(pat.strip('（）'))
            rest=rest.replace(pat,'')
    mts=re.search(r'天融信CHM:(.*?)(?=(?:；|$))', rest)
    if mts:
        ts_raw=mts.group(1).strip()
        out.append(parse_ts(ts_raw))
        rest=rest[mts.end():]
    rest=rest.lstrip('；;，, ').strip()
    if rest:
        for s in parse_pa(rest):
            out.append(s)
    if notes:
        out.append("NOTE:"+";".join(notes))
    return " ; ".join(o for o in out if o)

results=[]
for r in range(4,62):
    rid=ws.cell(r,1).value or str(r-3)
    ev=ws.cell(r,13).value or ""
    newv=norm_ev(str(rid), str(ev))
    results.append((r,str(rid),newv))
    print(f"[{str(rid)}] {newv}")

print("\n===== 机器校验 =====")
fails=0
for r,rid,newv in results:
    for mm in re.finditer(r'\[e:([^\]]+)\]', newv):
        nm=norm_id(mm.group(1))
        if nm+'.html' not in manual_files:
            print(f"  X {rid}: 天融信条目不存在 {nm}")
            fails+=1
    for tag,maxp in (('PA-ADMIN',ADMIN_PAGES),('PA-WEBHELP',WEBHELP_PAGES)):
        seg=re.search(re.escape(tag)+r':(.*?)(?: ; |$)', newv)
        if seg:
            for pn in re.findall(r'\[p:(\d+)(?:-\d+)?\]', seg.group(1)):
                if int(pn)>maxp:
                    print(f"  X {rid}: 页码超范围 {tag} p{pn}")
                    fails+=1
print(f"问题总数: {fails}")

# ===== 写回工作簿 需求对比表 M 列 =====
from openpyxl import load_workbook as _lw
wbw=_lw(path)  # 保留公式/样式
wsw=wbw["需求对比表"]
written=0
for r,rid,newv in results:
    old=wsw.cell(r,13).value
    if old is None: 
        continue
    wsw.cell(r,13).value=newv
    written+=1
# 更新说明备注到单元格批注附近？不做。保存
out=base+"/天融信vsPA功能对比.xlsx"
wbw.save(out)
print(f"已写回 {written} 行到 {out} 的 M 列")

# 校验页数
import json
json.dump([{"row":r,"rid":rid,"new":nv} for r,rid,nv in results], open(base+"/_scripts/_normed.json","w"), ensure_ascii=False)
wb.close()