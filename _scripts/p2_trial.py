from openpyxl import load_workbook
from openpyxl.styles import Font

F = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求\对比工作底稿.xlsx'
TODAY = '2026-08-15'
wb = load_workbook(F)
ws = wb['功能对比矩阵']

def fill_row(no, tsup, tdesc, psup, pdesc, evid):
    for r in range(4, ws.max_row + 1):
        if str(ws.cell(r, 1).value) == str(no):
            ws.cell(r, 5, tsup)
            ws.cell(r, 6, tdesc)
            ws.cell(r, 7, psup)
            ws.cell(r, 8, pdesc)
            ws.cell(r, 9, '试填校验行')
            ws.cell(r, 10, evid)
            ws.cell(r, 11, TODAY)
            ws.cell(r, 12, 'P2试填')
            for c in range(1, 13):
                ws.cell(r, c).font = Font(name='微软雅黑', size=10, bold=True, color='1F2937')
            return r
    return None

r1 = fill_row(39, '完全支持', '入侵防御模块内置规则集（预定义+自定义规则+白名单），规则库经升级中心在线/离线更新；特征条数官方文档未公布，P3 从产品彩页/规格页核实',
              '需订阅授权', 'Threat Prevention 订阅；Vulnerability Protection（防漏洞）+Anti-Spyware（防间谍软件）两配置文件；特征库经 Dynamic Updates 更新',
              '天融信CHM:入侵防御[topic10/200-202]+升级中心[topic61]；PA:Web帮助>安全配置文件>Vulnerability/Anti-Spyware+Device>Dynamic Updates')
r2 = fill_row(67, '完全支持', 'IPSecVPN 章含 IKE 配置（静态隧道/手工隧道两形态），IKEv1/IKEv2 支持细节 P3 读正文确认',
              '完全支持', 'IKE Gateways 网关对象显式支持 IKEv1/IKEv2（IKEv2 优先/仅 IKEv2 模式），IKE Crypto 独立配置套件',
              '天融信CHM:IPSecVPN[ipsec.html/静态隧道topic4]；PA:Web帮助>Network Profiles>IKE Gateways/IKE Crypto')
r3 = fill_row(87, '完全支持', '高可用模块（状态/配置/链路探测/链路备份），双机热备主备模式；双活经虚系统+集群实现（案例四），主主原生形态 P3 确认',
              '完全支持', 'Device>High Availability：主动/被动与主动/主动两种模式均原生支持（配置同步+会话同步+路径监控）',
              '天融信CHM:高可用[toc462734581]+案例一/案例四；PA:Web帮助>Device>High Availability(主动/主动配置)')
wb.save(F)
print('trial rows filled:', r1, r2, r3)
