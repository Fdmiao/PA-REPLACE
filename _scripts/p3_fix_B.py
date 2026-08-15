from openpyxl import load_workbook
from openpyxl.styles import PatternFill

F = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求\对比工作底稿.xlsx'
TODAY = '2026-08-15'
FULL, PART, SUB, NO, TBD = '完全支持', '部分支持', '需订阅授权', '不支持', '待验证'
FILL_BY_SUP = {FULL: PatternFill('solid', fgColor='E2EFDA'), PART: PatternFill('solid', fgColor='FFF2CC'),
               SUB: PatternFill('solid', fgColor='E4DFEC'), NO: PatternFill('solid', fgColor='FBDBDB'),
               TBD: PatternFill('solid', fgColor='EDEDED')}

MISPLACED = [28, 30, 31, 32, 34, 35]

ROWS = [
(28, FULL, '高级威胁防护含“恶意加密流量检测”（不解密检测恶意加密流量：禁止/告警/放行）+DGA检测（DNS行为分析）；SNI/证书维度识别未见手册专项——待验证',
    PART, 'App-ID 不解密识别加密应用（QUIC/HTTP3；QUIC 专有加密建议默认阻止）；不解密配置文件（证书状态控制：阻止过期/不可信CA会话）；TLS1.3 加密证书影响自动排除',
    '待验证', '天融信CHM:高级威胁防护策略[topic203]；PA:admin p1020/p1006/p1037', 'R40 双方加密流量不解密识别深度均需实测'),
(30, FULL, '代理策略解密类型“服务器SSL解密”：客户端侧解密→安全引擎检测→服务器；SSL服务器模板；根证书导出导入浏览器/客户端',
    FULL, 'SSL Inbound Inspection（保护内部服务器，解密策略规则+服务器证书）；SSL入站检查选项卡（不受支持模式/版本阻断）；ECC/DSA 证书解密',
    '仅命名不同', '天融信CHM:代理策略[sslxiezaidailicelue]+代理模板[sslxiezaidailimuban]；PA:webhelp p165/p315/admin p1001', ''),
(31, FULL, 'PKI 模块四部分：本地证书/可信CA机构/证书请求文件/CA管理；内置本地CA签发；CA中心管理根证书/CRL/签发证书；CRL 支持 HTTP/LDAP 获取',
    FULL, 'Certificate Management：证书+证书配置文件+OCSP响应者；企业CA导入；OCSP优先+CRL回退吊销检查；防火墙可自作OCSP响应者；转发代理自动生成证书',
    '仅命名不同', '天融信CHM:PKI[ref483380791]+CA管理[ca3]+CA中心[ca2]+可信CA[ca]；PA:webhelp p784-794/admin p314/p338', ''),
(32, FULL, '各安全引擎动作模型统一（禁止/告警/放行）；DNS NX防御模式：限速/黑名单；WAF暴力破解客户端IP阻断（0-1440分钟，0=永久封禁）；DDoS源/目的限流+会话检查+DOS动态黑名单',
    FULL, '安全配置文件操作：alert/drop/reset-client/reset-server/reset-both（UDP删除连接）；AV最佳实践 reset-both；漏洞保护 Block IP（带阻断期限）',
    '仅命名不同', '天融信CHM:DDoS模板[ref464832260]+DNS防护[topic99]+WAF规则集[topic219]；PA:admin p1206-1207/webhelp p258', ''),
(34, FULL, '外部认证服务器：RADIUS/LDAP/TACACS；LDAP用户同步（自动/手动+查询账户）；门户认证/SSLVPN认证（本地+外部）；管理员认证策略（转发外部认证）',
    FULL, 'LDAP服务器配置文件+组映射（AD/Novell eDirectory/Sun ONE）；User-ID代理（Windows/PAN-OS集成两种）；LDAP 389/全局目录3268端口',
    '仅命名不同', '天融信CHM:认证服务器[toc462834526]+用户管理[toc490041708]+添加Ldap[ref467487846]；PA:webhelp p832/p880-883/admin p776/p816', ''),
(35, FULL, '认证服务器 RADIUS（PAP/CHAP/Unix登录）+LDAP+TACACS；Portal门户认证；SSLVPN认证；SAML 对接未见手册命中——待验证',
    FULL, 'SAML 2.0 IdP（SSO/SLO单点登出）+RADIUS+Kerberos；身份验证配置文件；GlobalProtect MFA（第一因素 Kerberos/SAML SSO）',
    '待验证', '天融信CHM:认证服务器[toc462834526]；PA:admin p268/p281/webhelp p740/p743', '天融信 SAML 对接需实测确认（R42）'),
(37, FULL, 'GRE/6to4/ISATAP/6in4 隧道终结；防代理（报文回环检测识别代理）；隐蔽信道检测（高级威胁防护）；隧道内流量经IPS检测',
    FULL, 'Tunnel Inspection 独立策略：检测 GRE/GTP-U/HTTP2 等明文隧道内层流量；App-ID 隧道内应用识别；max_encap 最大封装级数防护',
    '仅命名不同', '天融信CHM:防代理[topic76]+高级威胁防护策略[topic203]；PA:webhelp p172-174', ''),
(38, FULL, '访问控制规则引用应用+安全配置文件（策略与检测联动）；安全事件/攻击者/受害者三视角+攻击链分析；威胁统计（14类引擎）；日志查看',
    FULL, 'ACC 应用+威胁联动视图；Log Forwarding 日志转发配置文件（SIEM/sniffer/syslog）；统一日志；自动关联引擎跨日志关联',
    '仅命名不同', '天融信CHM:安全事件[topic79]+攻击者[topic55]+威胁统计[toc450050590]；PA:webhelp p305+ACC', ''),
]

wb = load_workbook(F)
ws = wb['功能对比矩阵']
pos = {}
for r in range(4, ws.max_row + 1):
    v = ws.cell(r, 1).value
    if v is not None:
        pos[str(v)] = r

for no in MISPLACED:
    r = pos.get(str(no))
    if not r:
        continue
    for c in range(5, 13):
        ws.cell(r, c).value = None
        ws.cell(r, c).fill = PatternFill()

fixed = 0
for no, ts, td, ps, pd, diff, evid, note in ROWS:
    r = pos.get(str(no))
    if not r:
        print('MISS', no)
        continue
    ws.cell(r, 5, ts); ws.cell(r, 7, ps)
    ws.cell(r, 6, td); ws.cell(r, 8, pd)
    ws.cell(r, 9, diff); ws.cell(r, 10, evid)
    ws.cell(r, 11, TODAY); ws.cell(r, 12, note or 'P3填充(修正错位)')
    ws.cell(r, 5).fill = FILL_BY_SUP[ts]
    ws.cell(r, 7).fill = FILL_BY_SUP[ps]
    fixed += 1
wb.save(F)
print('B fixed rows:', fixed)
