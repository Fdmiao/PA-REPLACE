from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

F = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求\对比工作底稿.xlsx'
TODAY = '2026-08-15'
FULL, PART, SUB, NO, TBD = '完全支持', '部分支持', '需订阅授权', '不支持', '待验证'
GREEN = PatternFill('solid', fgColor='E2EFDA')
YELLOW = PatternFill('solid', fgColor='FFF2CC')
PURPLE = PatternFill('solid', fgColor='E4DFEC')
RED = PatternFill('solid', fgColor='FBDBDB')
GRAY = PatternFill('solid', fgColor='EDEDED')
FILL_BY_SUP = {FULL: GREEN, PART: YELLOW, SUB: PURPLE, NO: RED, TBD: GRAY}

# (no, tsup, tdesc, psup, pdesc, diff, evid, note)
A = [
(1, FULL, '工作原理章：五元组会话表查询+DDoS预检；TCP连接完整性状态检测（三次握手跟踪）[wangluocanshu]',
    FULL, '会话老化时间/会话所有者/会话分发等全局会话机制（Device>Setup>Session）',
    '仅命名不同', '天融信CHM:工作原理[toc183414233]+连接参数[wangluocanshu]；PA:Web帮助 p691 会话设置', ''),
(2, FULL, '应用类型24大类（父类/子类）；预定义应用不可改+自定义应用+应用组；应用作为策略过滤条件',
    FULL, 'App-ID 识别引擎（应用+容器+依赖应用+子应用分级）；应用程序过滤器/组；ACE 云引擎未知应用检测',
    '实现路径不同', '天融信CHM:应用[toc448912175]+预定义应用[topic194]；PA:admin p920-940 App-ID', ''),
(3, FULL, '认证策略+门户认证/SSLVPN认证（本地+外部）；防暴力破解/登录验证码/账号锁定',
    FULL, 'User-ID：AD探测/Syslog/XML API/终端服务器代理多源映射；组映射+Cloud Identity Engine',
    '实现路径不同', '天融信CHM:认证策略[topic210]+管理用户；PA:webhelp p866 User-ID', '天融信侧重接入认证，PA侧重透明用户映射'),
(4, FULL, '一体化安全策略（五元组+应用+用户一次匹配）；包过滤+UTM两级策略模型',
    FULL, 'Single-Pass 单遍架构：一次解析完成 App-ID/Content-ID/User-ID 所有检测（PA 标志性架构）',
    '仅命名不同', '天融信CHM:安全策略[topicoperatoranquancelue]；PA:Web帮助 p21 功能与优点', ''),
(5, FULL, '访问控制策略数量受 license 控制[toc450050600]；具体容量数值官方文档未公开',
    FULL, '策略/对象容量因型号而异（见各型号 Datasheet）；地址组嵌套等容量项在 datasheet 规格表',
    '待验证', '天融信CHM:配置访问控制策略；PA:Datasheet 规格表', '容量数值双方均需实测/询证，登记待验证'),
(6, FULL, 'SNAT 地址池/EIM转换(P2P)/Hairpin回流/仅源IP转换等',
    FULL, 'DIPP 动态IP+端口（每IP约64000会话）/静态IP/持久NAT/接口地址；过度订阅',
    '仅命名不同', '天融信CHM:源NAT[nat1111]+EIM[nateimzhuanhuan]；PA:webhelp p147-149', ''),
(7, FULL, '服务器映射(DNAT)：静态映射+服务器负载均衡两种模式；NAT46/NAT66 全系列',
    FULL, '目标NAT：静态IP/动态IP(会话分发)/端口转换；DNS Rewrite；FQDN 目标支持',
    '仅命名不同', '天融信CHM:服务器映射[toc398625673]；PA:webhelp p148-149', ''),
(8, FULL, '一对一转换：源/目的地址池数量相同且一一映射',
    FULL, '静态IP转换+双向转换（Bi-directional）',
    '仅命名不同', '天融信CHM:一对一转换[yiduiyinat]；PA:webhelp p148', ''),
(9, FULL, '地址转换独立策略库（安全策略>地址转换），与访问控制策略分离配置',
    FULL, 'NAT 策略库独立于安全策略库（Policies>NAT），先 NAT 后安全策略匹配',
    '仅命名不同', '天融信CHM:地址转换[toc450050602]；PA:webhelp p145 NAT策略', ''),
(10, FULL, '策略路由：按源/目的地址端口、ISP名称、协议、角色等选路；路由优先级：回环>黑洞>直连>策略路由>ISP路由>静态>动态',
    FULL, 'PBF 基于策略的转发：源区域/地址/用户/应用/服务+下一跳；支持路径监控与失效切换；BFD',
    '实现路径不同', '天融信CHM:策略路由[toc490041605]+路由[ref487714854]；PA:webhelp p156 PBF', ''),
(11, FULL, 'OSPFv2/OSPFv3；区域/Router-ID/路由引入；与BGP/RIP/静态互引',
    FULL, 'OSPFv2/v3：区域类型(Stub/NSSA/TS）、验证配置文件、BFD、RFC1583兼容',
    '仅命名不同', '天融信CHM:OSPF[ospf.html]+OSPFv3；PA:webhelp p445-451', ''),
(12, FULL, 'BGP：邻居/Router-ID/路由引入；动态路由服务管理（限制BGP邻居防攻击）',
    FULL, 'BGP：对端组/导入导出策略/条件通告/路由反射器/最大前缀/平稳重启',
    '部分差异', '天融信CHM:BGP[bgp.html]；PA:webhelp p456-462', 'PA 侧 BGP 特性粒度更细（对端组/条件通告）'),
(13, PART, '无独立多VRF对象；“虚拟路由”为IPSec隧道内多保护子网路由[topic67]；多路由实例依赖虚系统实现',
    FULL, '虚拟路由器(VR)/逻辑路由器多实例：每VR独立路由表+RIB+动态路由；vsys间路由泄漏',
    '实现路径不同', '天融信CHM:虚拟路由[topic67]+虚系统；PA:webhelp p436 虚拟路由器', '重要差异：PA原生多VRF，天融信经虚系统'),
(14, FULL, '接口五种工作模式：路由/交换/listening/虚拟线/聚合（默认路由模式）',
    FULL, '第3层接口：静态/DHCP/PPPoE 寻址，IPv4+IPv6 双栈',
    '仅命名不同', '天融信CHM:配置接口基本信息[toc490041592]；PA:webhelp p352', ''),
(15, FULL, '透明模式：交换模式接口+MAC表；案例二透明模式双机链路保护部署',
    FULL, '虚拟线路(Virtual Wire)：两接口绑定透传；vwire子接口VLAN标记；支持解密',
    '实现路径不同', '天融信CHM:工作模式[toc100644454]；PA:webhelp p435 虚拟线路', 'PA叫Virtual Wire，天融信叫透明/交换模式'),
(16, FULL, '混合模式：透明+路由同机工作（三种工作模式之一）',
    PART, '无“混合模式”专门配置项；靠L2/Vwire/L3接口同机组合实现等效部署',
    '实现路径不同', '天融信CHM:工作模式[toc100644454]；PA:接口类型组合（webhelp p332-406）', ''),
(17, FULL, 'IPv4/IPv6双栈防护：双栈访问控制/应用识别/AV/IPS/抗D等全功能；NAT64/NAT46/NAT66/DNS64/RADVD/DHCPv6',
    FULL, 'IPv6 全面支持：接口/路由(OSPFv3/BGP)/安全策略/区域保护IPv6丢弃选项/NAT66',
    '仅命名不同', '天融信CHM:功能和特点[toc462736546]；PA:webhelp p673/p603', ''),
(18, FULL, 'NAT64：静态转换+前缀转换+动态算法（IVI/普通前缀）；DNS64 关联客户端与前缀；NAT46 双向',
    PART, 'NAT64 有支持（会话设置NAT64 MTU、地址对象NAT64用途、NetFlow NAT64字段/RFC6146）；DNS64 配置未见专章——待验证',
    '天融信优势', '天融信CHM:NAT64[anquancelue_nat_nat64]+DNS64；PA:webhelp p693/admin p1233', ''),
(19, FULL, '虚系统：路由/非路由模式；虚接口+虚拟链路互联；虚系统管理员（无授权能力）；对象共享/独享',
    FULL, 'vsys：独立策略/接口/管理员；vsys间User-ID数据共享；外部区域/共享网关；高型号需vsyss许可证',
    '实现路径不同', '天融信CHM:虚拟系统[xunixitong/topic128]；PA:webhelp p780-781/admin p1312', '数量双方均按型号/许可，登记待验证'),
(20, FULL, 'TAG子接口（802.1Q）+MAC子接口+VLAN定义；QinQ（业务接入端dot1q/qinq匹配）',
    FULL, 'L2/L3/vwire 三类子接口；VLAN接口路由；802.1Q 标准支持',
    '仅命名不同', '天融信CHM:TAG子接口[toc490041596]+VLAN[vlan.html]；PA:webhelp p339-406', ''),
(21, FULL, '聚合接口（链路聚合组）+接口联动（2-8口）；MTU/MSS 独立设置',
    FULL, '聚合以太网 AE 接口组；HA 被动状态 LACP 预协商；MC-LAG（集群跨机箱）',
    '仅命名不同', '天融信CHM:聚合接口[toc490041599]；PA:webhelp p388+admin p463', ''),
(22, FULL, 'DHCP服务器（物理/VLAN/聚合口）+DHCP中继+DHCPv6服务器+地址池保留',
    FULL, 'DHCP 服务器/客户端/中继三角色；IoT DHCP日志提取（EAL）',
    '仅命名不同', '天融信CHM:DHCP[toc490041616-621]；PA:webhelp p556-560', ''),
(23, FULL, 'IPv4组播(PIM)+IPv6组播；组播转发表项；组播报文IPS检测（R38需求）',
    PART, 'IP多播：PIM-SM(ASM)/SSM+IGMP+静态RP/BSR/自动RP；MSDP；PIM-DM 未见——待验证',
    '天融信优势(待验证)', '天融信CHM:IPv4组播[ipv4.html]；PA:webhelp p469-472 多播', 'PIM-DM 支持性双方实测确认'),
(112, FULL, 'BFD：全局配置（缺省组播地址/震荡抑制）+BFD会话；与静态路由联动',
    FULL, 'BFD Profile 应用于静态路由/OSPF/BGP；PA-400系列起支持',
    '仅命名不同', '天融信CHM:BFD全局配置[topic218]；PA:webhelp p439/p445', ''),
(113, TBD, 'CHM 手册未见 ECMP 专章；可能在 OSPF 等路由配置内——待读正文确认',
    FULL, 'ECMP 等价多路径（虚拟路由器）；主动/主动 HA 模式下 ECMP 支持',
    '待验证', 'PA:admin p397 ECMP', ''),
(114, FULL, 'DNS服务器/DNS代理/域名记录/例外域名/DNS Doctoring（DNS改写）',
    FULL, 'DNS Proxy：映射规则/静态条目/TCP代理；DNS缓存；DNS重定向',
    '仅命名不同', '天融信CHM:网络管理>DNS；PA:webhelp p563 DNS代理', ''),
(115, FULL, 'VXLAN 二层网关+三层网关（VTEP封装解封装）；VXLAN MAC表（动/静态）；业务接入端',
    TBD, '11.1 Web帮助无 VXLAN VTEP 配置章节命中；PA 对 VXLAN 流量可识别（App-ID/Tunnel Inspection），隧道终结能力待验证',
    '天融信优势(待验证)', '天融信CHM:VXLAN[vxlan.html/topic253/254]', ''),
(116, NO, '手册未见 MACsec 相关内容',
    FULL, 'MACsec Profile（Network Profiles）：PA-400系列及以上，802.1AE 链路层加密',
    'PA优势', 'PA:Web帮助>Network Profiles>MACsec(11.1)', ''),
(117, FULL, 'PPPoE：路由模式物理接口拨号，最多8条线路',
    FULL, '第3层接口 PPPoE 客户端（接入集中器/被动模式）',
    '部分差异', '天融信CHM:PPPoE[pppoe.html]；PA:webhelp p355/p359', '天融信8线路并发，PA按接口'),
(118, FULL, 'ISATAP/6to4/6in4 三种 IPv6 过渡隧道专章（自动隧道/多点隧道）',
    NO, '双份手册均无 ISATAP/6to4/6in4 配置（否定证据：全文检索无命中）',
    '天融信优势', '天融信CHM:ISATAP[isatap]/6to4/6in4', ''),
(119, FULL, '服务器负载均衡：负载均衡策略+真实服务器+健康检查（资源管理>负载均衡）',
    NO, '无内置 SLB 功能（NAT+ECMP 可实现简单分发，但无服务器健康检查/会话保持等 SLB 特性）',
    '天融信优势', '天融信CHM:资源管理>负载均衡[toc490041635]', ''),
(120, FULL, '链路负载均衡：本地链路/链路池/出口策略（就近接入/权重调度）',
    PART, '无 LLB 功能模块；出向多链路选择经 SD-WAN 路径选择+PBF 实现（链路质量探测+应用感知选路）',
    '实现路径不同', '天融信CHM:链路负载均衡；PA:SD-WAN(webhelp p156+SD-WAN章节)', ''),
]

wb = load_workbook(F)
ws = wb['功能对比矩阵']
filled = 0
for no, ts, td, ps, pd, diff, evid, note in A:
    for r in range(4, ws.max_row + 1):
        if str(ws.cell(r, 1).value) == str(no):
            ws.cell(r, 5, ts); ws.cell(r, 7, ps)
            ws.cell(r, 6, td); ws.cell(r, 8, pd)
            ws.cell(r, 9, diff); ws.cell(r, 10, evid)
            ws.cell(r, 11, TODAY); ws.cell(r, 12, note or 'P3填充')
            ws.cell(r, 5).fill = FILL_BY_SUP[ts]
            ws.cell(r, 7).fill = FILL_BY_SUP[ps]
            filled += 1
            break
wb.save(F)
print('A category filled:', filled)
