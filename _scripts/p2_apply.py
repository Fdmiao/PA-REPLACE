from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求'
F = BASE + r'\对比工作底稿.xlsx'
TODAY = '2026-08-15'

F_BODY = Font(name='微软雅黑', size=10, color='1F2937')
F_HEAD = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
FILL_HEAD = PatternFill('solid', fgColor='1F4E79')
FILL_ZEBRA = PatternFill('solid', fgColor='F7F9FC')
FILL_NEW = PatternFill('solid', fgColor='FFF7E6')
THIN = Side(style='thin', color='D9DEE7')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

wb = load_workbook(F)

# ========== 1. 矩阵增补 28 功能点 ==========
ws = wb['功能对比矩阵']
additions = [
    # (大类, 模块, 功能点, 初步证据双侧来源)
    ('A', '路由', 'BFD 双向转发检测', '天融信CHM:BFD章[bfd.html]；PA:Network Profiles>BFD Profile'),
    ('A', '路由', 'ECMP 等价多路径', '天融信:待验证(CHM未见专章)；PA:虚拟路由器>ECMP'),
    ('A', '网络服务', 'DNS 代理与 DNS Doctoring', '天融信CHM:DNS服务器/域名记录/DNS Doctoring；PA:Network>DNS Proxy'),
    ('A', '二层网络', 'VXLAN（二层/三层网关）', '天融信CHM:VXLAN章[vxlan.html]+配置实例；PA:11.1待核实'),
    ('A', '网络服务', 'MACsec 链路加密', '天融信:未见；PA:Network Profiles>MACsec Profile(11.1)'),
    ('A', '网络服务', 'PPPoE 客户端接入', '天融信CHM:PPPoE章[pppoe.html]；PA:接口支持PPPoE'),
    ('A', 'IPv6', 'IPv6 过渡隧道（ISATAP/6to4/6in4）', '天融信CHM:隧道章三协议专节；PA:待核实'),
    ('A', '负载均衡', '服务器负载均衡 SLB', '天融信CHM:资源管理>负载均衡/服务器；PA:无内置(生态外置)'),
    ('A', '负载均衡', '链路负载均衡 LLB', '天融信CHM:链路负载均衡/链路池/本地链路；PA:无LLB(SD-WAN路径选择替代,实现路径不同)'),
    ('B', '识别技术', '移动网络协议保护（GTP/SCTP/MNP）', '天融信:无；PA:MNP/SCTP Protection Profile(电信场景)'),
    ('B', '应用控制', '网络数据包代理（NPBroker）', '天融信:无(端口镜像替代)；PA:Policies>Network Packet Broker(11.1)'),
    ('B', '识别技术', 'IoT 设备识别与资产清单', '天融信:资产发现(安全中心)；PA:IoT Security订阅+设备资产清单'),
    ('B', 'SSL 解密', '解密流量镜像与解密排除', '天融信:SSL卸载策略[R58]；PA:解密镜像接口+SSL解密排除(11.1)'),
    ('C', '威胁检测', '工控协议识别与防护', '天融信CHM:工业安全章(工控策略组/地址策略/值策略)；PA:无内置'),
    ('C', '威胁检测', '数据库访问防护', '天融信CHM:数据库安全章(防护/自学习/基线/风险语句)；PA:无内置'),
    ('C', '威胁检测', '弱口令检测', '天融信CHM:高级威胁防护>弱口令[topic205]；PA:无直接对应'),
    ('C', '威胁检测', '隧道检测（Tunnel Inspection）', '天融信:IPS兼做；PA:独立策略类型(11.1)'),
    ('E', '组网', 'SD-WAN（路径选择/链路质量探测/流量分发）', '天融信:无SD-WAN(LLB+策略路由部分实现)；PA:完整SD-WAN体系(策略/链路管理/接口)'),
    ('E', 'SSL VPN', '终端合规检查（HIP）', '天融信:无；PA:GlobalProtect HIP对象/配置文件/MDM'),
    ('E', 'IPSec VPN', '抗量子 VPN（PQ KEM）', '天融信:无；PA:PAN-OS 12.1新增(HQ KEM算术套件)'),
    ('F', '配置管理', '策略优化器与宽泛策略分析', '天融信CHM:策略分析及管理/宽泛策略分析；PA:Security Policy Optimizer+规则使用统计'),
    ('F', '密钥管理', 'HSM 硬件安全模块与主密钥', '天融信:未见；PA:Device>Setup>HSM+Master Key'),
    ('F', '运维诊断', '网络诊断与抓包', '天融信CHM:网络诊断/网络抓包；PA:Packet Capture+Ping/Traceroute'),
    ('F', '配置管理', '配置锁与并发编辑控制', '天融信:待验证(对应R25)；PA:Web界面锁定配置(11.1)'),
    ('G', '形态', '5G 蜂窝接口', '天融信:待核实；PA:PA-415-5G+蜂窝接口(11.1)'),
    ('G', '接口', 'PoE 供电输出', '天融信:未见；PA:PA-1400/PA-500 PoE(330W)'),
    ('G', '形态', '工业加固型号', '天融信:工业防火墙产品线；PA:PA-400R系列'),
    ('G', '扩展', '专用日志转发卡', '天融信:无此形态(集中管理平台替代)；PA:Log Forwarding Card(PA-5400/7000)'),
]

start_row = ws.max_row + 1
for i, (cat, mod, fp, src) in enumerate(additions):
    r = start_row + i
    vals = [112 + i, cat, mod, fp, '', '', '', '', '', src, TODAY, 'P2增补']
    for c, v in enumerate(vals, 1):
        cell = ws.cell(r, c, v)
        cell.font = F_BODY
        cell.alignment = WRAP
        cell.border = BORDER
        cell.fill = FILL_NEW
print('matrix rows after add:', ws.max_row - 3)

# ========== 2. 试填 3 点（#39 IPS / #67 IKE / #87 HA）==========
def fill_row(ws, no, tsup, tdesc, psup, pdesc, evid):
    for r in range(4, ws.max_row + 1):
        if ws.cell(r, 1).value == no:
            ws.cell(r, 5, tsup)
            ws.cell(r, 6, tdesc)
            ws.cell(r, 7, psup)
            ws.cell(r, 8, pdesc)
            ws.cell(r, 9, '试填校验行')
            ws.cell(r, 10, evid)
            ws.cell(r, 11, TODAY)
            ws.cell(r, 12, 'P2试填')
            for c in range(1, 13):
                ws.cell(r, c).font = Font(name='微软雅黑', size=10, bold=True, color='1F2937')
            return r
    return None

r1 = fill_row(ws, 39, '完全支持', '入侵防御模块内置规则集（预定义+自定义规则+白名单），规则库经升级中心在线/离线更新；特征条数官方文档未公布，P3 从产品彩页/规格页核实',
              '需订阅授权', 'Threat Prevention 订阅；Vulnerability Protection 配置文件（防漏洞）+Anti-Spyware（防间谍软件）两模块；特征库经 Dynamic Updates 更新',
              '天融信CHM:入侵防御[topic10/200-202]+升级中心[topic61]；PA:Web帮助>安全配置文件>Vulnerability/Anti-Spyware+Device>Dynamic Updates')
r2 = fill_row(ws, 67, '完全支持', 'IPSecVPN 章含 IKE 配置（静态隧道/手工隧道两形态），IKEv1/IKEv2 支持细节 P3 读正文确认',
              '完全支持', 'IKE Gateways 网关对象显式支持 IKEv1/IKEv2（IKEv2 优先/仅 IKEv2 模式），IKE Crypto 独立配置套件',
              '天融信CHM:IPSecVPN[ipsec.html/静态隧道topic4]；PA:Web帮助>Network Profiles>IKE Gateways/IKE Crypto')
r3 = fill_row(ws, 87, '完全支持', '高可用模块（状态/配置/链路探测/链路备份），双机热备主备模式；双活经虚系统+集群实现（案例四），主主原生形态 P3 确认',
              '完全支持', 'Device>High Availability：主动/被动与主动/主动两种模式均原生支持（配置同步+会话同步+路径监控）',
              '天融信CHM:高可用[toc462734581]+案例一/案例四；PA:Web帮助>Device>High Availability(主动/主动配置)')
print('trial rows:', r1, r2, r3)

# ========== 3. 需求映射 Sheet ==========
ws6 = wb.create_sheet('需求映射')
ws6['A1'] = 'PA替代需求 58 条 → 功能对比矩阵功能点映射（P2 产出）'
ws6['A1'].font = Font(name='微软雅黑', size=12, bold=True, color='1F2937')
ws6.merge_cells('A1:F1')
heads = ['需求号', '一级/二级/三级需求', '映射功能点(矩阵序号)', '映射依据', '状态', '备注']
for c, h in enumerate(heads, 1):
    ws6.cell(3, c, h)

req_map = [
    (1, '完全支持映射 F84', '邮件告警(SMTPS)属日志外发/告警子项'),
    (2, 'F83', 'FireMon 经 REST/XML API 对接配置管理'),
    (3, 'F83', 'Ansible 经 API+Playbook 对接'),
    (4, 'F83', 'RESTful API 覆盖对象/策略/NAT/路由/接口'),
    (5, 'F84', '日志过滤字段属日志能力'),
    (6, 'F84', '同上'),
    (7, 'F84', '同上'),
    (8, 'F84', '同上'),
    (9, 'F84', '连接日志字段（会话开始时间）'),
    (10, 'F84', '连接日志字段（会话结束时间）'),
    (11, 'F84', '连接日志字段（结束原因）'),
    (12, 'F91', '授权平滑升级属授权管理'),
    (13, 'B24+B36', '港澳国际应用识别：识别库规模+覆盖度（R13 Facebook）'),
    (14, 'B24+B36', 'YouTube'),
    (15, 'B24+B36', 'Twitter(X)'),
    (16, 'B24+B36', 'WhatsApp'),
    (17, 'B24+B36', 'TikTok'),
    (18, 'B24+B36', 'Zoom'),
    (19, 'B24+B36', 'Steam'),
    (20, 'B24+B36', 'Office365（子应用识别 B27）'),
    (21, 'B24+B36', 'TeamViewer'),
    (22, 'E74', 'SSL VPN 双因子认证（密码+证书）'),
    (23, 'A5', '地址组嵌套 8 层属对象容量'),
    (24, 'F92', '代理服务器升级规则库属升级方式'),
    (25, 'F135', '配置锁（P2 增补功能点）'),
    (26, 'A10+F80', '带外管理路由 syslog（管理平面策略路由）'),
    (27, 'A10+F80', '带外管理路由认证'),
    (28, 'F90', 'Fortinet 配置转换工具（配置迁移能力）'),
    (29, 'F90', 'PA 配置转换工具'),
    (30, 'F90', 'Hillstone 配置转换工具'),
    (31, 'B33', 'RED 拥塞处理属 QoS'),
    (32, 'B33', '尾丢弃同上'),
    (33, 'A23', 'IGMP 属组播路由'),
    (34, 'A23', 'MLD（IPv6 组播）'),
    (35, 'A23', 'PIM-SM'),
    (36, 'A23', 'PIM-DM'),
    (37, 'A2+A5', '策略源/目的域名（地址对象域名支持）'),
    (38, 'C39+A23', '组播报文 IPS 检测'),
    (39, 'D57', '域名分类访问控制（URL 分类）'),
    (40, 'C53+B28', 'HTTP3/QUIC WAF（QUIC 识别+解密）'),
    (41, 'C53+F84', 'XFF 原始 IP 解析上报'),
    (42, 'C46', '云沙箱联动'),
    (43, 'F80', 'Web 界面英文模式'),
    (44, 'B24+F92', '规则库英文（应用/IPS/AV/IP归属地）'),
    (45, 'F80', '英文版用户手册'),
    (46, 'F84', '日志英文'),
    (47, 'F86', '报表英文'),
    (48, 'G95/G97', '网络层硬件加速（ASIC）——仅能力有无对比，不做性能数值对比'),
    (49, 'G97', 'IPS 内容加速卡'),
    (50, 'G97', 'DLP 内容检查加速卡'),
    (51, 'G97', 'IPSec 硬件加速卡'),
    (52, 'F87+F88', 'HA 配置同步（整机/策略）'),
    (53, 'A19+F87', '虚系统 HA（双活虚系统互备）'),
    (54, 'A5', 'NAT 策略组容量 512 条'),
    (55, 'G96', '25G/40G/2.5G/5G/10G 接口支持'),
    (56, 'G96', '接口光衰信息查看'),
    (57, 'E77+A10', 'IPSec 链路自动切换（DPD+策略路由）'),
    (58, 'B124', '解密端口镜像（P2 增补功能点）'),
]

ws1 = wb['需求对比表']
for i, (reqno, fp, basis) in enumerate(req_map):
    r = 4 + i
    l1, l2, l3 = [ws1.cell(r, c).value for c in (1, 2, 3)]
    ws6.cell(r - 1, 1, f'R{reqno}')
    ws6.cell(r - 1, 2, f'{l1} / {l2} / {l3}')
    ws6.cell(r - 1, 3, fp)
    ws6.cell(r - 1, 4, basis)
    ws6.cell(r - 1, 5, '待P3填充')

for rr in range(3, 3 + len(req_map)):
    for c in range(1, 7):
        cell = ws6.cell(rr, c)
        cell.font = F_BODY if rr > 3 else F_HEAD
        cell.alignment = WRAP
        cell.border = BORDER
        if rr > 3 and (rr - 4) % 2 == 1:
            cell.fill = FILL_ZEBRA
for c in range(1, 7):
    ws6.cell(3, c).fill = FILL_HEAD
    ws6.cell(3, c).alignment = CENTER
for c, w in zip(range(1, 7), [8, 44, 16, 50, 12, 20]):
    ws6.column_dimensions[chr(64 + c)].width = w
ws6.freeze_panes = 'A4'

wb.save(F)
print('saved. sheets:', wb.sheetnames)
