from openpyxl import load_workbook
from collections import defaultdict, Counter
import json

BASE = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求'
wb = load_workbook(BASE + r'\天融信vsPA功能对比.xlsx')
mx = wb['功能对比矩阵']

CAT = {'A': '基础网络与部署', 'B': '应用识别与控制', 'C': '威胁防护', 'D': '内容安全',
       'E': 'VPN与加密', 'F': '管理与运维', 'G': '硬件与平台', 'H': '合规与生态'}
data = defaultdict(lambda: {'n': 0, 'ts': Counter(), 'pa': Counter(), 'diff': Counter()})
for r in range(4, mx.max_row + 1):
    no = mx.cell(r, 1).value
    if no is None or not str(no).strip().isdigit():
        continue
    c = str(mx.cell(r, 2).value or '')[:1]
    data[c]['n'] += 1
    data[c]['ts'][str(mx.cell(r, 5).value or '')] += 1
    data[c]['pa'][str(mx.cell(r, 7).value or '')] += 1
    data[c]['diff'][str(mx.cell(r, 9).value or '')] += 1

out = {'cats': {}, 'ts_total': {}, 'pa_total': {}, 'diff_total': {}}
for c in 'ABCDEFGH':
    d = data[c]
    out['cats'][c] = {'name': CAT[c], 'n': d['n'],
                      'ts': dict(d['ts']), 'pa': dict(d['pa']),
                      'ts_full_rate': round(d['ts'].get('完全支持', 0) / d['n'] * 100, 1),
                      'pa_full_rate': round(d['pa'].get('完全支持', 0) / d['n'] * 100, 1)}
    for k, v in d['ts'].items():
        out['ts_total'][k] = out['ts_total'].get(k, 0) + v
    for k, v in d['pa'].items():
        out['pa_total'][k] = out['pa_total'].get(k, 0) + v
    for k, v in d['diff'].items():
        out['diff_total'][k] = out['diff_total'].get(k, 0) + v

with open(BASE + r'\_scripts\p4_stats.json', 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=1)
print(json.dumps(out, ensure_ascii=False, indent=1))
