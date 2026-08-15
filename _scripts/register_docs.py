from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求'
F = BASE + r'\天融信vsPA功能对比.xlsx'
TODAY = '2026-08-15'

F_BODY = Font(name='微软雅黑', size=10, color='1F2937')
FILL_ZEBRA = PatternFill('solid', fgColor='F7F9FC')
THIN = Side(style='thin', color='D9DEE7')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top')

wb = load_workbook(F)
ws = wb['资料登记表']

docs = [
    [6, 'PAN-OS 12.1 官方文档门户（英文，特性索引）', 'Palo Alto', 'L2 手册/文档门户', '12.1', '持续更新',
     'https://docs.paloaltonetworks.com/pan-os', '全大类', TODAY, 'P3 按功能点逐页取证'],
    [7, 'PAN-OS 管理员指南（简体中文版 PDF）', 'Palo Alto', 'L2 手册', '11.0（中文版；11.1+ 在线版）', '—',
     'https://docs.paloaltonetworks.com/content/dam/techdocs/zh_CN/pdf/pan-os/11-0/pan-os-admin-11-0-zh-cn.pdf；中文索引 docs.paloaltonetworks.com/translated/zh-cn', '全大类', TODAY,
     '中文版为 11.x；正式证据以 12.1 英文文档为准，版本差异在矩阵中标注'],
    [8, 'PA NGFW 硬件产品页（产品线全集 PA-400R~7500）', 'Palo Alto', 'L1 产品规格', '在售全集', '2026-08-15 检索',
     'https://www.paloaltonetworks.com/network-security/next-generation-firewall-hardware', 'G', TODAY, '—'],
    [9, 'Single-Pass Parallel Processing 架构白皮书（入口页）', 'Palo Alto', 'L1 原理', '—', '—',
     'https://www.paloaltonetworks.com/resources/whitepapers/single-pass-parallel-processing-architecture', 'A/C/F', TODAY, 'P3 需引用时再取全文'],
    [10, 'PA Certifications 官方页（CC/FIPS 汇总）', 'Palo Alto', 'L3 合规', '—', '2026-08-15 检索',
     'https://docs.paloaltonetworks.com/ngfw/administration/certifications', 'H', TODAY, '—'],
    [11, 'Common Criteria Validation Report（NDcPP v3.0e + IPS/ST/VPN PP）', '第三方权威（Leidos）', 'L3 合规', 'PP-Configuration v2.0（2024-04-25）', '—',
     'https://www.commoncriteriaportal.org/files/epfiles/st_vid11482-vr.pdf（覆盖 PA-400~7000 及 VM 系列）', 'H', TODAY, '—'],
    [12, 'FIPS 140-3 Level 2 证书（PAN-OS 11.1/11.2 @ PA-400~7500）', '第三方权威（NIST CMVP）', 'L3 合规', 'FIPS 140-3 L2', '2026-06-10 验证（sunset 2031-09-06）',
     'https://sec-certs.org/fips/2e64f1bab2a90810/', 'H', TODAY, '—'],
    [13, '天融信《网络安全专用产品安全检测证书》新闻（网专检测，三类防火墙）', '天融信', 'L3 合规', '第一批证书（万兆/千兆/百兆全覆盖）', '—',
     'https://www.topsec.com.cn/newsx/4441.html', 'H', TODAY, '等保/销售许可体系佐证'],
    [14, '天融信 IPv6 Ready Logo / 互联互通检测证书（新闻页）', '天融信', 'L3 合规', '—', '—',
     'https://www.topsec.com.cn/newsx/5604（IPv6 Ready）；/newsx/5437（首批互联互通）', 'A/H', TODAY, '—'],
]

start = 4 + 5
for i, d in enumerate(docs):
    r = start + i
    ws.cell(r, 1, d[0])
    for c, v in enumerate(d, 1):
        cell = ws.cell(r, c, v)
        cell.font = F_BODY
        cell.alignment = WRAP
        cell.border = BORDER
        if i % 2 == 1:
            cell.fill = FILL_ZEBRA

wb.save(F)
print('registered docs 6-14')
