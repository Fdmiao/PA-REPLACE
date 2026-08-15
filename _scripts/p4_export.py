import shutil
from openpyxl import load_workbook

BASE = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求'
SRC = BASE + r'\对比工作底稿.xlsx'
DST = BASE + r'\天融信vsPA功能对比.xlsx'

shutil.copyfile(SRC, DST)
wb = load_workbook(DST)

# 标题行标注最终版
ws = wb['需求对比表']
ws.cell(1, 1).value = '天融信 NGFW × Palo Alto NGFW 功能对比 · 最终交付 v1（2026-08-15，港澳 PA 替代需求 58 条 × 功能矩阵 139 点）'
ws.cell(2, 1).value = '范围：产品功能层面；数据截止 2026-08-15；证据=文档名+章节/页码；"待验证"项未实测不得视为已支持'

# sheet 顺序核对（计划：需求对比/矩阵/待验证/资料登记/范围基线，+需求映射审计）
order = ['需求对比表', '功能对比矩阵', '待验证移交清单', '资料登记表', '范围基线', '需求映射']
assert wb.sheetnames == order, wb.sheetnames
wb.active = 0
wb.save(DST)
print('最终交付已生成:', DST)
print('sheets:', order)
