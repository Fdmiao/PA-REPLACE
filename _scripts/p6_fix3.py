import re
from openpyxl import load_workbook

BASE = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求'
# V清单: #编号 → V编号
wb0 = load_workbook(BASE + r'\天融信vsPA功能对比.xlsx')
wv = wb0['待验证移交清单']
vid = {}
for r in range(4, wv.max_row + 1):
    if wv.cell(r, 2).value:
        v = str(wv.cell(r, 1).value)
        for m in re.findall(r'#(\d+)', str(wv.cell(r, 2).value)):
            vid.setdefault(m, v)
print('映射:', {k: vid[k] for k in ('60', '107', '110')})

for F in [BASE + r'\天融信vsPA功能对比.xlsx', BASE + r'\对比工作底稿.xlsx']:
    wb = load_workbook(F)
    mx = wb['功能对比矩阵']
    fixed = []
    for r in range(4, mx.max_row + 1):
        no = str(mx.cell(r, 1).value or '')
        if no in ('60', '107', '110'):
            cur = str(mx.cell(r, 12).value or '')
            v = vid.get(no)
            if v and f'（{v}）' not in cur:
                mx.cell(r, 12).value = f'{cur}（{v}）'
                fixed.append(f'#{no}+{v}')
    try:
        wb.save(F)
        print(F.split('\\')[-1], '→', fixed)
    except PermissionError:
        print(F.split('\\')[-1], '→', fixed, '[保存失败: 文件被占用]')
