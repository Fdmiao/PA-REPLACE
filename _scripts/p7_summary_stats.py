# -*- coding: utf-8 -*-
# P7 交付物汇总报告数据提取：从主交付 Excel 提取关键统计
import os
from collections import Counter
from openpyxl import load_workbook

BASE = r"e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求"
XLSX = os.path.join(BASE, "天融信vsPA功能对比.xlsx")

wb = load_workbook(XLSX, data_only=True)
print("== Sheets ==", wb.sheetnames)

# Sheet1 需求对比表
rq = wb['需求对比表']
labels1 = [str(rq.cell(3, c).value or '') for c in range(1, 15)]
print("\n== Sheet1 表头 ==")
for i, l in enumerate(labels1, 1):
    print(f"  col{i}: {l}")

status = Counter()
support = Counter()
diff_conc = []
for r in range(4, 62):
    status[str(rq.cell(r, 6).value or '').split('（')[0].strip()] += 1
    support[str(rq.cell(r, 8).value or '').strip()] += 1
    d = str(rq.cell(r, 12).value or '').strip()
    if d:
        diff_conc.append((f"R{r-3}", d))
print("\n== 需求状态分布 ==", dict(status))
print("== Sheet1 天融信支持度(简化计数) ==")
sup2 = Counter()
for r in range(4, 62):
    v = str(rq.cell(r, 8).value or '').strip()
    key = '完全支持' if v.startswith('完全支持') else ('部分支持' if v.startswith('部分支持') else ('不支持' if v.startswith('不支持') else ('待验证' if '待验证' in v else ('N/A' if not v else '其他'))))
    sup2[key] += 1
print(dict(sup2))
print("== 差异结论非空条数 ==", len(diff_conc))
for rid, d in diff_conc:
    if '缺口' in d:
        print(f"  {rid}: {d[:60]}")

# Sheet2 功能对比矩阵
mx = wb['功能对比矩阵']
cats = Counter()
ts = Counter()
pa = Counter()
diff = Counter()
prio = Counter()
for r in range(4, mx.max_row + 1):
    if not mx.cell(r, 2).value:
        continue
    cats[str(mx.cell(r, 3).value or '')] += 1
    ts[str(mx.cell(r, 5).value or '')] += 1
    pa[str(mx.cell(r, 6).value or '')] += 1
    diff[str(mx.cell(r, 7).value or '')] += 1
    prio[str(mx.cell(r, 8).value or '')] += 1
print("\n== 矩阵大类分布 ==")
for k, v in cats.items():
    print(f"  {k}: {v}")
print("== TS 支持度 ==", dict(ts))
print("== PA 支持度 ==", dict(pa))
print("== 差异定性 ==", dict(diff))
print("== 优先级 ==", dict(prio))

# Sheet3 待验证清单
wv = wb['待验证移交清单']
vp = Counter()
vt = Counter()
hi = []
for r in range(4, wv.max_row + 1):
    if not wv.cell(r, 2).value:
        continue
    p = str(wv.cell(r, 6).value or '')
    t = str(wv.cell(r, 5).value or '')
    vp[p] += 1
    vt[t.split('（')[0].strip()] += 1
    if '高' in p:
        hi.append((str(wv.cell(r, 1).value), str(wv.cell(r, 4).value or '')[:40], t[:30]))
print("\n== V清单优先级 ==", dict(vp))
print("== V清单验证方式 ==", dict(vt))
print("== 高优 14 项 ==")
for v in hi:
    print(f"  {v[0]}: {v[1]} | {v[2]}")

# Sheet4 资料登记
src = wb['资料登记表']
n_src = sum(1 for r in range(4, src.max_row + 1) if src.cell(r, 2).value)
print("\n== 资料登记份数 ==", n_src)

# Sheet5/6
try:
    rb = wb['范围基线']
    print("== 范围基线行数 ==", sum(1 for r in range(1, rb.max_row + 1) if any(rb.cell(r, c).value for c in range(1, 6))))
except Exception as e:
    print("范围基线:", e)
try:
    rm = wb['需求映射']
    n_map = sum(1 for r in range(4, rm.max_row + 1) if rm.cell(r, 2).value)
    print("== 需求映射条数 ==", n_map)
except Exception as e:
    print("需求映射:", e)

# 交付物文件大小
print("\n== 交付物文件 ==")
for name in ["天融信vsPA功能对比.xlsx", "对比工作底稿.xlsx", "对比执行计划.md",
             r"fw-comparison-report\fw-comparison-report.html"]:
    p = os.path.join(BASE, name)
    if os.path.exists(p):
        kb = os.path.getsize(p) / 1024
        print(f"  {name}: {kb:.0f} KB")
    else:
        print(f"  {name}: MISSING")

for d in ["_sources", "_scripts"]:
    p = os.path.join(BASE, d)
    if os.path.isdir(p):
        n = sum(len(fs) for _, _, fs in os.walk(p))
        sz = sum(os.path.getsize(os.path.join(dp, f)) for dp, _, fs in os.walk(p) for f in fs) / 1024 / 1024
        print(f"  {d}\\: {n} 个文件, {sz:.1f} MB")
