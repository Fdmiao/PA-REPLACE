from openpyxl import load_workbook

BASE = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求'
F = BASE + r'\天融信vsPA功能对比.xlsx'
TODAY = '2026-08-15'

wb = load_workbook(F)
ws = wb['资料登记表']

# 按编号修订
updates = {
    1: {10: 'docs/topsec/天融信防火墙系统一本通.chm（原件备份）+ topsec_manual/（431页解包）'},
    3: {10: 'docs/topsec/topsec_ngfw_product_doc.pdf（26.7MB/2页）'},
    6: {10: '12.1 新特性页快照已取证；其余按功能点逐页在线取证（URL 留档本表）'},
    7: {4: 'L2 手册', 5: '11.1 & later（官方中文）', 6: '2026-08-15 归档',
        7: 'https://docs.paloaltonetworks.com（官方中文 PDF，用户渠道获取）',
        10: 'docs/pa/pan-os-admin-11-1-zh-cn.pdf（59.7MB/1390页）'},
    8: {10: 'URL 取证（快照经目录整理移除）'},
    10: {10: 'URL 取证（快照经目录整理移除）'},
    11: {10: 'URL 取证（PDF 快照经目录整理移除）'},
    13: {10: 'URL 取证（快照经目录整理移除）'},
    15: {10: 'URL 取证（PDF 快照经目录整理移除）'},
}
for r in range(4, 32):
    no = ws.cell(r, 1).value
    if no in updates:
        for col, v in updates[no].items():
            ws.cell(r, col, v)

# 追加 #16 #17
new_docs = [
    [16, 'PAN-OS Web 界面帮助（简体中文版）', 'Palo Alto', 'L2 手册（按 Web UI 菜单组织）', '11.1', '2026-08-15 归档',
     'https://docs.paloaltonetworks.com（官方中文 PDF，用户渠道获取）', '全大类（功能项全集索引）', TODAY,
     'docs/pa/pan-os-web-interface-help-11-1-zh-cn.pdf（10.4MB/1104页）'],
    [17, '天融信防火墙系统一本通（CHM 原件备份）', '天融信', 'L2 手册原件', '文档编号 V3.2406.37098', '©2025',
     '本地既有文件归档', '全大类', TODAY, 'docs/topsec/天融信防火墙系统一本通.chm（21.8MB）'],
]
start = 4 + 15
for i, d in enumerate(new_docs):
    r = start + i
    for c, v in enumerate(d, 1):
        ws.cell(r, c, v)

# 范围基线：PAN-OS 软件版本行补充中文手册说明
ws5 = wb['范围基线']
for r in range(4, 20):
    if ws5.cell(r, 1).value == 'Palo Alto · 软件版本':
        ws5.cell(r, 2, 'PAN-OS 12.1（2025-08 发布的最新大版本；最新维护版 12.1.5，2026-03）；中文官方手册《管理员指南》《Web 界面帮助》11.1 & later 为对照证据（12.1 新增特性以 12.1 英文文档为准）')
        break

wb.save(F)
print('registration updated to 17 docs')
