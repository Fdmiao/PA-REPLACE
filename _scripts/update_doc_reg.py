from openpyxl import load_workbook

BASE = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求'
F = BASE + r'\天融信vsPA功能对比.xlsx'

wb = load_workbook(F)
ws = wb['资料登记表']

# 行 6 是编号3（已下载完成，更新归档状态与类型描述）
for r in range(4, 30):
    if ws.cell(r, 1).value == 3:
        ws.cell(r, 4, 'L1 产品彩页（规格）')
        ws.cell(r, 5, 'NGTOS 平台（2页彩页）')
        ws.cell(r, 10, '_sources/docs/topsec_ngfw_product_doc.pdf（26.7MB）')
        break

# 追加 PA-3400 datasheet（编号15）
docs = [
    [15, 'PA-3400 Series Datasheet（PA-3410/3420/3430/3440）', 'Palo Alto', 'L1 Datasheet', 'PA-3400 系列', '2026-02-26',
     'https://www.paloaltonetworks.com/resources/datasheets/pa-3400-series', 'G/A/C', '2026-08-15', 'PDF 动态下载链接，按需引用页面'],
]
start = 4 + 14
for i, d in enumerate(docs):
    r = start + i
    for c, v in enumerate(d, 1):
        ws.cell(r, c, v)

wb.save(F)
print('updated')
