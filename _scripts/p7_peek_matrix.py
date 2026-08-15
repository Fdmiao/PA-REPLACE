# -*- coding: utf-8 -*-
import os
from openpyxl import load_workbook

BASE = r"e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求"
wb = load_workbook(os.path.join(BASE, "天融信vsPA功能对比.xlsx"), data_only=True)
mx = wb['功能对比矩阵']
print("max_row =", mx.max_row)
for rr in range(1, 5):
    print(f"--- row {rr} ---")
    for c in range(1, 16):
        v = mx.cell(rr, c).value
        if v is not None:
            print(f"  col{c}: {str(v)[:50]}")
