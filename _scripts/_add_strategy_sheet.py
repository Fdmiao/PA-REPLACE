# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import math

base="/Users/个人资料/trx/2026/PA替代"
path=base+"/PA替代需求列表.xlsx"

# 数据模型：每行为 [一级,二级,三级,需求规格,优先级,交付/说明,状态,负责人]
ROWS=[
 # —— 策略基础管理 ——
 ["策略","策略基础管理","支持策略类型划分（Security/NAT/QoS/PBF/Decryption/NPBroker/Tunnel/AppOverride/Auth/DoS/SD-WAN）","PA Web界面『策略』章节定义多种策略类型，按用途各自独立配置；对应功能矩阵#2-9/33/52/122/128/129","高","","已支持","产品线"],
 ["","","支持策略规则移动与克隆（Move/Clone）","PA支持将策略规则移动/克隆到虚拟系统或Panorama设备组（含共享位置）；天融信访问控制策略支持移动策略与克隆模板[toc450050600]","中","","已支持",""],
 ["","","支持审核注释存档（Audit Comment Archive）","PA记录规则的审核注释历史/配置日志/规则更改历史，可CSV导出，用于策略变更审计；天融信仅规则描述字段，无审核注释存档","中","","不支持",""],
 ["","","支持规则使用点击数查询（Rule Usage Hit Count）","PA支持按时间段查询规则命中数，标识未使用规则以清理；含Reset Rule Hit Count重置功能；天融信策略统计开关可显示规则匹配情况[toc450050600]","中","","已支持",""],
 ["","","支持Panorama前置/后置规则层级拆解（Pre-Rules/Post-Rules）","PA在Panorama管理下策略分为前置规则（设备组共享）/本地规则/后置规则三层按序匹配；天融信为自上而下单层规则链[toc450050600]；迁移评估项：需拆解全局继承策略并映射到天融信策略表顺序","中","","待评估",""],
 # —— 安全策略 Security ——
 ["策略","安全策略(Security)","支持基于应用、源/目的区域和地址、用户或用户组、服务（端口/协议）的安全策略","PA安全策略规则基于App-ID应用、用户(User-ID)、区域地址、服务多维匹配并允许/拒绝/跟踪流量；对应矩阵#2/3/4","高","","已支持","产品线"],
 ["","","支持安全策略规则构建块：源用户/HIP/设备/应用/服务/日志等","PA规则构建块含预登录用户、HIP配置文件、设备对象、应用默认端口规避、会话结束日志等；对应矩阵#3/4","高","","已支持",""],
 ["","","支持安全策略优化器（Policy Optimizer）","PA新应用查看器+端口规则迁移至App-ID规则工作流+规则使用统计，减少攻击面；对应矩阵#132","中","","已支持",""],
 ["","","支持配置锁（Config Lock/Commit Lock）","PA可锁定待选配置或提交，防止他管理员并发更改；对应矩阵#135（天融信不支持，PA优势）","中","","待评估",""],
 ["","","支持应用默认端口绑定（Service: application-default）","PA安全策略默认服务为application-default（应用走标准端口，非标准端口不匹配即拦截）；天融信策略为五元组+应用识别组合[toc450050600]；迁移评估项：策略转换若仅匹配应用不限端口将放宽安全边界，需制定非标准端口转换细则；对应矩阵#2","高","","待评估",""],
 ["","","支持安全策略绑定安全配置文件组（Profile Group）","PA安全策略Allow动作关联Profile组（反病毒/漏洞防护/URL过滤/文件阻止/WildFire等）；天融信一体化策略引擎自带检测项；迁移评估项：需评估Profile组到天融信各检测引擎的映射与缺省动作差异；对应矩阵#4","中","","待评估",""],
 # —— NAT ——
 ["策略","NAT策略","支持源NAT（动态/PAT静态IP）与Hairpin回流","PA源NAT含DIPP动态IP+端口/静态IP/接口地址/持久NAT等；对应矩阵#6","高","","已支持","产品线"],
 ["","","支持目的NAT（端口映射/服务器发布）","PA目标NAT含静态/动态IP、端口转换、DNS Rewrite、FQDN目标；对应矩阵#7","高","","已支持",""],
 ["","","支持静态/双向NAT一对一转换","PA支持静态IP转换与双向(Bi-directional)转换；对应矩阵#8","高","","已支持",""],
 ["","","支持NAT与安全策略解耦独立配置","PA NAT策略库独立于安全策略库(Policies>NAT)，先NAT后安全匹配；对应矩阵#9","高","","已支持",""],
 ["","","支持IPv6/NAT64转换（nat64类型）","PA NAT规则支持IPv4/IPv6及NAT64转换类型（RFC6146）；对应矩阵#18","中","","已支持",""],
 ["","","支持NAT后目的区域路由回查匹配（Dest Zone by route lookup）","PA NAT策略目的区域按NAT后地址路由回查确定、源区域按NAT前区域；天融信地址转换为独立策略库常规匹配（安全策略>地址转换）[toc398625673]；迁移评估项：Zone/IP匹配计算差异易引发业务不通，需专项评估NAT策略转换映射；对应矩阵#9","高","","待评估",""],
 # —— QoS ——
 ["策略","QoS策略","支持按应用/源地址/用户等定义QoS策略并划分优先级类","PA QoS策略规则分配QoS类（8类+4级优先级），按Egress Guaranteed/Max控制带宽；对应矩阵#33/66","中","","已支持","产品线"],
 ["","","支持基于DSCP/ToS值匹配的QoS","PA QoS规则可选按DSCP值或IP优先级/ToS匹配；天融信有流量管理模块（虚拟通道+限制/保证带宽+每IP/每用户限速+通道优先级，匹配维度为应用/用户/区域/地址/服务/时间[toc450050609]），DSCP/CoS识别与重标记在现有NGFW手册未见（ACM产品/新版手册待补证）","低","","不支持",""],
 # —— 基于策略转发 PBF ——
 ["策略","基于策略转发(PBF)","支持按源区域/地址/用户/应用/服务+下一跳的基于策略转发","PA PBF策略绕过常规路由表选路，支持路径监控与失效切换、BFD、强制对称返回；对应矩阵#10","高","","已支持","产品线"],
 ["","","支持PBF强制对称返回（非对称路由环境）","PA Enforce Symmetric Return确保返回流量对称","中","","已支持",""],
 # —— 解密策略 Decryption ——
 ["策略","解密策略(Decryption)","支持SSL/TLS解密策略（出站正向/入站反向）","PA解密策略区分SSL Forward Proxy/Inbound Inspection/SSH，可解密以获得流量可见性与深度检测；对应矩阵#29/30","高","","已支持","产品线"],
 ["","","支持解密目标选择与证书配置","PA解密目标允许指定数据包/会话；入站含Certificates证书配置；对应矩阵#31","中","","已支持",""],
 ["","","支持解密流量镜像到专用接口（Decryption Port Mirroring）","PA解密后将明文流量镜像到专用接口供IDS/抓包，需激活免费许可证；天融信无解密流量镜像专用功能（仅SSL卸载+WAF镜像联动）","中","","不支持",""],
 # —— 网络数据包代理 NPBroker ——
 ["策略","网络数据包代理(NPBroker)","支持按应用/用户/区域/设备/IP定义NPBroker规则转发到第三方安全链","PA网络数据包代理策略将已解密/未解密TLS/非TLS流量内联转发至第三方安全工具；对应矩阵#122","高","","部分支持","产品线"],
 ["","","支持NPBroker转发流量类型选择（TLS解密/未解密/非TLS）","PA规则可选Forward TLS(Decrypted)/未解密/非TLS多种流量类型；需免费许可证；天融信无网络数据包代理模块（矩阵#122）","中","","不支持",""],
 # —— 隧道检测 Tunnel Inspection ——
 ["策略","隧道检测(Tunnel Inspection)","支持GRE/GTP-U/HTTP2等明文隧道内层流量检测","PA隧道检测策略解封装并深度检测GRE/GTP-U/HTTP2等明文隧道内容；对应矩阵#128/37","高","","部分支持","产品线"],
 ["","","支持最大封装级数防护（max_encap）","PA对隧道最大封装级数超限丢包防护，防隧道嵌套绕过；天融信手册无封装级数防护（检索无命中）","中","","不支持",""],
 # —— 应用替代 Application Override ——
 ["策略","应用覆盖(App Override)","支持通过Application Override重定义应用识别","PA应用替代策略改变防火墙对网络通信的应用分类方式（如指定明文协议/端口识别）；对应矩阵#26","中","","已支持","产品线"],
 ["","","支持覆盖类型（端口映射类）识别","PA支持端口式应用覆盖识别（configurable自定义应用）","低","","已支持",""],
 # —— 身份验证 Authentication ——
 ["策略","身份验证(Authentication)","支持基于用户身份的身份验证策略（User-ID）","PA身份验证策略在访问网络资源前对用户鉴权，配合User-ID多源映射（AD/XML/Syslog等）；对应矩阵#3/35","高","","已支持","产品线"],
 ["","","支持认证策略超时与单次认证（SSO）","PA支持认证超时减少重复质询、Kerberos/SAML/RADIUS等对接","中","","已支持",""],
 # —— DoS 防护 DoS Protection ——
 ["策略","DoS防护(DoS Protection)","支持针对关键资源的分类/聚合DoS防护","PA DoS策略基于源/目的接口、区域、地址、服务定义允许/拒绝/告警，含分类与聚合双层阈值配置文件；对应矩阵#52/42","高","","已支持","产品线"],
 ["","","支持DoS聚合配置文件（连接/秒阈值）","PA聚合DoS配置文件按传入连接数/秒触发告警与最大速率限制","中","","已支持",""],
 # —— SD-WAN ——
 ["策略","SD-WAN策略","支持按应用配置链路路径管理与链路质量探测","PA SD-WAN策略按应用/抖动/延迟/丢包指标分配链路，流量分发；对应矩阵#129/120","高","","部分支持","产品线"],
 ["","","支持SD-WAN路径选择与链路质量监测（jitter/latency/loss）","PA路径质量配置文件基于抖动/延迟/丢包等运行状况指标选路；天融信有链路探测(ping)+智能选路(最小延迟优先)[toc490041605]，无jitter/latency/loss多维指标","中","","部分支持",""],
]

wb=load_workbook(path)
# 重建"策略需求"sheet，避免残留合并/样式状态
if "策略需求" in wb.sheetnames:
    del wb["策略需求"]
ws=wb.create_sheet("策略需求")

# 写入数据
headers=["一级需求","二级需求","三级需求","需求规格","优先级","交付/规划说明","状态","负责人"]
for c,h in enumerate(headers,1):
    ws.cell(1,c,h)
for r,row in enumerate(ROWS,2):
    for c,val in enumerate(row,1):
        ws.cell(r,c,val)

# 向下填充 A/B 值，便于整组合并
for r in range(2, ws.max_row+1):
    if not ws.cell(r,1).value:
        ws.cell(r,1).value=ws.cell(r-1,1).value
    if not ws.cell(r,2).value:
        ws.cell(r,2).value=ws.cell(r-1,2).value

# 样式
thin=Side(style="thin", color="000000")
border=Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_font=Font(name="等线", size=12, bold=True)
hdr_align=Alignment(horizontal="center", vertical="center", wrap_text=True)
for c in range(1, 9):
    ws.cell(1,c).font=hdr_font
    ws.cell(1,c).alignment=hdr_align

center_cols={1,2,5,7,8}
data_font=Font(name="等线", size=11)
for r in range(2, ws.max_row+1):
    for c in range(1, 9):
        cell=ws.cell(r,c)
        cell.font=data_font
        if c in center_cols:
            cell.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            cell.alignment=Alignment(horizontal="left", vertical="center", wrap_text=True)

# 合并：A/B 整组合并；E/F/G/H 组内合并连续相同值
def merge_runs(col, within_group=False):
    r=2
    while r<=ws.max_row:
        val=ws.cell(r, col).value
        end=r
        while end+1<=ws.max_row and ws.cell(end+1, col).value==val:
            if within_group and ws.cell(end+1, 2).value!=ws.cell(end, 2).value:
                break
            end+=1
        if end>r:
            ws.merge_cells(start_row=r, start_column=col, end_row=end, end_column=col)
        r=end+1

merge_runs(1)
merge_runs(2)
for col in (5,6,7,8):
    merge_runs(col, within_group=True)

# 合并后统一补边框
for r in range(1, ws.max_row+1):
    for c in range(1, 9):
        ws.cell(r,c).border=border

# 行高：按 C/D 列最长文本估算，最小 30
chars_per_line={3:24, 4:27}
for r in range(2, ws.max_row+1):
    max_lines=1
    for c, cpl in chars_per_line.items():
        v=ws.cell(r, c).value
        if v:
            lines=0
            for seg in str(v).split("\n"):
                lines+=max(1, math.ceil(len(seg)/cpl))
            max_lines=max(max_lines, lines)
    ws.row_dimensions[r].height=max(30, max_lines*15+6)
ws.row_dimensions[1].height=20

# 列宽
widths={"A":11.7,"B":19.7,"C":48,"D":55,"E":7.7,"F":20,"G":12,"H":9.7}
for k,v in widths.items():
    ws.column_dimensions[k].width=v
ws.freeze_panes="A2"

wb.save(path)
print("已重建并格式化策略需求sheet，共", len(ROWS), "行")
print("合并范围数:", len(ws.merged_cells.ranges))
for rng in sorted(ws.merged_cells.ranges, key=lambda x:(x.min_col,x.min_row)):
    print(rng)
