# -*- coding: utf-8 -*-
from openpyxl import load_workbook
base="/Users/个人资料/trx/2026/PA替代"
wb=load_workbook(base+"/天融信vsPA功能对比.xlsx", data_only=True)
ws=wb["功能对比矩阵"]
# 策略相关功能点编号
targets=[2,3,4,5,6,7,8,9,10,11,33,52,66,120,122,125,128,129,132,134,135]
for r in range(4, ws.max_row+1):
    no=ws.cell(r,1).value
    if not no or str(no).strip().isalpha(): continue
    try: n=int(float(str(no)))
    except: continue
    if n not in targets: continue
    e=ws.cell(r,5).value; g=ws.cell(r,7).value; i=ws.cell(r,9).value
    fn=ws.cell(r,4).value
    print(f"#{n} {fn} | 天融信={e} | PA={g} | 差异={str(i)[:40] if i else ''}")
wb.close()