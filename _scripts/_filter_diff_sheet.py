# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import math
from collections import Counter

base="/Users/个人资料/trx/2026/PA替代"
path=base+"/PA替代需求列表.xlsx"
KEEP={"不支持","部分支持","待评估"}

wb=load_workbook(path)
src=wb["策略需求"]

def src_val(r,c):
    v=src.cell(r,c).value
    if v is not None:
        return v
    for rng in src.merged_cells.ranges:
        if rng.min_row<=r<=rng.max_row and rng.min_col<=c<=rng.max_col:
            return src.cell(rng.min_row,rng.min_col).value
    return v

rows=[]; dropped=Counter()
for r in range(2, src.max_row+1):
    row=[src_val(r,c) for c in range(1,9)]
    if not any(str(x).strip() for x in row):
        continue
    st=str(row[6]).strip()
    if st in KEEP:
        rows.append(row)
    else:
        dropped[st]+=1

# 重建差异sheet：先解除合并、删多余行
if "策略需求差异" in wb.sheetnames:
    ws=wb["策略需求差异"]
    old_max=ws.max_row
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    new_max=1+len(rows)
    if old_max>new_max:
        ws.delete_rows(new_max+1, old_max-new_max)
else:
    ws=wb.create_sheet("策略需求差异")
    old_max=1

for r in range(1, ws.max_row+1):
    for c in range(1,9):
        ws.cell(r,c).value=None
    if r>1 and r in ws.row_dimensions:
        del ws.row_dimensions[r]

headers=["一级需求","二级需求","三级需求","需求规格","优先级","交付/规划说明","状态","负责人"]
for c,h in enumerate(headers,1):
    ws.cell(1,c,h)
for i,row in enumerate(rows,2):
    for c,v in enumerate(row,1):
        ws.cell(i,c, str(v) if v is not None else "")

thin=Side(style="thin", color="000000")
border=Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_font=Font(name="等线", size=12, bold=True)
hdr_align=Alignment(horizontal="center", vertical="center", wrap_text=True)
for c in range(1,9):
    ws.cell(1,c).font=hdr_font
    ws.cell(1,c).alignment=hdr_align

center_cols={1,2,5,7,8}
data_font=Font(name="等线", size=11)
for r in range(2,ws.max_row+1):
    for c in range(1,9):
        cell=ws.cell(r,c)
        cell.font=data_font
        if c in center_cols:
            cell.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            cell.alignment=Alignment(horizontal="left", vertical="center", wrap_text=True)

def merge_runs(col, within_group=False):
    r=2
    while r<=ws.max_row:
        v=ws.cell(r,col).value
        end=r
        while end+1<=ws.max_row and ws.cell(end+1,col).value==v:
            if within_group and ws.cell(end+1,2).value!=ws.cell(end,2).value:
                break
            end+=1
        if end>r:
            ws.merge_cells(start_row=r, start_column=col, end_row=end, end_column=col)
        r=end+1

merge_runs(1)
merge_runs(2)
for col in (5,6,7,8):
    merge_runs(col, within_group=True)

for r in range(1,ws.max_row+1):
    for c in range(1,9):
        ws.cell(r,c).border=border

chars_per_line={3:24, 4:27}
for r in range(2,ws.max_row+1):
    max_lines=1
    for c,cpl in chars_per_line.items():
        v=ws.cell(r,c).value
        if v:
            lines=0
            for seg in str(v).split("\n"):
                lines+=max(1, math.ceil(len(seg)/cpl))
            max_lines=max(max_lines, lines)
    ws.row_dimensions[r].height=max(30, max_lines*15+6)
ws.row_dimensions[1].height=20

widths={"A":11.7,"B":19.7,"C":48,"D":55,"E":7.7,"F":20,"G":12,"H":9.7}
for k,v in widths.items():
    ws.column_dimensions[k].width=v
ws.freeze_panes="A2"

wb.save(path)
print("策略需求差异：保留", len(rows), "行 / 剔除", sum(dropped.values()), "行", dict(dropped))
print("保留状态分布:", dict(Counter(str(r[6]) for r in rows)))
