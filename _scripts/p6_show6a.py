from openpyxl import load_workbook

BASE = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求'
wb = load_workbook(BASE + r'\天融信vsPA功能对比.xlsx')
mx = wb['功能对比矩阵']
for r in range(4, mx.max_row + 1):
    if str(mx.cell(r, 5).value) == '待验证':
        no = str(mx.cell(r, 1).value)
        note = str(mx.cell(r, 12).value or '')
        ok = ('待验证' in note) or ('V' in note) or ('查' in note)
        print(f'{"OK " if ok else "BAD"} #{no} | {note[:70]}')
