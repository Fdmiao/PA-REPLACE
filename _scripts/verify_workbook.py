from openpyxl import load_workbook

wb = load_workbook(r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求\天融信vsPA功能对比.xlsx')
for name in wb.sheetnames:
    ws = wb[name]
    print(name, '| max_row:', ws.max_row, '| max_col:', ws.max_column)

ws = wb['功能对比矩阵']
cnt_data = sum(1 for r in range(4, ws.max_row + 1) if ws.cell(r, 1).value)
cnt_section = sum(1 for r in range(4, ws.max_row + 1) if ws.cell(r, 2).value and ws.cell(r, 1).value and not ws.cell(r, 3).value)
print('矩阵数据行(含大类):', cnt_data, '| 大类标题行:', cnt_section)
print('矩阵最后一行:', [ws.cell(ws.max_row, c).value for c in range(1, 5)])
