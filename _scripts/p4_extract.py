from openpyxl import load_workbook
from collections import defaultdict, Counter

BASE = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求'
wb = load_workbook(BASE + r'\天融信vsPA功能对比.xlsx')
mx = wb['功能对比矩阵']
wv = wb['待验证移交清单']

CAT = {'A': '基础网络与部署', 'B': '应用识别与控制', 'C': '威胁防护', 'D': '内容安全',
       'E': 'VPN 与加密', 'F': '管理与运维', 'G': '硬件与平台', 'H': '合规与生态'}
rows = []
for r in range(4, mx.max_row + 1):
    no = mx.cell(r, 1).value
    if no is not None and str(no).strip().isdigit():
        rows.append({
            'no': int(str(no)), 'cat': str(mx.cell(r, 2).value or ''), 'mod': str(mx.cell(r, 3).value or ''),
            'fp': str(mx.cell(r, 4).value or ''), 'ts': str(mx.cell(r, 5).value or ''),
            'tsd': str(mx.cell(r, 6).value or ''), 'pa': str(mx.cell(r, 7).value or ''),
            'pad': str(mx.cell(r, 8).value or ''), 'diff': str(mx.cell(r, 9).value or ''),
            'ev': str(mx.cell(r, 10).value or '')})

out = []
out.append(f'==== 大类统计（共{len(rows)}点）====')
stat = defaultdict(lambda: {'n': 0, 'ts': Counter(), 'diff': Counter()})
for x in rows:
    c = x['cat'][:1]
    stat[c]['n'] += 1
    stat[c]['ts'][x['ts']] += 1
    stat[c]['diff'][x['diff']] += 1
for c in 'ABCDEFGH':
    s = stat[c]
    out.append(f"{c} {CAT[c]}: {s['n']}点 | TS:{dict(s['ts'].most_common())} | 差异:{dict(s['diff'].most_common())}")

out.append('\n==== 天融信侧不支持 11 点（缺口）====')
for x in rows:
    if x['ts'] == '不支持':
        out.append(f"#{x['no']} [{x['cat']}/{x['mod']}] {x['fp']} | TS:{x['tsd'][:70]} | PA:{x['pad'][:60]}")

out.append('\n==== PA优势 31 点（含待验证类）====')
for x in rows:
    if 'PA优势' in x['diff']:
        out.append(f"#{x['no']} [{x['mod']}] {x['fp']} | {x['pad'][:75]}")

out.append('\n==== 天融信优势 16+2 点 ====')
for x in rows:
    if '天融信优势' in x['diff']:
        out.append(f"#{x['no']} [{x['mod']}] {x['fp']} | {x['tsd'][:75]}")

out.append('\n==== 高优先级待验证项 ====')
for r in range(4, wv.max_row + 1):
    if wv.cell(r, 2).value and str(wv.cell(r, 6).value or '') == '高':
        out.append(f"{wv.cell(r,1).value} {str(wv.cell(r,2).value)[:12]} {str(wv.cell(r,3).value)[:30]} | {str(wv.cell(r,4).value)[:60]} | {wv.cell(r,5).value}")

with open(BASE + r'\_scripts\p4_report_material.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(out))
print('素材已导出', len(out), '行 → p4_report_material.txt')
