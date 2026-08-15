from openpyxl import load_workbook
from collections import Counter

F = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求\对比工作底稿.xlsx'
wb = load_workbook(F)

# ===== 1. 功能对比矩阵（表头第3行：1序号2大类3功能模块4功能点5TS支持度6TS说明7PA支持度8PA说明9差异10证据11日期12备注） =====
ws = wb['功能对比矩阵']
n, empty, ts_dist, diff_dist = 0, [], Counter(), Counter()
for r in range(4, ws.max_row + 1):
    no = ws.cell(r, 1).value
    if no is None or not str(no).strip().isdigit():
        continue
    n += 1
    holes = []
    if not ws.cell(r, 5).value: holes.append('TS支持度')
    if not str(ws.cell(r, 6).value or '').strip(): holes.append('TS说明')
    if not ws.cell(r, 7).value: holes.append('PA支持度')
    if not str(ws.cell(r, 8).value or '').strip(): holes.append('PA说明')
    if not str(ws.cell(r, 10).value or '').strip(): holes.append('证据')
    if holes: empty.append((no, holes))
    ts_dist[str(ws.cell(r, 5).value)] += 1
    diff_dist[str(ws.cell(r, 9).value or '空')] += 1

print('===== 1. 功能对比矩阵 =====')
print(f'功能点: {n} | 空洞: {len(empty)}')
if empty: print(empty)
print('TS支持度:', dict(ts_dist.most_common()))
print('差异分布:', dict(diff_dist.most_common()))

# ===== 2. 需求映射（R1标题行，R2空，R3起数据：1R编号2需求3映射F4映射说明5结论6取证说明） =====
wm = wb['需求映射']
n_req, miss, cd = 0, [], Counter()
for r in range(3, 3 + 58):
    rid = wm.cell(r, 1).value
    if rid is None: break
    n_req += 1
    cv = wm.cell(r, 5).value
    if not str(cv or '').strip(): miss.append(rid)
    else: cd[str(cv).split('（')[0].strip()] += 1
print('\n===== 2. 需求映射 =====')
print(f'需求: {n_req} | 无结论: {miss if miss else "无"}')
print('结论分布:', dict(cd.most_common()))

# ===== 3. 待验证清单（R1标题R2表头R3起数据） =====
wv = wb['待验证移交清单']
hdr = [str(wv.cell(2, c).value or '') for c in range(1, wv.max_column + 1)]
print('\n===== 3. 待验证清单 =====')
print('表头:', hdr)
items = []
for r in range(4, wv.max_row + 1):
    if wv.cell(r, 2).value or wv.cell(r, 3).value:
        items.append([str(wv.cell(r, c).value or '') for c in range(1, wv.max_column + 1)])
print(f'条数: {len(items)}')
mcol, pcol = 5, 6  # 验证方法/优先级列（按V11行观察: 1编号2来源3功能点4原因5方法6优先级7状态8空）
print('验证方法分布:', dict(Counter(x[mcol-1] for x in items).most_common()))
print('优先级分布:', dict(Counter(x[pcol-1] for x in items).most_common()))
fp = [x for x in items if 'P3填充' in ' '.join(x) or '修正错位' in ' '.join(x)]
print('假阳性残留:', len(fp), fp[:3] if fp else '')

# ===== 4. 质量门判定 =====
print('\n===== P3 质量门 =====')
g1 = n == 139 and not empty
g2 = n_req == 58 and not miss
g3 = len(items) > 0 and not fp
print(f'G1 矩阵139点全填充: {g1}')
print(f'G2 需求58行全结论: {g2}')
print(f'G3 待验证清单干净: {g3}')
print('P3 质量门:', 'PASS' if (g1 and g2 and g3) else 'FAIL')
