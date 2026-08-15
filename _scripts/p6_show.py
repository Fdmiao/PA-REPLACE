from openpyxl import load_workbook

BASE = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求'
wb = load_workbook(BASE + r'\天融信vsPA功能对比.xlsx')
mx = wb['功能对比矩阵']
wv = wb['待验证移交清单']

print('=== #113 / #5 完整行 ===')
for r in range(4, mx.max_row + 1):
    no = str(mx.cell(r, 1).value or '')
    if no in ('5', '113'):
        for c, h in [(3, '模块'), (4, '功能点'), (5, 'TS'), (6, 'TS说明'), (7, 'PA'), (9, '差异'), (10, '证据'), (12, '备注')]:
            print(f'  #{no} {h}: {str(mx.cell(r, c).value or "")[:110]}')

print('\n=== V清单中 #113 / #5 相关条目 ===')
for r in range(4, wv.max_row + 1):
    if wv.cell(r, 2).value and '#113' in str(wv.cell(r, 2).value):
        print(' ', [str(wv.cell(r, c).value or '')[:60] for c in range(1, 8)])

print('\n=== 资料登记表 Datasheet 条目 ===')
rg = wb['资料登记表']
for r in range(1, rg.max_row + 1):
    row = ' '.join(str(rg.cell(r, c).value or '') for c in range(1, rg.max_column + 1))
    if 'Datasheet' in row or 'datasheet' in row or '彩页' in row:
        print(f'  R{r}:', row[:150])
