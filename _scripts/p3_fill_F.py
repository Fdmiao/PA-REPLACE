from openpyxl import load_workbook
from openpyxl.styles import PatternFill

F = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求\对比工作底稿.xlsx'
TODAY = '2026-08-15'
FULL, PART, SUB, NO, TBD = '完全支持', '部分支持', '需订阅授权', '不支持', '待验证'
FILL_BY_SUP = {FULL: PatternFill('solid', fgColor='E2EFDA'), PART: PatternFill('solid', fgColor='FFF2CC'),
               SUB: PatternFill('solid', fgColor='E4DFEC'), NO: PatternFill('solid', fgColor='FBDBDB'),
               TBD: PatternFill('solid', fgColor='EDEDED')}

R = [
(78, FULL, '联动下一代策略集中管理平台（系统设置>集中管理）+安全控制管理平台（SD-WAN：策略/隧道/运维/性能/告警/报表一体）；平台为独立产品，需单独部署授权',
    FULL, 'Panorama集中管理系统（M系列物理/虚拟设备双形态）；Panorama Web界面与防火墙Web界面同体验；上下文切换管理多设备',
    '实现路径不同', '天融信CHM:集中管理[topic207]+安全控制管理平台[topic208]；PA:webhelp p969/p976/p1053，admin p200', 'R54 双方集中管理平台管理规模（受管设备数）需查平台规格'),
(79, PART, 'CHM仅含平台联动入口；策略模板推送、设备分组、管理规模等属集中管理平台侧能力，不在本CHM范围',
    FULL, '设备组（Device Groups，按分支/部门分组）+模板/模板堆栈（Template Stack）；计划配置推送调度器（定时/频率/范围）；预/后处理规则分层',
    'PA优势', '天融信CHM:集中管理[topic207]；PA:webhelp p988/p1032/p1048/p1053', 'R55 天融信集中管理平台模板与规模能力需查平台产品手册'),
(80, FULL, 'Web界面：浏览器HTTPS登录（feth0默认192.168.1.x）；自动保存/系统切换/本地告警/CLI入口等界面功能',
    FULL, 'Web界面：全局查找（跨配置搜索对象/策略/威胁ID）；多语言切换；Panorama与防火墙界面同构',
    '仅命名不同', '天融信CHM:登录系统[ngfw2]+通过浏览器登录[topic142]+熟悉WEB管理界面[toc490041566]；PA:webhelp p21/p26/p42，admin p89', ''),
(81, FULL, 'CLI三通道：Console（本地）/SSH/Telnet（远程）；独立管理路由表仅CLI配置（#network route-indep-mgmt）；Web界面与CLI并行',
    FULL, 'CLI（终端模拟器PuTTY/SSH密钥认证）；操作命令（request restart system等）；ZTP零接触配置模式（Panorama插件）',
    '仅命名不同', '天融信CHM:通过Console口登录[console]+通过SSH方式登录[topic37]+独立管理[topic179]；PA:admin p87/p124，webhelp p49', ''),
(82, FULL, '账号管理：预置superman+自定义管理员；权限模板按功能模块三档（读写/只读/无权限）；根系统管理员/虚拟系统管理员分型；登录安全策略（防暴力破解/验证码/失败锁定）',
    FULL, '管理员角色：动态角色（超级用户/设备管理员只读/虚拟系统管理员）+自定义角色按选项卡逐功能三态（启用/只读/禁用）；提交/保存/恢复范围受角色约束',
    'PA优势', '天融信CHM:账号管理[topic53]+权限模板[topic54]+登录安全策略[ref465087238]；PA:webhelp p731/p1025/p1027，admin p110-145', 'PA角色粒度到选项卡级（admin p132-145粒度表），天融信为模块级三档'),
(83, PART, '无公开XML/REST API文档（CHM仅"第三方管理接口"概述一句；态势感知联动经SSH superman拉取配置，非标准API）',
    FULL, 'XML API+REST API双通道；API密钥生命周期管理（0-525600分钟）；角色可分别控制Web UI/XML API/CLI/REST API权限；User-ID XML API用户映射集成',
    'PA优势', '天融信CHM:功能和特点[toc462736546]第三方管理接口+与态势感知系统联动[-ngfw]；PA:admin p125/p854-858，webhelp p628/p731-732', 'R56 天融信开放API能力需向原厂确认（自动化运维场景关键缺口）'),
(84, FULL, '日志配置（本地存储与外发级别分别设置）；日志服务器syslog外发；日志存储管理（容量上限/每日每周备份/FTP备份/按天导出/加密GB2312-UTF8）',
    FULL, '日志设置+日志转发配置文件（挂安全/DoS/隧道策略）；转发目标：Panorama/日志记录服务/syslog/SNMP陷阱/邮件/HTTP(S)；LFC日志转发卡（硬件加速）',
    '仅命名不同', '天融信CHM:日志配置[toc490041759]+日志服务器配置[ref465084227]+日志存储管理[topic212]；PA:webhelp p305/p808/p817/p721，admin p579/p589', ''),
(85, FULL, 'syslog外发（TCP/UDP、syslog/welf双格式、UTF8/GB2312编码）+态势感知系统联动（日志采集+SNMP性能监控）+SNMP Trap',
    FULL, '日志转发目标：syslog服务器/SNMP陷阱/邮件通知/HTTP(S)负载（自动格式转换）；Panorama/Cortex Data Lake云端集中日志',
    '实现路径不同', '天融信CHM:探针[topic94]+与态势感知系统联动[-ngfw]；PA:admin p579/p759，webhelp p808', '天融信welf格式面向日志审计厂商，PA面向SIEM/云数据湖'),
(86, FULL, '报表模块：报表配置+报表模板自定义（模板管理页签）；威胁统计多视图；攻击者/受害者关联分析（攻击链支持）+数据导出',
    FULL, '40+预定义报告+自定义报告（按需/每夜计划运行）；报告组聚合PDF（含标题页）+计划邮件投递；报告存储约200MB配额管理',
    '实现路径不同', '天融信CHM:报表配置[topic21]+模板管理[topic23]+攻击者[topic55]；PA:webhelp p107/p109，admin p553-570', ''),
(87, FULL, 'HA四模式：主备（AS）/连接保护/负载均衡/集群模式；心跳口（本端/对端IP）+虚拟ID+监控角色主/备',
    FULL, '主动/被动（A/P）+主动/主动（A/A，会话所有者机制）；HA1控制链路+HA2数据链路（备份链路可选）；HA1加密；专用HA端口（按型号）',
    '实现路径不同', '天融信CHM:高可用性[toc462834583]+配置[topic15/136/140/141]；PA:admin p377-399/p411，webhelp p706-717', ''),
(88, PART, '心跳口同步工作状态；状态页支持安全策略/整机配置/管理员配置/实时配置同步+运行比较；链路探测（IP探测）+BFD（单跳/多跳）触发主从切换；会话级同步未见表述',
    FULL, 'HA2数据链路同步会话（会话所有者/会话设置加载共享）；链路监视+路径监视（ping组故障判定）触发故障转移；抢占行为（优先级恢复）；内部健康检查（FPGA/CPU）',
    '待验证', '天融信CHM:配置[topic15]+状态[toc462834584]+BFD[topic7]；PA:admin p378-395，webhelp p715/p716', 'R57 天融信HA会话保持（切换不断连）能力需实测'),
(89, FULL, '配置维护：启动配置/当前配置双轨；配置文件备份/替换/删除/导入导出；恢复出厂配置；远程备份FTP/SFTP（手动+自动）；远程导入',
    FULL, '待选/运行配置双轨+保存/提交/恢复更改（按管理员/位置/vsys过滤部分保存）；命名配置快照（保存/加载/导出XML）；加载Panorama配置版本',
    '实现路径不同', '天融信CHM:配置维护[toc462834560]+远程备份[topic216]+远程导入[topic217]；PA:webhelp p32-35/p650-652，admin p106-107', ''),
(90, FULL, '配置文件导入导出（导入可设为启动配置）；用户批量导入（TXT/CSV模板）；证书导入导出（PKCS12/加密密码）；日志导出CSV/TXT',
    FULL, '导出设备状态/设备配置包（FW主密钥加密）；导出命名配置快照（XML）；导入设备状态（迁移/更换设备）；阻止私钥导出选项',
    '实现路径不同', '天融信CHM:配置维护[toc462834560]+管理用户[ref467487317]+CA管理[ca3]；PA:webhelp p653，admin p106-107/p1049', ''),
(91, FULL, 'License许可证文件模式：受限制模块授权（升级中心>License升级）；功能模块化授权；到期后行为未见说明',
    SUB, '订阅制：授权码激活订阅（威胁防护/DNS安全/WildFire/UEBA/SaaS安全等）；到期前30天告警；到期后订阅功能停止执行、基础防火墙继续运行；解密镜像免费许可证机制',
    '实现路径不同', '天融信CHM:选项[topic206]License升级；PA:webhelp p851，admin p76-78', 'R58 天融信License到期后功能降级行为需实测（替换场景商务关键）'),
(92, FULL, '升级中心四件套：固件维护（当前固件/Patch列表/固件升级/版本说明）+引擎管理（引擎模块单独升级/新模块安装免整机升级）+规则库升级（应用识别/IPS/URL/APT/僵木蠕/病毒）+License升级',
    FULL, 'Device>Software（检查更新/下载/上传/安装/删除映像，发行说明）；动态内容更新（防病毒/应用程序/威胁内容版本）；Panorama统一软件更新管理（验证+分批）',
    '仅命名不同', '天融信CHM:固件维护[topic83]+引擎管理[topic84]+规则库升级[toc462834565]；PA:webhelp p846-847/p1091，admin p33', ''),
(93, PART, 'SNMP V1/V2/V3（管理主机查询+Trap主动告警；V3用户认证加密）；SNMP跨三层IP-MAC绑定应用场景；NetFlow无命中（否定证据）',
    FULL, 'SNMP监控+陷阱（V2c/V3，MIB文件支持，接口统计）；NetFlow导出（物理/子接口/回环/隧道接口流量统计，标准+企业模板）',
    'PA优势', '天融信CHM:SNMP[toc462834551]+V3用户[toc462834555]+配置SNMP管理/陷阱主机[snmp_]；NetFlow检索无命中（否定证据）；PA:webhelp p815/p825，admin p743-762', '天融信无NetFlow导出，流量可视化场景需syslog/态势感知替代'),
(94, PART, '虚拟系统VSYS（一台物理设备划分多逻辑设备：虚系统接口/虚系统管理员/资源共享独享）；云安全中心（云检测：恶意域名/IP/文件MD5，连天融信云）；无公有云形态/云管理平台说明',
    FULL, '虚拟系统vsys（独立策略/接口/管理员，数量依平台+许可证）；VM信息源监控（AWS/Azure/GCP虚拟机IP变更自动注册）；虚拟系统服务路由自定义；Panorama插件云管理',
    'PA优势', '天融信CHM:虚拟系统[xunixitong]+虚系统[topic128]+云安全中心[topic52]；PA:webhelp p780-781，admin p1268/p1311-1327', 'R59 天融信虚拟化版本（NGFW-VM）与云管理能力需查产品线资料'),
]

wb = load_workbook(F)
ws = wb['功能对比矩阵']
pos = {}
for r in range(4, ws.max_row + 1):
    v = ws.cell(r, 1).value
    if v is not None:
        pos[str(v)] = r

filled = 0
for no, ts, td, ps, pd, diff, evid, note in R:
    r = pos.get(str(no))
    if not r:
        print('MISS', no)
        continue
    ws.cell(r, 5, ts); ws.cell(r, 7, ps)
    ws.cell(r, 6, td); ws.cell(r, 8, pd)
    ws.cell(r, 9, diff); ws.cell(r, 10, evid)
    ws.cell(r, 11, TODAY); ws.cell(r, 12, note or 'P3填充')
    ws.cell(r, 5).fill = FILL_BY_SUP[ts]
    ws.cell(r, 7).fill = FILL_BY_SUP[ps]
    filled += 1
wb.save(F)
print('F filled:', filled)
