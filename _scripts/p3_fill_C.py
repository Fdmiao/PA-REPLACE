from openpyxl import load_workbook
from openpyxl.styles import PatternFill

F = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求\对比工作底稿.xlsx'
TODAY = '2026-08-15'
FULL, PART, SUB, NO, TBD = '完全支持', '部分支持', '需订阅授权', '不支持', '待验证'
FILL_BY_SUP = {FULL: PatternFill('solid', fgColor='E2EFDA'), PART: PatternFill('solid', fgColor='FFF2CC'),
               SUB: PatternFill('solid', fgColor='E4DFEC'), NO: PatternFill('solid', fgColor='FBDBDB'),
               TBD: PatternFill('solid', fgColor='EDEDED')}

C = [
(40, FULL, 'IPS 自定义规则（topic201）：规则库未收录的新攻击可自行定义规则；前提是自定义规则被规则集引用且规则集被访问控制规则引用；WAF/数据库安全/工控同样支持自定义规则',
    FULL, 'Custom Objects>自定义间谍软件/漏洞签名：正则表达式模式标识 Phone Home 通信或漏洞利用；上下文/方向/CVE/Bugtraq 元数据；自定义防病毒签名与数据模式签名',
    '仅命名不同', '天融信CHM:IPS自定义规则[topic201]+WAF自定义[topic221]；PA:webhelp p247-249', ''),
(41, FULL, '入侵防御规则集含漏洞攻击特征（缓冲区溢出/非法代码执行类），按规则集选择性启用实现漏洞屏蔽；规则动作可逐条调整',
    FULL, 'Vulnerability Protection 漏洞保护配置文件：预定义 default/strict；按签名严重性分级操作；签名含 CVE/Bugtraq 标识符（区域保护 L3/L4 头检查含 CVE 标识）',
    '仅命名不同', '天融信CHM:入侵防御[topic10]+规则集[topic200]；PA:webhelp p267/p270/p608', '虚拟补丁场景需按 CVE 逐条核对双方特征覆盖（实测项）'),
(42, FULL, 'IPS/僵木蠕引擎=模式匹配+异常检测技术；DDoS 策略模板含畸形报文防御（TCP ABNORMAL FLAG FLOOD）、特殊控制报文防御、IP/端口扫描阈值检测',
    FULL, '威胁异常（Threat Exception/免除配置文件）；区域保护：泛滥/侦察/基于数据包攻击/非IP协议攻击四类；DoS 保护（分类+聚合两层阈值）',
    '仅命名不同', '天融信CHM:僵木蠕防御+DDoS模板[ref464832260]；PA:webhelp p588/p287/admin p1338', ''),
(43, FULL, '病毒过滤六协议（HTTP/FTP/POP3/SMTP/IMAP/IM）；基于SSL解密的代理杀毒；文件过滤模块（类型/方向控制）；高级文件类型未见手册说明——待实测',
    FULL, 'Antivirus 协议解码器逐项操作；标准签名+WildFire 签名双动作列；WildFire 订阅扩展 APK/Flash/PDF/Office/Java/JAR/DMG 高级文件类型',
    'PA优势', '天融信CHM:病毒过滤[topic11]；PA:webhelp p257-258/admin p73/p79', '天融信高级文件类型覆盖需实测规则库（R36）'),
(44, FULL, '规则库升级（toc462834565）：病毒防御引擎等六大规则库在线自动/手动升级，更新频繁',
    FULL, 'Dynamic Updates 计划调度（webhelp p848-850）；AV 内容更新；WildFire 签名：订阅用户每5分钟、威胁防护用户24-48小时；可回退之前版本',
    '仅命名不同', '天融信CHM:规则库升级[toc462834565]；PA:webhelp p848-850/admin p68/p79', ''),
(45, PART, '无本地虚拟执行沙箱（手册仅联动第三方沙箱）；静态检测：MD5 特征比对+云安全中心云查杀（域名/IP/文件MD5）',
    SUB, 'WildFire 设备（WF系列）可作本地私有云分析；私有云不支持 APK/MacOS/archive/Linux 文件分析；需 WildFire 订阅+单独硬件',
    'PA优势', '天融信CHM:APT联动[apt]（无本地执行证据）；PA:webhelp p283-284/p685-686', '本地虚拟执行天融信需确认沙箱产品线配合（R37）'),
(46, FULL, 'APT联动配置：按工作模式（深度/智能模式）决定是否送沙箱；沙箱返回 MD5+判定结果，恶意文件 MD5 记入本地特征库；MD5值查询界面；云安全中心云检测',
    FULL, 'WildFire 分析配置文件：按文件类型/应用/传输方向转发公共云或私有云；免费转发 PE 分析；判定恶意/灰色/良性/钓鱼；新恶意软件生成签名实时分发',
    '仅命名不同', '天融信CHM:APT联动[apt]+MD5查询[md5]+云安全中心[topic52]；PA:webhelp p283/admin p67-68', ''),
(47, FULL, '威胁情报模块：内置威胁情报库+云端威胁情报库双源；规则集配置恶意域名/恶意IP动作与方向；白名单（域名+IPv4）',
    SUB, 'AutoFocus 威胁情报门户（需订阅许可）；外部动态列表 EDL（含恶意域/IP源）；威胁防护订阅含内置 EDL；情报摘要视图',
    '实现路径不同', '天融信CHM:威胁情报[topic62]+规则集[topic60]；PA:webhelp p45/p240/admin p72-74', '天融信云端情报库来源与更新频率需实测（R38）'),
(48, FULL, '威胁情报规则集动作：告警/拦截；僵木蠕恶意IP检测；URL过滤/病毒/WAF/漏扫/DLP 等命中“加入黑名单”策略自动生成动态黑名单（含动态五元组/动态域名黑名单）',
    FULL, 'Anti-Spyware C2 签名阻断+DNS Sinkhole（伪造响应识别受感染主机）；漏洞保护规则 Action=Block IP（阻断期限）；EDL 恶意域自动生成 DNS 阻断签名',
    '仅命名不同', '天融信CHM:威胁情报[topic60]+黑白名单[toc490041706]；PA:webhelp p260/p264/p95/admin p552/p1246', ''),
(49, FULL, '僵木蠕防御：信息检查依赖动态僵尸网络信息库检测 DNS 请求恶意域名；恶意IP检测；访问跟踪（DNS/后续流量跟踪）；域名白名单',
    FULL, 'Botnet 报告：行为机制识别受感染主机（置信度1-5）；Anti-Spyware C2 检测+内联云分析（深度学习高级 C2 实时分析）；DNS Security 基于 DNS 的 C2 签名',
    'PA优势', '天融信CHM:僵木蠕防御；PA:webhelp p97-98/p260/p265', ''),
(50, FULL, '行为分析引擎（检测内网主机未知威胁+日志告警）；行为基线（资产行为基线：基线粒度/容忍度/异常阈值，小时/天级）；违规访问行为分析（对象+用户维度）',
    FULL, 'Automated Correlation Engine：关联对象跟踪多日志源可疑行为升级；关联事件（高/中/低置信度：C2活动/botnet活动/横向移动关联）',
    '仅命名不同', '天融信CHM:行为分析[topic18]+行为基线[topic98]+违规[topic96]；PA:webhelp p76-77/admin p521-524', ''),
(51, FULL, '高级威胁防护模块：APT攻击链描述；沙箱联动+DGA检测/恶意加密流量检测/隐蔽信道检测/扫描探测检测/暴力破解检测/弱口令检测七项；检测结果特征码入库',
    SUB, 'WildFire+Advanced Threat Prevention 订阅：深度学习 C2 检测（全网流量）；威胁防护内联云分析（最大延迟控制）；威胁防护许可包含基本 WildFire',
    '实现路径不同', '天融信CHM:高级威胁防护+策略[topic203]；PA:admin p72-74/webhelp p684', 'PA ATP/DNS Security 均为独立订阅项，采购成本需评估'),
(52, FULL, 'DDoS防御模块：策略模板按协议分类（SYN Flood/TCP ABNORMAL FLAG/连接耗尽/畸形报文等）；阈值 pps/bps 双单位；源限流/目的限流/会话检查；DOS动态黑名单',
    FULL, 'Zone Protection 区域保护（泛滥/侦察/包攻击/非IP协议）+DoS Protection 策略（分类/聚合双配置文件：CPS阈值/警报/最大速率）；三层防御粒度',
    '仅命名不同', '天融信CHM:DDoS模板[ref464832260]；PA:webhelp p189/p287/p588/admin p1338-1358', ''),
(53, FULL, 'WAF 模块：预定义核心规则库（多语言通用+CMS专项 phpwind/wordpress/discuz）+自定义规则；爬虫管理（内置爬虫库+自定义）；登录页面防护（basic/get/form/digest 暴力登录拦截）；IP黑白名单；SSL解密代理WAF；与TopWAF联动',
    PART, '无独立 WAF 模块；Web 攻击防护经漏洞保护签名（SQL注入/命令注入）+内联云分析（深度学习实时检测 SQL/命令注入）；XSS/爬虫/登录防护专项未见；钓鱼防护依赖高级URL过滤',
    '天融信优势', '天融信CHM:WAF[anquancelue_waf]+爬虫[topic220]+登录页面[topic222]；PA:webhelp p270（否定证据：中文手册无WAF专章）', 'Web应用深度防护需求场景建议 PA 侧加配 WAF 或实测内联云分析覆盖（R44）'),
(54, FULL, 'DNS安全模块：用户端/服务端分开配置防护策略；NX异常检测（NX攻击防御：限速/黑名单）；DNS QUERY FLOOD 防御；动态防护列表+静态可信地址表',
    SUB, 'DNS Security 订阅（需先购威胁防护许可）：云预测分析实时查询；Advanced DNS Security 订阅扩展；Anti-Spyware 配置文件内 DNS 签名来源（本地内容/云端）',
    '实现路径不同', '天融信CHM:DNS安全[dns2]+防护策略[topic99]+动态防护列表[topic100]；PA:webhelp p263/admin p73', ''),
(55, FULL, '安全配置文件模型：IPS/僵木蠕/威胁情报/病毒/高级威胁防护/WAF/URL/审计/数据过滤/文件过滤/邮件安全/工业安全/防代理/数据库安全 14类挂接访问控制规则；黑白名单先行匹配；单遍处理',
    FULL, '安全配置文件附加到安全策略规则（7类配置文件）；安全配置文件组聚合复用；App-ID+Content-ID 单遍架构；策略允许后扫描模型',
    '仅命名不同', '天融信CHM:安全配置文件[topic17123321]；PA:webhelp p253/p304', ''),
(56, FULL, '协同联动：与 TA-DB/防病毒网关/WAF/蜜罐/DLP/EDR 联动（检测结果自动生成黑名单）；安全态势感知联动（下发封堵策略）；探针（流量+威胁探针上报态势感知平台）；漏扫联动；防病毒网关联动',
    FULL, '自有云生态联动：WildFire 云/AutoFocus 情报/Panorama 集中管理；设备遥测共享威胁情报；EDL+自动化（API）；与 Cortex XDR 生态配合',
    '实现路径不同', '天融信CHM:协同联动+态势感知[topic129]+探针[topic94]+漏扫联动[1-.html]；PA:admin p74/p216', '天融信为第三方设备联动模式，PA 为自有云平台生态模式'),
(125, FULL, '工业安全模块：MODBUS-TCP/OPC-DA/OPC-UA/IEC104/S7/DNP3/IEC61850 等工控协议；工控策略组+地址策略（地址范围）+值策略（控制地址及变量值）细粒度控制；与访问控制策略结合',
    PART, '中文手册无工控安全专章（检索无命中）；IoT Security 订阅可识别部分 SCADA/工控协议（B123行）；无地址/值级细粒度策略',
    '天融信优势', '天融信CHM:工业安全[topic74]+工控策略组[topic238]+地址策略[topic239]+值策略[topic240]；PA:中文手册无命中（否定证据）', '工控场景为天融信差异化优势项'),
(126, FULL, '数据库安全模块：数据库防护（黑/白/灰名单三模式）+数据库基线+风险语句防护（预定义规则库+自定义规则）+白名单+解码器',
    NO, '无数据库访问防护模块（检索命中均为本地用户数据库 Local User Database，非数据库安全防护；否定证据）',
    '天融信优势', '天融信CHM:数据库安全[topic101]+风险语句[topic113-115]；PA:中文手册无命中（否定证据）', ''),
(127, FULL, '弱口令检测（高级威胁防护策略项）：检测 telnet/ftp/smtp/pop3/imap/rlogin 明文协议登录，判断所用口令是否弱口令；告警动作',
    NO, '无弱口令审计功能（检索仅弱密码套件限制与后量子密码，非口令审计；否定证据）',
    '天融信优势', '天融信CHM:弱口令[topic205]+策略[topic203]；PA:中文手册无命中（否定证据）', ''),
(128, PART, '无独立隧道检测策略；B30行已记：GRE/6to4/ISATAP/6in4 隧道终结+防代理（回环检测）；隧道内检测经 IPS',
    FULL, 'Tunnel Inspection 独立策略类型：检测 GRE/GTP-U/HTTP2 等明文隧道内层流量；最大封装级数防护（超限丢包）；专用隧道检测日志；非加密隧道会话独立记录',
    'PA优势', '天融信CHM:防代理[topic76]+GRE隧道（B30行证据）；PA:webhelp p172-174/admin p544/p655', 'GTP-U/HTTP2 隧道内层检测天融信覆盖需实测（R45）'),
]

wb = load_workbook(F)
ws = wb['功能对比矩阵']
pos = {}
for r in range(4, ws.max_row + 1):
    v = ws.cell(r, 1).value
    if v is not None:
        pos[str(v)] = r

filled = 0
for no, ts, td, ps, pd, diff, evid, note in C:
    r = pos.get(str(no))
    if not r:
        print('MISS', no)
        continue
    ws.cell(r, 5, ts); ws.cell(r, 7, ps)
    ws.cell(r, 6, td); ws.cell(r, 8, pd)
    ws.cell(r, 9, diff); ws.cell(r, 10, evid)
    ws.cell(r, 11, TODAY); ws.cell(r, 12, note or 'P3填充')
    ws.cell(r, 5).fill = FILL_BY_SUP[ts]
    ws.cell(r, 7).fill = FILL_BY_SUP[ps]
    filled += 1
wb.save(F)
print('C category filled:', filled)
