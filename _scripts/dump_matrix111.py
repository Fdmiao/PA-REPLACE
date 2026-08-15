from openpyxl import load_workbook

BASE = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求'
wb = load_workbook(BASE + r'\对比工作底稿.xlsx')
ws = wb['功能对比矩阵']

lines = []
for r in range(4, ws.max_row + 1):
    no, cat, mod, fp = [ws.cell(r, c).value for c in (1, 2, 3, 4)]
    if no is None:
        continue
    if fp:
        lines.append(f"{no} | {cat} | {mod} | {fp}")
    else:
        lines.append(f"== {no} {mod} ==")

with open(BASE + r'\_scripts\matrix_111.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))
print('rows:', len(lines))
