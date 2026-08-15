from openpyxl import load_workbook
from openpyxl.styles import PatternFill

F = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求\对比工作底稿.xlsx'
TODAY = '2026-08-15'
FULL, PART, SUB, NO, TBD = '完全支持', '部分支持', '需订阅授权', '不支持', '待验证'
FILL_BY_SUP = {FULL: PatternFill('solid', fgColor='E2EFDA'), PART: PatternFill('solid', fgColor='FFF2CC'),
               SUB: PatternFill('solid', fgColor='E4DFEC'), NO: PatternFill('solid', fgColor='FBDBDB'),
               TBD: PatternFill('solid', fgColor='EDEDED')}

B = [
(24, FULL, '应用类型24大类（父类/子类）；预定义应用不可改；应用识别/IPS/URL/APT/僵木蠕/病毒引擎规则库频繁更新；应用总数官方文档未公布',
    FULL, 'App-ID 每周内容更新；应用程序+过滤器+组三级；ACE 云引擎经 ML 为未知应用生成新 App-ID；具体数量未在手册公布',
    '待验证', '天融信CHM:应用[toc448912175]；PA:admin p889-940 App-ID', '双方应用总数均需实测规则库版本，登记待验证'),
(25, FULL, 'IPS章：模式匹配+异常检测技术；自定义应用特征规则（规则配置）；流重组',
    FULL, 'App-ID 多元识别：应用签名+协议解码器+上下文签名（协议内隧道检测）+行为特征',
    '仅命名不同', '天融信CHM:入侵防御[topic10]+自定义应用[topic195]；PA:admin p882', ''),
(26, FULL, '自定义应用：应用类型/名称/协议+特征规则配置；应用组聚合',
    FULL, '自定义应用程序（签名式）；Application Override 端口式识别；pcap 生成自定义签名',
    '仅命名不同', '天融信CHM:自定义应用[topic195]+应用组[topic196]；PA:webhelp p178/admin p888', ''),
(27, PART, '应用24大类含子类；手册无 SaaS 治理专项（无 sanctioned 应用标记/SaaS 报告）',
    FULL, 'SaaS App-ID 策略建议导入；sanctioned 应用标记；SaaS 使用报告；HTTP 标头管理企业版 SaaS；Enterprise DLP 联动',
    'PA优势', 'PA:admin p955-967 SaaS App-ID', ''),
(28, FULL, '加密流量识别：SSL代理策略（透明代理/服务器SSL解密）；QUIC 识别能力未见手册说明——待验证',
    FULL, '解密策略全类型：SSL转发代理/入站检查/SSH代理；QUIC/HTTP3 App-ID 识别；TLSv1.3',
    '待验证', '天融信CHM:代理策略[sslxiezaidailicelue]；PA:webhelp p161-175/admin p1009', '天融信QUIC不解密场景覆盖待实测（R40）'),
(29, FULL, '代理策略+代理模板（SSL客户端/服务器模板）+CA管理（RSA/ECC/SM2）；基于SSL解密的代理杀毒/WAF',
    FULL, 'SSL Forward Proxy/Inbound Inspection/SSH解密；解密配置文件（协议/密码套件控制）；SSL解密排除列表+本地排除缓存；HSM 签署',
    'PA优势', '天融信CHM:解密[sslxiezaidailimuban]+CA[ca3]；PA:admin p1013-1067', 'PA解密工程化程度更高（排除缓存/最佳实践向导），天融信支持国密SM2为差异化优势'),
(30, FULL, 'GRE隧道/6to4/ISATAP/6in4隧道终结；防代理（报文回环检测识别代理）；隧道内检测经IPS',
    FULL, 'GRE/IPSec 隧道；Tunnel Inspection 独立策略：检测 GRE/GTP-U/HTTP2 等明文隧道内层流量；最大封装层级防护',
    'PA优势', '天融信CHM:GRE隧道+防代理[topic76]；PA:webhelp p172-175 隧道检测', ''),
(31, FULL, '流量控制：区域间虚拟链路→虚拟通道；通道类型（保证/限制带宽）；按应用/用户/地址/服务/时间段；通道优先级',
    FULL, 'QoS：8个QoS类+实时/高/中/低优先级队列；接口出口限速；DSCP标记与基于DSCP分类；无锁QoS',
    '实现路径不同', '天融信CHM:流量管理[toc450050609]；PA:admin p1116-1139', '天融信区域间模型 vs PA接口出向队列模型；R31/R32 RED/尾丢弃在需求表细化'),
(32, FULL, '用户/用户组管理；角色控制（筛选规则动态划分角色）；认证策略（本地/外部）；用户访问策略；用户流量统计',
    FULL, 'User-ID 源用户作为所有策略匹配条件；组映射；XFF 标头用户识别（代理后用户）；pre-logon 用户',
    '仅命名不同', '天融信CHM:角色[topic75]+管理用户；PA:admin p780-789', ''),
(33, FULL, '虚拟通道：保证带宽（最低可用带宽）+最大带宽限制；虚拟链路上/下行总带宽策略',
    FULL, 'QoS 配置文件：每类 Egress Guaranteed（保障速率%）+Egress Max（最大速率）+突发',
    '仅命名不同', '天融信CHM:配置流量策略[toc399440500]；PA:webhelp p609/admin p1125', ''),
(34, FULL, '邮件安全综合检测（发件人/收件人/主题/附件）；病毒过滤六协议（HTTP/FTP/POP3/SMTP/IMAP/IM）；文件过滤+内容过滤',
    FULL, 'File Blocking（文件类型阻断+方向控制）；Data Filtering（数据模式：信用卡/社保号）；AV 协议解码器操作',
    '仅命名不同', '天融信CHM:邮件安全[anquancelue_youjiananquan]+病毒过滤[topic11]；PA:webhelp p285/p241', ''),
(35, FULL, '规则库升级：应用识别/入侵防御/URL过滤/APT防御/僵木蠕防御/病毒防御引擎；在线自动/手动升级',
    FULL, 'Dynamic Updates：Applications and Threats/AV/URL/WildFire 内容；计划调度；Review Policies 评估策略影响；气隙环境 SCP 手动上传',
    '仅命名不同', '天融信CHM:规则库升级[toc462834565]；PA:webhelp p848-850/admin p33', ''),
(36, TBD, '手册不含应用清单（Facebook/YouTube/WhatsApp/TikTok等检索无命中）；应用覆盖需实测规则库——港澳场景核心验证项',
    FULL, 'facebook 容器应用（facebook-base/chat/video/posting）；YouTube/O365 login 域名清单（HTTP标头条目预定义）；App-ID 国际应用生态成熟',
    '待验证', 'PA:admin p929/p967/p907', 'R13-R21 九大国际应用全部登记实测（T7用例）'),
(121, NO, '手册无 GTP/SCTP 移动网络协议防护内容（否定证据：全文检索无相关命中）',
    SUB, 'MNP 移动网络保护：GTPv1-C/GTPv2-C/GTP-U/PFCP 检测+5G SBA HTTP/2；SCTP Protection（电信信令/Diameter）；需启用 GTP 安全+Threat Prevention 订阅',
    'PA优势', 'PA:webhelp p291-298 GTP/SCTP', '电信运营商场景差异，港澳一般企业场景影响有限'),
(122, PART, '无网络数据包代理模块；旁路检测经“交换机镜像+WAF/漏扫联动”实现（联动阻断）',
    FULL, 'Network Packet Broker 策略：按应用/用户/设备转发已解密TLS/未解密TLS/非TLS流量至第三方安全链（内联）；数据包代理配置文件；免费许可证',
    'PA优势', '天融信CHM:与WAF联动[1-wafngfw]；PA:webhelp p167-171/p320', ''),
(123, PART, '资产发现（主动扫描+被动发现）+资产管理（EDR/漏扫/手动导入六来源）+智能防护自动策略；无专门 IoT 协议画像与设备指纹库',
    SUB, 'IoT Security 订阅：Device-ID 设备对象入策略匹配；设备资产清单；策略建议自动推送；ERSPAN 支持；ML 设备画像',
    '实现路径不同', '天融信CHM:资产发现[topic81]+资产管理；PA:webhelp p100-101/admin p974', ''),
(124, PART, 'SSL卸载策略（解密后送安全引擎检测）；“解密流量镜像到外部接口”专用功能未见手册说明——待验证（R58）',
    FULL, '解密镜像接口（Decryption Port Mirroring）：解密后流量镜像到专用接口供 IDS/抓包；需激活免费许可证',
    '待验证', 'PA:webhelp p312 解密镜像接口+admin p1114', 'R58 需求项，天融信实测确认'),
]

wb = load_workbook(F)
ws = wb['功能对比矩阵']
filled = 0
for no, ts, td, ps, pd, diff, evid, note in B:
    for r in range(4, ws.max_row + 1):
        if str(ws.cell(r, 1).value) == str(no):
            ws.cell(r, 5, ts); ws.cell(r, 7, ps)
            ws.cell(r, 6, td); ws.cell(r, 8, pd)
            ws.cell(r, 9, diff); ws.cell(r, 10, evid)
            ws.cell(r, 11, TODAY); ws.cell(r, 12, note or 'P3填充')
            ws.cell(r, 5).fill = FILL_BY_SUP[ts]
            ws.cell(r, 7).fill = FILL_BY_SUP[ps]
            filled += 1
            break
wb.save(F)
print('B category filled:', filled)
