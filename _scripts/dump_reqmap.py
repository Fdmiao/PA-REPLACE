from openpyxl import load_workbook

F = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求\对比工作底稿.xlsx'
wb = load_workbook(F)
wm = wb['需求映射']
lines = []
for r in range(3, 61):
    rid = wm.cell(r, 1).value
    if rid is None:
        continue
    c3 = str(wm.cell(r, 3).value or '')
    c5 = str(wm.cell(r, 5).value or '')
    c6 = str(wm.cell(r, 6).value or '')
    lines.append(f'{rid} | {c3} | {c5} | {c6[:70]}')
with open(r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求\_scripts\req_map_dump.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))
print('rows:', len(lines))
