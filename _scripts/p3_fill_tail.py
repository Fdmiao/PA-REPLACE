from openpyxl import load_workbook
from openpyxl.styles import PatternFill

F = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求\对比工作底稿.xlsx'
TODAY = '2026-08-15'
FULL, PART, SUB, NO, TBD = '完全支持', '部分支持', '需订阅授权', '不支持', '待验证'
FILL_BY_SUP = {FULL: PatternFill('solid', fgColor='E2EFDA'), PART: PatternFill('solid', fgColor='FFF2CC'),
               SUB: PatternFill('solid', fgColor='E4DFEC'), NO: PatternFill('solid', fgColor='FBDBDB'),
               TBD: PatternFill('solid', fgColor='EDEDED')}

R = [
(132, FULL, '策略分析及管理模块（内置）：策略关系分析（冗余/冲突策略筛选）+宽泛策略分析（源/目的/服务范围超阈值识别）+重复对象检测，提供针对性配置优化建议',
    FULL, 'Policy Optimizer策略优化器：New App Viewer（已见应用迁移）+端口规则→App-ID规则库迁移工作流（减少攻击面/安全启用）+规则使用情况统计（含NPBroker规则）',
    '实现路径不同', '天融信CHM:策略分析及管理[topic88]+策略关系分析[topic89]+宽泛策略分析[topic90]；PA:webhelp p136/p142/p171，admin p910-917', '天融信侧重策略治理（冗余/冲突/宽泛），PA侧重端口→应用迁移'),
(133, NO, '无HSM硬件安全模块/主密钥章节（检索无命中，否定证据）；私钥经USBKey或本地文件导入管理',
    FULL, 'HSM集成（SafeNet Network/Thales CipherTrust Manager，支持HSM高可用）+主密钥加密配置中全部密码与私钥（Master Key and Diagnostics）+HSM包装密钥+主密钥自动更新',
    'PA优势', '天融信CHM:无命中（否定证据）；PA:webhelp p659-662/p854，admin p318/p365/p372', '密钥合规场景（金融/关基）需评估HSM缺口'),
(134, FULL, '网络诊断（ping/traceroute工具，Web界面）+网络抓包（抓包条件/最大可抓包数/自动停止）+链路探测（周期ping，HA依赖）',
    FULL, '内置pcap（管理口/网络接口流量捕获；自定义捕获构建块：捕获阶段/字节计数/文件下载）+威胁触发捕获（AV/防间谍/漏洞配置文件单包或扩展捕获）+未知应用捕获',
    '仅命名不同', '天融信CHM:网络抓包[wangluozhuabao]+验证[topic133]+链路探测[ref487627830]；PA:webhelp p79-83，admin p526-533', ''),
(135, NO, '无配置锁/并发编辑控制机制（检索无命中，否定证据；仅登录失败锁定等账户策略）',
    FULL, '配置更改限制锁：Config Lock（阻止他人更改待选配置）+Commit Lock（阻止提交）两种类型；锁定备注+锁定图标与计数；多管理员部分保存/提交范围控制',
    'PA优势', '天融信CHM:登录安全策略[ref465087238]（仅账户级）；PA:webhelp p40，admin p100-101', ''),
(136, NO, '无蜂窝/5G接口章节（检索无命中，否定证据）',
    FULL, '蜂窝接口（主SIM插槽选择+GPS设置；需支持型号）',
    'PA优势', '天融信CHM:无命中（否定证据）；PA:webhelp p428', 'R70 PA蜂窝接口仅限特定型号，需查datasheet确认型号清单'),
(137, NO, '无PoE供电输出（检索无命中，否定证据）',
    FULL, 'PoE接口供电（等级过流/双重签名/切断电源等管理术语体系；需支持型号）',
    'PA优势', '天融信CHM:无命中（否定证据）；PA:webhelp p425', 'PoE仅限特定型号（PA-400系列部分），需查datasheet'),
(138, TBD, '软件侧工业安全能力完备（工控协议MODBUS-TCP/OPC-DA/OPC-UA/IEC104/S7/DNP3+工业安全数据库报表）；工业加固硬件型号（宽温/防尘/DIN导轨）未见手册',
    TBD, '本手册集无加固型号描述；PA-400R加固系列（宽温/DIN）属产品目录范围，需查硬件资料',
    '待验证', '天融信CHM:工业安全[topic74]+地址策略[topic239]（软件能力）；硬件加固双方手册均无', 'R71 双方工业加固型号（宽温/防尘规格）需查产品目录'),
(139, NO, '无专用日志转发卡（检索无命中，否定证据）',
    FULL, 'LFC日志转发卡（高性能日志卡：数据面全量日志转发至Panorama/syslog；LFC子接口+vsys分配）；PA-7000系列双卡体系（LPC本地存储/LFC高速转发）',
    'PA优势', '天融信CHM:无命中（否定证据）；PA:webhelp p385/p721-722，admin p910/p1329-1330', ''),
]

wb = load_workbook(F)
ws = wb['功能对比矩阵']
pos = {}
for r in range(4, ws.max_row + 1):
    v = ws.cell(r, 1).value
    if v is not None:
        pos[str(v)] = r

filled = 0
for no, ts, td, ps, pd, diff, evid, note in R:
    r = pos.get(str(no))
    if not r:
        print('MISS', no)
        continue
    ws.cell(r, 5, ts); ws.cell(r, 7, ps)
    ws.cell(r, 6, td); ws.cell(r, 8, pd)
    ws.cell(r, 9, diff); ws.cell(r, 10, evid)
    ws.cell(r, 11, TODAY); ws.cell(r, 12, note or 'P3填充')
    ws.cell(r, 5).fill = FILL_BY_SUP[ts]
    ws.cell(r, 7).fill = FILL_BY_SUP[ps]
    filled += 1
wb.save(F)
print('tail filled:', filled)
