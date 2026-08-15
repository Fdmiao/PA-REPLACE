from openpyxl import load_workbook

BASE = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求'
wb = load_workbook(BASE + r'\天融信vsPA功能对比.xlsx')

# 1) V清单优先级分布与高优计数
wv = wb['待验证移交清单']
prio = {}
for r in range(4, wv.max_row + 1):
    if wv.cell(r, 2).value:
        p = str(wv.cell(r, 6).value or '')
        prio[p] = prio.get(p, 0) + 1
print('Sheet3 待验证移交清单: 优先级分布 =', prio, '| 数据行 4-95, 列1=V编号 列2=来源 列6=优先级 列7=状态')

# 2) Sheet1 需求对比表: 待交付需求计数与定位方式
rq = wb['需求对比表']
td = []
gap = []
for r in range(4, 62):
    s = str(rq.cell(r, 6).value or '')
    d = str(rq.cell(r, 12).value or '')
    if '待交付' in s:
        td.append(f'R{r-3}')
    if d.startswith('天融信缺口'):
        gap.append(f'R{r-3}')
print('Sheet1 需求对比表: C6含"待交付" =', len(td), '条 | C12"天融信缺口" =', gap)

# 3) 矩阵缺口行
mx = wb['功能对比矩阵']
for r in range(4, mx.max_row + 1):
    if str(mx.cell(r, 1).value) in ('83', '135'):
        print(f'矩阵 #{mx.cell(r,1).value} 行{r}: {mx.cell(r,4).value} | TS={mx.cell(r,5).value}')
