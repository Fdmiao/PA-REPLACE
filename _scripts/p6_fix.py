from openpyxl import load_workbook

BASE = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求'
for F in [BASE + r'\天融信vsPA功能对比.xlsx', BASE + r'\对比工作底稿.xlsx']:
    wb = load_workbook(F)
    mx = wb['功能对比矩阵']
    fixed = []
    for r in range(4, mx.max_row + 1):
        no = str(mx.cell(r, 1).value or '')
        if no == '113':
            mx.cell(r, 12).value = '天融信ECMP待读CHM路由正文确认（V61）'
            fixed.append('#113备注')
        elif no == '5':
            mx.cell(r, 10).value = '天融信CHM:配置访问控制策略[toc450050600]；PA:PA-3400 Series Datasheet 2026-02-26（资料登记表#15，URL留档）'
            fixed.append('#5证据')
        elif no == '60':
            mx.cell(r, 10).value = '双侧手册全文检索无量化数据（P3否定证据留痕 report_D.txt）；建议50站点抽样（V清单）'
            fixed.append('#60证据')
        elif no == '104':
            mx.cell(r, 10).value = '双侧手册均无资质证书信息（P3全文检索无命中）；天融信查公安部目录、港澳准入另行核实（V54）'
            fixed.append('#104证据')
        elif no == '107':
            mx.cell(r, 10).value = '双侧手册均无信创目录信息（P3全文检索无命中，否定证据）；天融信需查官方信创目录（V57）'
            fixed.append('#107证据')
    try:
        wb.save(F)
        print(F.split('\\')[-1], '→', fixed)
    except PermissionError:
        print(F.split('\\')[-1], '→', fixed, '[保存失败: 文件被占用，请关闭 Excel 后重跑]')
