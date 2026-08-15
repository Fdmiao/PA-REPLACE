from openpyxl import load_workbook
from openpyxl.styles import PatternFill

F = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求\对比工作底稿.xlsx'
TODAY = '2026-08-15'
FULL, PART, SUB, NO, TBD = '完全支持', '部分支持', '需订阅授权', '不支持', '待验证'
FILL_BY_SUP = {FULL: PatternFill('solid', fgColor='E2EFDA'), PART: PatternFill('solid', fgColor='FFF2CC'),
               SUB: PatternFill('solid', fgColor='E4DFEC'), NO: PatternFill('solid', fgColor='FBDBDB'),
               TBD: PatternFill('solid', fgColor='EDEDED')}

R = [
# ===== G 硬件与平台规格 =====
(95, TBD, 'CHM不列型号梯度（系统组成仅描述1U/2U形态与3.2294软件平台）；天融信NGFW全系列型号与规格需查产品手册',
    FULL, '手册全篇引用型号体系：PA-220/400/800/3200/3400/5200/5400/7000系列+PA-7500+CN系列（容器防火墙集群）+VM系列；型号间能力差异显式标注（BFD最小间隔/vsys基数/HA3聚合等）',
    '待验证', '天融信CHM:系统组成[topic32]（型号信息缺失）；PA:webhelp p526/p1018，admin p379/p1312', 'R60 天融信NGFW型号梯度与规格需查产品手册/官网'),
(96, FULL, '系统组成章：前/后面板布局（指示灯/串口/心跳口/扩展插槽/电源）；接口高级信息区分电口/光口（光口不支持修改双工，速率10/100/1000M自适应）',
    FULL, '接口体系：物理/子接口/AE聚合/HA/解密镜像/旁接/集群以太网/蜂窝（SIM双卡）多类型；接口速度10/100Mbps可选',
    '实现路径不同', '天融信CHM:系统组成[topic32]+配置接口高级信息[ref487722465]；PA:webhelp p332-341/p428，admin p477', 'R61 双方各型号接口数量与光口规格需查datasheet'),
(97, FULL, '系统组成：扩展插槽用于扩展卡安装（不同NGFW产品支持不同数量的扩展模块）',
    PART, 'PA-7000系列模块化机箱（NPC/LPC插槽，日志卡LFC槽位）；中低端型号为一体化固定接口，扩展性需查硬件参考指南',
    '待验证', '天融信CHM:系统组成[topic32]；PA:admin p329/p477/p1329（插槽表述仅限PA-7000）', 'R62 双方扩展卡类型（10G/40G/100G）与槽位规格需查硬件手册'),
(98, FULL, '双电源监控（系统资源监控：双电源异常告警）；电源线插槽按设备电源插槽数量配1-2根电源线；后面板电源开关+接地螺丝',
    FULL, '双电源冗余（安装指南：具有双电源的型号连接第二电源实现冗余）；电源状态监控（正常/总数A/B计数）；ps-failure电源告警日志',
    '仅命名不同', '天融信CHM:系统资源监控[toc462834567]+系统组成[topic32]；PA:admin p21/p736，webhelp p1043/p1045', '各型号是否标配双电源需查datasheet'),
(99, FULL, '硬件形态1U或2U机架型设备；19英寸标准机柜尺寸设计；四周10cm散热空间要求；上架挂耳配件',
    TBD, '手册不含机箱尺寸/重量规格；PA按系列提供硬件参考指南（1U/2U/5U chassis，手册外部资料）',
    '待验证', '天融信CHM:系统组成[topic32]+硬件设备安装[topic34]；PA:admin p21（引用硬件参考指南）', 'R63 双方各型号机箱尺寸/重量/功耗需查硬件手册'),
(100, FULL, 'Console串口（本地CLI）；feth0管理口（默认192.168.1.x，HTTPS服务）；独立管理功能（管理路由与业务路由分离，管理口仅作管理用途）',
    FULL, '专用MGT带外管理口（数据/管理平面分离）+PA-5200辅助口AUX-1/AUX-2；串行控制台；USB闪存盘自举（init-cfg.txt/bootstrap.xml）',
    '仅命名不同', '天融信CHM:通过Console口登录[console]+独立管理[topic179]；PA:webhelp p671，admin p20/p207-212', ''),
(101, NO, 'CHM无虚拟机镜像/虚拟化版本章节（检索无命中，否定证据）；虚拟系统VSYS为逻辑分区非虚拟化形态',
    FULL, 'VM-Series虚拟防火墙（ESXi/KVM/Hyper-V私有云部署）；VM信息源联动（ESXi/vCenter动态地址组）；VM-Series NSX版本（NSX Manager自动编排）；解密镜像许可证',
    'PA优势', '天融信CHM:无命中（否定证据）；PA:webhelp p755-757/p974，admin p1060/p1268', 'R64 天融信NGFW-VM虚拟化版本需查产品线资料'),
(102, NO, '无公有云镜像/云市场部署能力（检索无命中，否定证据）；云安全中心仅为云检测情报服务（恶意域名/IP/MD5）',
    FULL, '公有云全栈：AWS/AWS GovCloud/Azure/GCP部署（MGT接口IPv6支持差异说明）；Azure Security Center集成；VM信息源监控云平台虚拟机属性（AWS插件/Azure插件/GCE）',
    'PA优势', '天融信CHM:云安全中心[topic52]（仅情报云）；PA:webhelp p758-760/p808/p994，admin p1276-1281', ''),
(103, TBD, '国密算法栈（IPSec国密1.0/1.1+SM2证书+国密SSL模式）具备信创技术基础；CPU/操作系统平台适配（飞腾/鲲鹏/麒麟/统信）未见手册说明',
    NO, '无信创/国产化平台适配（境外产品；检索无命中）',
    '天融信优势', '天融信CHM:静态隧道[topic4]+CA管理[ca3]+管理服务[topic215]（国密证据）；平台适配需查信创目录', 'R65 天融信信创平台适配清单需查官方信创目录'),
# ===== H 合规与生态 =====
(104, TBD, '国内厂商按规应有计算机信息系统安全专用产品销售许可证；CHM未见证书编号——需查公安部认证目录',
    TBD, '境外产品在中国大陆销售历史上有销售许可，港澳市场另有准入体系；手册无资质信息',
    '待验证', '双侧手册均无资质证书信息', 'R66 天融信销售许可证编号需查公安部目录；港澳准入需另行核实'),
(105, TBD, '等保场景能力项有覆盖（上网行为审计/日志留存/国密——见D65行）；等保测评证书/销售许可增项未见手册',
    NO, '无中国等级保护资质（境外产品；检索无命中）；对应国际体系为NIAP/FIPS',
    '天融信优势', '天融信CHM:审计策略[topic77]（能力侧证据）；资质证书需查等保目录', 'R67 天融信等保资质与测评报告需查官方目录'),
(106, TBD, '国密算法栈具备商用密码应用基础（SM2/SM3/SM4——E70行）；商用密码产品认证证书未见手册',
    NO, '无中国商用密码产品认证；对应FIPS 140-2/CC体系（FIPS-CC模式专章）',
    '天融信优势', '天融信CHM:国密算法栈（E70行证据）；PA:admin p1382-1388 FIPS-CC', 'R47(已有) 查天融信商用密码产品认证证书编号'),
(107, TBD, '信创目录入围与国产化适配清单未见手册；国密能力为信创技术基础',
    NO, '无信创目录入围（境外产品）',
    '天融信优势', '双侧手册均无信创目录信息；天融信需查官方信创目录', 'R65(已有) 天融信信创目录状态需核实'),
(108, TBD, '无CC EAL国际认证信息（检索无命中）；境内认证体系为主',
    TBD, 'FIPS-CC模式为FIPS 140-2+CC联合合规模式（手册专章：模式切换/安全功能/自检日志）；CC EAL等级证书状态需查NIAP PCL目录',
    'PA优势', 'PA:admin p1382-1388 FIPS-CC专章；天融信无命中（否定证据）', 'R68 PA CC EAL4+证书有效期需查NIAP目录（港澳项目关注项）'),
(109, NO, '无FIPS能力（检索无命中，否定证据）',
    FULL, 'FIPS-CC模式：启用流程（MRT维护恢复工具切换）；安全功能（TLS1.2强制/RSA≥2048位/串口受限/并发会话≤4）；FIPS自检日志（fips-selftest事件族）',
    'PA优势', '天融信CHM:无命中（否定证据）；PA:admin p1382-1388/p734-735，webhelp p629/p787/p799', ''),
(110, TBD, '手册无行业准入资质信息；运营商场景有GTP等价能力待核实（天融信ALG/工控协议见C类）；金融/关基准入需查资质包',
    PART, '运营商场景能力有据：GTP检测配置文件（APN/MCC/MNC字段）+SCTP保护+移动网络运营商专属报告/角色；行业准入资质手册不载',
    '待验证', '天融信CHM:ALG[toc490041707]；PA:webhelp p295，admin p673/p141-148', 'R69 双方行业准入资质包（金融/运营商/关基）需商务渠道核实'),
(111, PART, '联动生态：EDR/沙箱/态势感知/IDS/DLP/蜜罐/防病毒网关/WAF联动（黑名单自动封堵）；TopPolicy/CA中心/USBKey（K5/GM3000/K7/TF五款型）；无第三方认证目录',
    FULL, '虚拟化/云生态：VMware NSX服务定义/AWS/Azure/GCP插件（Panorama插件框架）；VM信息源（ESXi/vCenter）；XML API/REST开放集成',
    '实现路径不同', '天融信CHM:协同联动[anquanzhongxinxietongliandong]+系统证书[toc462834534]USBKEY；PA:webhelp p755/p974/p1081，admin p1268', '天融信偏安全设备联动生态，PA偏虚拟化/云平台生态'),
]

wb = load_workbook(F)
ws = wb['功能对比矩阵']
pos = {}
for r in range(4, ws.max_row + 1):
    v = ws.cell(r, 1).value
    if v is not None:
        pos[str(v)] = r

filled = 0
for no, ts, td, ps, pd, diff, evid, note in R:
    r = pos.get(str(no))
    if not r:
        print('MISS', no)
        continue
    ws.cell(r, 5, ts); ws.cell(r, 7, ps)
    ws.cell(r, 6, td); ws.cell(r, 8, pd)
    ws.cell(r, 9, diff); ws.cell(r, 10, evid)
    ws.cell(r, 11, TODAY); ws.cell(r, 12, note or 'P3填充')
    ws.cell(r, 5).fill = FILL_BY_SUP[ts]
    ws.cell(r, 7).fill = FILL_BY_SUP[ps]
    filled += 1
wb.save(F)
print('G/H filled:', filled)
