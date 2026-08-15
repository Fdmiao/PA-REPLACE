from openpyxl import load_workbook
from openpyxl.styles import PatternFill

F = r'e:\zt\ggz\研发部\3.7项目\需求分析\PA-替代需求\对比工作底稿.xlsx'
TODAY = '2026-08-15'
FULL, PART, SUB, NO, TBD = '完全支持', '部分支持', '需订阅授权', '不支持', '待验证'
FILL_BY_SUP = {FULL: PatternFill('solid', fgColor='E2EFDA'), PART: PatternFill('solid', fgColor='FFF2CC'),
               SUB: PatternFill('solid', fgColor='E4DFEC'), NO: PatternFill('solid', fgColor='FBDBDB'),
               TBD: PatternFill('solid', fgColor='EDEDED')}

R = [
# ===== D 内容安全 =====
(57, FULL, 'URL分类库：系统内置+自定义URL组；预定义分类页签+全局查询；分类数/库规模未在手册公布——待实测',
    SUB, 'PAN-DB 订阅数据库（URL 过滤已并入高级URL过滤订阅）；预定义类别列表+URL/IP云数据库；PAN-DB 私有云选项',
    '实现路径不同', '天融信CHM:URL分类[toc448912180]+预定义分类[topic191]；PA:webhelp p273-275/admin p73', '双方分类库规模与分类数均需实测规则库版本（R51）'),
(58, FULL, 'URL自定义分类（URL组：多条URL归组，一条URL可属多分类）；URL黑白名单（字符串含域名，不支持正则）；IP黑白名单；WAF IP黑白名单',
    FULL, '自定义URL类别（逐条添加+列表文件导入）；URL过滤配置文件 block/allow 列表；EDL 外部动态列表（服务器托管免手工）',
    'PA优势', '天融信CHM:自定义分类[topic192]+黑白名单[toc490041706]；PA:webhelp p251/p273', 'PA EDL 动态列表运维成本更低'),
(59, FULL, 'SSL代理策略（透明代理/服务器SSL解密）解密后经URL过滤引擎实现HTTPS站点过滤；不解密时无法按内容过滤',
    FULL, '解密策略（转发代理/入站/SSH）+URL过滤配置文件组合；SSL解密排除（预定义+自定义）；高级URL过滤实时分析',
    '仅命名不同', '天融信CHM:代理策略[sslxiezaidailicelue]；PA:webhelp p161/p800-801/admin p1013', '双方均依赖解密实现 HTTPS 内容过滤'),
(60, TBD, '国内厂商本地分类库（天融信定义维护）；中文/港澳站点分类准确性未实测——无量化数据',
    TBD, 'PAN-DB 国际库为主；中国区域站点分类准确性未见本地化说明——无量化数据',
    '待验证', '双侧手册均无量化数据', 'R46 抽样实测国内站点分类准确率（港澳场景关键项，建议50站点抽样）'),
(61, FULL, '文件过滤：过滤策略+过滤规则+应用组三对象；按文件类型/传输方向/动作控制；与数据过滤组合防泄密',
    FULL, 'File Blocking：按会话流方向（入站/出站/两者）阻止所选文件类型；动作 alert/block/continue；预定义 basic/strict 配置文件',
    '仅命名不同', '天融信CHM:文件过滤[ref487726325]+过滤规则[ref448933869]；PA:webhelp p281/admin p1220', ''),
(62, PART, '数据过滤模块（关键字组+数据过滤策略）；与第三方DLP设备联动（事件上报+黑名单）；无预置合规模板',
    FULL, 'Data Filtering 预定义模式（信用卡/社保号；HIPAA/GDPR/金融服务现代化法案合规）；文件属性/第三方DLP标签；Enterprise DLP 订阅扩展',
    'PA优势', '天融信CHM:数据过滤[ref487726281]+关键字组[ref467487077]+DLP联动[dlp]；PA:admin p1214-1218/webhelp p241', '天融信模板化 DLP 能力弱于 PA，深度合规场景需联动外部DLP'),
(63, PART, '关键字组自定义（机密/违规关键词+分组）；URL路径正则（WAF例外项支持正则）；文件指纹未见支持',
    FULL, '数据模式三类型：预定义/正则表达式/文件属性；正则语法专章（经典+增强语法）；文件属性匹配（文档标题/作者）；第三方DLP标签',
    'PA优势', '天融信CHM:关键字组[ref467487077]+WAF规则集[topic219]；PA:webhelp p241-243/admin p1216', ''),
(64, FULL, '邮件安全模块：综合检测策略（发件人/收件人/主题/附件）；静态+动态黑名单（阻断记录自动生成，默认3600秒）；白名单（IPv4/IPv6）',
    PART, '无独立邮件安全模块；邮件链接钓鱼检测经 WildFire（SMTP/POP3链接判定网络钓鱼，含发件人/收件人记录）；附件经AV/FileBlocking；无垃圾邮件过滤',
    '天融信优势', '天融信CHM:邮件安全[anquancelue_youjiananquan]+综合检测[topic225]+黑名单[topic226]；PA:webhelp p689/admin p67', '邮件网关场景 PA 需评估反垃圾能力缺口'),
(65, FULL, '审计策略（网页内容/网页标题/邮件内容/FTP传输审计）+上网行为审计（网站访问/邮件收发/FTP）+日志存储管理（每日/每周备份，CSV/TXT，GB2312/UTF-8）',
    PART, '日志存储配额+过期期限（各日志类型）；Panorama 日志收集器（保留1-2000天）；无网页内容/上网行为专项审计引擎',
    '天融信优势', '天融信CHM:审计策略[topic77]+上网行为[topic27]+日志存储[topic212]；PA:webhelp p636/p1066/admin p550', '上网行为审计与内容留存为天融信差异化项（等保合规）'),
(66, FULL, '流量管理两级模型：虚拟链路（区域间上/下行总带宽）→虚拟通道（按应用/用户/地址/服务/时间段细分；保证带宽+最大带宽限制；通道优先级）',
    FULL, 'QoS 8类+4级优先级（实时/高/中/低）；每类 Egress Guaranteed（保障速率）+Egress Max+突发；接口级（物理/子接口/AE）+DSCP',
    '实现路径不同', '天融信CHM:流量管理[toc450050609]+配置流量策略[toc399440500]；PA:webhelp p609/admin p1120-1126', '区域间模型 vs 接口出向队列模型（同B33行判读）'),
# ===== E VPN =====
(68, FULL, 'IPSecVPN：静态隧道（IKE自动协商）+手工隧道；管理服务按区域开放；隧道流量统计（上/下行速率/流量）；隧道数容量规格未见手册',
    FULL, 'IPSec Tunnels（标准IKE/IPSec）；隧道接口 tunnel.1-9999；LSVPN大规模VPN（GlobalProtect卫星点对多点）；Panorama自动VPN拓扑',
    '仅命名不同', '天融信CHM:IPSecVPN[ipsec]+静态隧道[topic4]+管理服务[topic177]；PA:webhelp p548/p550/admin p1149', '双方隧道数规格均需查产品 datasheet（R52）'),
(69, FULL, '支持标准 IKE/IPSec 协议，可与任何支持IKE标准的第三方设备建隧道；配置案例含 3DES/MD5/DH1 等通用算法组合',
    FULL, '标准 IKE/IPSec；GRE 隧道章明确“如何提供与其他供应商隧道端点的互操作”；IPSec+GRE 封装组合',
    '仅命名不同', '天融信CHM:IPSecVPN[ipsec]+配置IPSec VPN[ipsec-vpn]；PA:webhelp p548/p554', '与PA设备互通是本替代项目核心场景，需优先POC验证'),
(70, FULL, 'IPSec 协议类型三选项：国际/国密1.0/国密1.1（国密固定数字证书认证+主模式）；CA签发 RSA/ECC/SM2 证书；Web管理SSL模式：国际标准/国密标准/自适应',
    NO, '无国密算法套件（检索无命中；FIPS/CC 国际密码体系）',
    '天融信优势', '天融信CHM:静态隧道[topic4]+CA管理[ca3]+管理服务[topic215]；PA:中文手册无命中（否定证据）', '国密为信创/关基场景关键差异项'),
(71, TBD, '国密算法栈具备商用密码认证基础；手册未列资质证书清单——需查官方认证目录',
    NO, '无商用密码产品认证（境外产品）；对应国际认证为 FIPS 140-2/CC（FIPS-CC模式专章）',
    '天融信优势', 'PA:admin p1382-1388 FIPS-CC；天融信资质需查官方证书（H105/H106行联动）', 'R47 查天融信商用密码产品认证证书编号'),
(72, TBD, 'SSLVPN模块（ACL管理+访问控制综合授权）；并发能力受 license/连接配额控制（虚系统连接配额1024-1600000）；SSLVPN并发授权档位未见手册',
    SUB, 'GlobalProtect 订阅授权（授权码激活；按平台最大用户数授权）；IPSec为主+SSL-VPN回退模式',
    '待验证', '天融信CHM:SSLVPN授权+虚系统[topic128]；PA:webhelp p851/p937/admin p76', 'R48 天融信 SSLVPN 并发授权档位需查报价单'),
(73, TBD, 'SSLVPN章存在但客户端形态（浏览器/客户端/免插件）未在检索证据中明确——需读SSLVPN专章确认',
    FULL, 'GlobalProtect 应用程序（Windows/macOS/Linux/移动）+无客户端VPN（Clientless：浏览器HTML5/JS访问Web应用，无需装客户端）',
    '待验证', 'PA:webhelp p953 Clientless/p926/p966；天融信CHM:SSLVPN章', 'R49 读天融信SSLVPN章节确认客户端形态'),
(74, FULL, '单因子（密码/证书）+双因子（密码+证书/密码+UKey/密码+短信）；短信网关配置；UKey USB认证（需驱动）',
    FULL, 'MFA 供应商API集成（Okta/RSA SecurID/PingID等，随内容更新扩充）；RADIUS/SAML/Kerberos；GlobalProtect网关入站MFA提示',
    'PA优势', '天融信CHM:账号管理[topic53]+Local服务器[toc462834527]+短信[topic68]；PA:webhelp p839/p741/admin p236-240', 'PA MFA 云供应商生态更丰富'),
(75, FULL, 'IPSecVPN章明确：对VPN隧道内流量进行访问控制、流量带宽控制、NAT处理、病毒查杀（解密后全引擎检测）',
    FULL, '隧道解密后流量统一过安全策略（全部安全配置文件）；GlobalProtect隧道流量同策略引擎；明文隧道独立检测策略',
    '仅命名不同', '天融信CHM:IPSecVPN[ipsec]；PA:webhelp p548+安全策略体系', ''),
(76, PART, 'GRE隧道专章（封装IP/IPX）；ISATAP/6to4/6in4/PPPoE隧道；PPTP ALG；L2TP 未见专章命中',
    PART, 'GRE Tunnels 专章（含GRE-in-IPSec封装）；隧道接口tunnel.n；L2TP 同样未见专章命中',
    '仅命名不同', '天融信CHM:GRE隧道[toc490041625]+隧道章[toc490041589]；PA:webhelp p554/p420', 'R50 双方 L2TP 支持需实测或查datasheet'),
(77, FULL, 'DPD（对端失效检测：启用/间隔/超时时间，默认启用）；链路备份（接口级）；策略路由智能选路（负载均衡+链路备份）；HA+链路探测+BFD',
    FULL, 'Tunnel Monitor 隧道监控（失败故障转移至备份网关）；静态路由 Path Monitoring（ICMP探测）；IKEv1 失效对等检测',
    '仅命名不同', '天融信CHM:静态隧道[topic4]+链路备份[toc462834587]+高可用[toc462834581]；PA:webhelp p439/admin p1168', ''),
(129, PART, '无原生SD-WAN模块；策略路由智能选路（按ISP/应用/角色选路+负载均衡+链路备份）+链路负载均衡LLB（就近接入/权重调度）；SD-WAN需联动天融信安全控制管理平台（外部平台）',
    FULL, 'PAN-OS原生SD-WAN：SD-WAN策略+路径质量配置文件（抖动/延迟/丢包指标）+流量分布；Panorama SD-WAN插件集中管理+自动VPN（受Panorama管理）',
    'PA优势', '天融信CHM:策略路由[toc490041605]+LLB+SD-WAN平台联动[topic208]；PA:webhelp p195/p324/p422/admin p72', ''),
(130, PART, '无HIP等价物；EDR联动准入（监控范围内资产须在资产列表且在线否则阻断）+资产发现准入控制',
    FULL, 'HIP对象+HIP配置文件（磁盘加密/防病毒/补丁/主机防火墙/DLP状态检查）；HIP匹配日志；GlobalProtect代理收集端点状态',
    'PA优势', '天融信CHM:EDR联动[edr]+资产管理；PA:webhelp p955-964/admin p633', '终端合规准入场景需评估 EDR 联动 vs HIP 覆盖差异'),
(131, NO, '无抗量子能力（检索无命中，否定证据）',
    FULL, 'IKEv2 后量子预共享密钥 PQ PPK（高级选项卡）；后量子密码检测与控制（PQC/混合PQC算法会话可见性+解密日志）',
    'PA优势', '天融信CHM:无命中（否定证据）；PA:webhelp p576/p581/admin p1046-1048', ''),
]

wb = load_workbook(F)
ws = wb['功能对比矩阵']
pos = {}
for r in range(4, ws.max_row + 1):
    v = ws.cell(r, 1).value
    if v is not None:
        pos[str(v)] = r

filled = 0
for no, ts, td, ps, pd, diff, evid, note in R:
    r = pos.get(str(no))
    if not r:
        print('MISS', no)
        continue
    ws.cell(r, 5, ts); ws.cell(r, 7, ps)
    ws.cell(r, 6, td); ws.cell(r, 8, pd)
    ws.cell(r, 9, diff); ws.cell(r, 10, evid)
    ws.cell(r, 11, TODAY); ws.cell(r, 12, note or 'P3填充')
    ws.cell(r, 5).fill = FILL_BY_SUP[ts]
    ws.cell(r, 7).fill = FILL_BY_SUP[ps]
    filled += 1
wb.save(F)
print('D/E filled:', filled)
